"""Excel导出服务

实现凭证数据的Excel导出功能，包括：
- 科目代码前缀过滤
- 数据格式转换
- Excel文件生成
"""

import pandas as pd
import os
import tempfile
from typing import List, Dict, Optional
from datetime import datetime
from app.core.logger import logger
from app.core.database import get_session
from app.core.config import settings
from app.services.oss_service import OSSService
from app.models.voucher import VoucherData


class ExcelExportService:
    """Excel导出服务"""
    
    def __init__(self):
        self.subject_prefixes = settings.SUBJECT_CODES.split(',')
        self.oss_service = OSSService()
        logger.info(f"科目代码前缀过滤规则: {self.subject_prefixes}")
    
    async def _download_template_from_oss(self) -> Optional[str]:
        """从OSS下载最新模板到本地templates/{TEMPLATE_FILENAME}
        - 使用ETag缓存，若ETag未变化则跳过下载
        - 下载失败时回退本地模板
        """
        try:
            # OSS中的模板文件路径
            template_oss_key = settings.TEMPLATE_OSS_KEY
            local_template_path = f"templates/{settings.TEMPLATE_FILENAME}"
            
            # 确保本地templates目录存在
            os.makedirs("templates", exist_ok=True)
            
            # 检查OSS中模板文件是否存在
            if not self.oss_service.file_exists(template_oss_key):
                logger.warning(f"OSS中模板文件不存在: {template_oss_key}，使用本地备用文件")
                if os.path.exists(local_template_path):
                    logger.info(f"使用本地备用模板文件: {local_template_path}")
                    return local_template_path
                else:
                    logger.error(f"本地备用模板文件也不存在: {local_template_path}")
                    return None
            
            # 使用ETag缓存判断是否需要重新下载
            meta = self.oss_service.head_object(template_oss_key)
            etag_path = f"{local_template_path}.etag"
            need_download = True
            if meta and meta.get('etag') and os.path.exists(local_template_path) and os.path.exists(etag_path):
                try:
                    cached = open(etag_path, 'r', encoding='utf-8').read().strip()
                    if cached == meta['etag']:
                        need_download = False
                        logger.info("模板ETag未变化，跳过下载")
                except Exception:
                    pass

            success = True
            if need_download:
                # 从OSS下载模板文件到本地templates目录
                success = self.oss_service.download_file(template_oss_key, local_template_path)
                if success and meta and meta.get('etag'):
                    try:
                        with open(etag_path, 'w', encoding='utf-8') as f:
                            f.write(meta['etag'])
                    except Exception:
                        pass
            if not success:
                logger.warning(f"从OSS下载模板文件失败: {template_oss_key}，使用本地备用文件")
                if os.path.exists(local_template_path):
                    logger.info(f"使用本地备用模板文件: {local_template_path}")
                    return local_template_path
                else:
                    logger.error(f"本地备用模板文件也不存在: {local_template_path}")
                    return None
            
            logger.info(f"从OSS下载最新模板文件成功: {local_template_path}")
            return local_template_path
            
        except Exception as e:
            logger.error(f"下载模板文件异常: {str(e)}")
            # 异常时尝试使用本地备用文件
            local_template_path = f"templates/{settings.TEMPLATE_FILENAME}"
            if os.path.exists(local_template_path):
                logger.info(f"异常时使用本地备用模板文件: {local_template_path}")
                return local_template_path
            else:
                logger.error(f"本地备用模板文件也不存在: {local_template_path}")
                return None
    
    async def export_voucher_data_to_excel(self, company_codes: List[str], period: str, output_dir: str = "output") -> Optional[str]:
        """导出凭证数据到Excel文件（基于模板）"""
        try:
            logger.info(f"开始导出Excel，公司: {company_codes}, 期间: {period}")
            
            # 从数据库获取数据
            excel_data = self._get_filtered_voucher_data(company_codes, period)
            
            if not excel_data:
                logger.warning("没有找到符合条件的数据")
                return None
            
            # 使用配置的Excel目录
            if output_dir is None:
                from app.core.config import settings
                output_dir = settings.EXCEL_DIR
            
            # 创建输出目录
            os.makedirs(output_dir, exist_ok=True)
            
            # 生成Excel文件
            excel_file = os.path.join(output_dir, f"月度财务报表_{period}.xlsx")
            
            # 定义表头
            headers = [
                "科目编码",
                "科目名称", 
                "日期",
                "费用项目编码",
                "费用项目名称",
                "核算账簿",
                "凭证号",
                "摘要",
                "借方本币",
                "贷方本币"
            ]
            
            # 创建DataFrame
            df = pd.DataFrame(excel_data, columns=headers)
            
            # 确保数值列的数据类型正确
            df['借方本币'] = pd.to_numeric(df['借方本币'], errors='coerce').fillna(0)
            df['贷方本币'] = pd.to_numeric(df['贷方本币'], errors='coerce').fillna(0)
            
            # 从OSS下载模板文件
            template_path = await self._download_template_from_oss()
            if not template_path:
                logger.error("无法获取模板文件")
                return None
            
            # 确定目标sheet名称（根据月份）
            month = period.split('-')[1]  # 从 "2025-09" 提取 "09"
            month_int = int(month)
            target_sheet = f"{month_int}月数据源"
            
            logger.info(f"使用模板: {template_path}, 目标sheet: {target_sheet}")
            
            # 复制模板文件
            import shutil
            shutil.copy2(template_path, excel_file)
            
            # 打开复制的文件并写入数据
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=target_sheet, index=False)
                
                # 设置列宽
                worksheet = writer.sheets[target_sheet]
                for i, col in enumerate(headers):
                    max_length = max(
                        df[col].astype(str).map(len).max(),
                        len(col)
                    )
                    worksheet.column_dimensions[chr(65 + i)].width = min(max_length + 2, 50)
            
            logger.info(f"Excel文件已生成: {excel_file}")
            logger.info(f"共导出 {len(excel_data)} 条记录到 {target_sheet} sheet")
            
            return excel_file
            
        except Exception as e:
            logger.error(f"导出Excel失败: {str(e)}")
            raise
    
    def _get_filtered_voucher_data(self, company_codes: List[str], period: str) -> List[List]:
        """获取凭证数据（数据库中已经是过滤后的数据）"""
        db = get_session()
        excel_data = []
        
        try:
            # 查询指定公司和期间的数据（数据库中已经是过滤后的数据）
            vouchers = db.query(VoucherData).filter(
                VoucherData.company_code.in_(company_codes),
                VoucherData.period == period
            ).all()
            
            logger.info(f"从数据库查询到 {len(vouchers)} 条记录（已过滤）")
            
            for voucher in vouchers:
                # 格式化制单日期
                maketime_str = ""
                if voucher.maketime:
                    maketime_str = voucher.maketime.strftime("%Y-%m-%d")
                
                # 构建Excel行数据
                row_data = [
                    voucher.accsubject_code or "",  # 科目编码
                    voucher.accsubject_name or "",  # 科目名称
                    maketime_str,  # 日期
                    voucher.auxiliary_code or "",  # 费用项目编码
                    voucher.auxiliary_name or "",  # 费用项目名称
                    voucher.accbook_name or "",  # 核算账簿
                    voucher.displayname or "",  # 凭证号
                    voucher.description or "",  # 摘要
                    voucher.debit_org or 0,  # 借方本币
                    voucher.credit_org or 0   # 贷方本币
                ]
                
                excel_data.append(row_data)
            
            logger.info(f"准备导出 {len(excel_data)} 条记录")
            
            return excel_data
            
        except Exception as e:
            logger.error(f"获取数据失败: {str(e)}")
            return []
        finally:
            db.close()
    
    def _filter_related_transactions(self, excel_data: List[List]) -> List[List]:
        """过滤关联交易数据 - 摘要中包含'关联交易'四个字的凭证"""
        related_transactions = []
        
        logger.info(f"开始筛选关联交易，总数据量: {len(excel_data)}")
        
        for i, row in enumerate(excel_data):
            # 检查摘要字段是否包含"关联交易"四个字
            summary = row[7] if len(row) > 7 else ""  # 摘要字段
            if "关联交易" in summary:
                related_transactions.append(row)
                logger.info(f"第{i+1}条 - 发现关联交易: {summary}")
        
        logger.info(f"筛选完成，过滤出 {len(related_transactions)} 条关联交易")
        return related_transactions
    
    def _generate_consolidation_offset_data(self, related_transactions: List[List], period: str) -> List[List]:
        """生成合并抵消数据"""
        offset_data = []
        
        # 从期间提取月份
        month = period.split('-')[1]  # 从 "2025-09" 提取 "09"
        month_int = int(month)
        month_name = f"{month_int}月"
        
        for row in related_transactions:
            # 构建抵消数据行
            offset_row = [
                month_name,  # 月份
                row[0],      # 科目编码
                row[1],      # 科目名称
                row[2],      # 日期
                row[3],      # 费用项目编码
                row[4],      # 费用项目名称
                row[5],      # 核算账簿
                row[6],      # 凭证号
                row[7],      # 摘要
                row[8],      # 借方本币
                row[9]       # 贷方本币
            ]
            offset_data.append(offset_row)
        
        logger.info(f"生成 {len(offset_data)} 条合并抵消数据")
        return offset_data
    
    def _sort_consolidation_data_by_month(self, consolidation_data: List[List]) -> List[List]:
        """按月份排序合并抵消数据"""
        if not consolidation_data:
            return consolidation_data
        
        # 按月份排序（第0列是月份）
        def month_sort_key(row):
            month_str = row[0] if row else ""
            # 提取月份数字进行排序
            if month_str:
                try:
                    # 处理"1月"、"2月"等格式
                    month_num = int(month_str.replace("月", ""))
                    return month_num
                except:
                    return 999  # 无法解析的放在最后
            return 999
        
        sorted_data = sorted(consolidation_data, key=month_sort_key)
        logger.info(f"合并抵消数据已按月份排序，共 {len(sorted_data)} 条")
        return sorted_data
    
    def _sort_consolidation_dataframe_by_month(self, df: pd.DataFrame) -> pd.DataFrame:
        """按月份排序合并抵消DataFrame"""
        if df.empty or '月份' not in df.columns:
            return df
        
        def month_sort_key(month_str):
            if pd.isna(month_str):
                return 999
            try:
                # 处理"1月"、"2月"等格式
                month_num = int(str(month_str).replace("月", ""))
                return month_num
            except:
                return 999
        
        # 按月份排序（从小到大）
        df_sorted = df.sort_values('月份', key=lambda x: x.map(month_sort_key))
        
        # 记录排序结果
        if not df_sorted.empty:
            months = df_sorted['月份'].unique()
            logger.info(f"合并抵消DataFrame已按月份排序，共 {len(df_sorted)} 行，月份顺序: {sorted(months)}")
        else:
            logger.info("合并抵消DataFrame为空")
        
        return df_sorted
    
    async def export_with_consolidation(self, company_codes: List[str], period: str, output_dir: str = None) -> Optional[str]:
        """导出数据并处理合并抵消"""
        try:
            logger.info(f"开始导出数据并处理合并抵消，公司: {company_codes}, 期间: {period}")
            
            # 从数据库获取数据
            excel_data = self._get_filtered_voucher_data(company_codes, period)
            
            if not excel_data:
                logger.warning("没有找到符合条件的数据")
                return None
            
            # 使用配置的Excel目录
            if output_dir is None:
                from app.core.config import settings
                output_dir = settings.EXCEL_DIR
            
            # 创建输出目录
            os.makedirs(output_dir, exist_ok=True)
            
            # 生成Excel文件
            excel_file = os.path.join(output_dir, f"月度财务报表_{period}.xlsx")
            
            # 定义表头
            headers = [
                "科目编码",
                "科目名称", 
                "日期",
                "费用项目编码",
                "费用项目名称",
                "核算账簿",
                "凭证号",
                "摘要",
                "借方本币",
                "贷方本币"
            ]
            
            # 创建DataFrame
            df = pd.DataFrame(excel_data, columns=headers)
            
            # 确保数值列的数据类型正确
            df['借方本币'] = pd.to_numeric(df['借方本币'], errors='coerce').fillna(0)
            df['贷方本币'] = pd.to_numeric(df['贷方本币'], errors='coerce').fillna(0)
            
            # 从OSS下载模板文件
            template_path = await self._download_template_from_oss()
            if not template_path:
                logger.error("无法获取模板文件")
                return None
            
            # 确定目标sheet名称（根据月份）
            month = period.split('-')[1]  # 从 "2025-09" 提取 "09"
            month_int = int(month)
            target_sheet = f"{month_int}月数据源"
            
            logger.info(f"使用模板: {template_path}, 目标sheet: {target_sheet}")
            
            # 复制模板文件
            import shutil
            shutil.copy2(template_path, excel_file)
            
            # 过滤关联交易
            related_transactions = self._filter_related_transactions(excel_data)
            
            # 生成合并抵消数据
            consolidation_data = self._generate_consolidation_offset_data(related_transactions, period)
            
            # 先处理合并抵消数据（只处理一次）
            current_month = f"{int(period.split('-')[1])}月"
            logger.info(f"处理合并抵消数据，当前月份: {current_month}")
            
            # 创建当月关联交易数据
            consolidation_df = pd.DataFrame(consolidation_data, columns=[
                "月份", "科目编码", "科目名称", "日期", "费用项目编码",
                "费用项目名称", "核算账簿", "凭证号", "摘要", "借方本币", "贷方本币"
            ])
            
            if not consolidation_df.empty:
                # 确保数值列格式正确
                consolidation_df['借方本币'] = pd.to_numeric(consolidation_df['借方本币'], errors='coerce').fillna(0)
                consolidation_df['贷方本币'] = pd.to_numeric(consolidation_df['贷方本币'], errors='coerce').fillna(0)
                logger.info(f"当月关联交易数据 {len(consolidation_df)} 条")

            # 处理合并抵消数据（只处理一次）
            final_consolidation_df = None
            if not consolidation_df.empty:
                # 从模板文件读取现有合并抵消数据
                try:
                    existing_df = pd.read_excel(template_path, sheet_name="合并抵消")
                    # 删除当月所有数据，保留其他月份
                    other_months_df = existing_df[existing_df['月份'] != current_month]
                    logger.info(f"从模板文件保留其他月份数据 {len(other_months_df)} 条")
                except Exception as e:
                    logger.info(f"模板文件合并抵消sheet不存在或读取失败: {e}，将创建新的")
                    other_months_df = pd.DataFrame(columns=[
                        "月份", "科目编码", "科目名称", "日期", "费用项目编码",
                        "费用项目名称", "核算账簿", "凭证号", "摘要", "借方本币", "贷方本币"
                    ])

                # 合并其他月份数据和当月新数据
                combined_df = pd.concat([other_months_df, consolidation_df], ignore_index=True)
                # 按月份排序（从小到大）
                final_consolidation_df = self._sort_consolidation_dataframe_by_month(combined_df)
                logger.info(f"合并抵消数据已按月份排序，共 {len(final_consolidation_df)} 条")

            # 将数据写入（1）模板文件 与（2）复制出的报表，两份都写
            def _write_into(target_path: str):
                # 当月数据源：覆盖写入
                with pd.ExcelWriter(target_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=target_sheet, index=False)

                # 合并抵消：使用已处理好的数据，完全替换
                if final_consolidation_df is not None and not final_consolidation_df.empty:
                    with pd.ExcelWriter(target_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        final_consolidation_df.to_excel(writer, sheet_name="合并抵消", index=False)

                # 设置样式与列宽
                from openpyxl import load_workbook
                from openpyxl.styles import Font
                from openpyxl.styles.numbers import FORMAT_NUMBER_00
                wb = load_workbook(target_path)
                # 当月sheet样式
                if target_sheet in wb.sheetnames:
                    ws = wb[target_sheet]
                    header_font = Font(name='宋体', size=10, bold=False)
                    data_font = Font(name='宋体', size=10, bold=False)
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=1, column=col).font = header_font
                    for r in range(2, ws.max_row + 1):
                        for c in range(1, len(headers) + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.font = data_font
                            if headers[c-1] in ['借方本币', '贷方本币']:
                                cell.number_format = FORMAT_NUMBER_00
                    for i, col_name in enumerate(headers):
                        ws.column_dimensions[chr(65 + i)].width = min(max(len(col_name), 15), 50)
                # 合并抵消样式
                if "合并抵消" in wb.sheetnames:
                    ws2 = wb["合并抵消"]
                    consolidation_headers = ["月份", "科目编码", "科目名称", "日期", "费用项目编码", 
                                           "费用项目名称", "核算账簿", "凭证号", "摘要", "借方本币", "贷方本币"]
                    header_font = Font(name='宋体', size=10, bold=False)
                    data_font = Font(name='宋体', size=10, bold=False)
                    for col in range(1, len(consolidation_headers) + 1):
                        ws2.cell(row=1, column=col).font = header_font
                    for r in range(2, ws2.max_row + 1):
                        for c in range(1, len(consolidation_headers) + 1):
                            cell = ws2.cell(row=r, column=c)
                            cell.font = data_font
                            if consolidation_headers[c-1] in ['借方本币', '贷方本币']:
                                cell.number_format = FORMAT_NUMBER_00
                    for i, col_name in enumerate(consolidation_headers):
                        ws2.column_dimensions[chr(65 + i)].width = min(max(len(col_name), 15), 50)
                wb.save(target_path)

            # 写入模板与复制出的报表
            _write_into(template_path)  # 更新本地模板文件
            
            # 复制模板文件到报表文件，然后只更新当月数据源
            import shutil
            shutil.copy2(template_path, excel_file)
            
            # 只更新报表文件的当月数据源
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=target_sheet, index=False)
            
            # 将更新后的模板文件重新上传到OSS，并刷新ETag缓存
            try:
                from app.core.config import settings
                template_oss_key = settings.TEMPLATE_OSS_KEY
                upload_url = await self.oss_service.upload_file(template_path, template_oss_key)
                if upload_url:
                    logger.info(f"模板文件已更新到OSS: {upload_url}")
                    # 刷新本地etag缓存
                    meta = self.oss_service.head_object(template_oss_key)
                    etag_path = f"{template_path}.etag"
                    if meta and meta.get('etag'):
                        try:
                            with open(etag_path, 'w', encoding='utf-8') as f:
                                f.write(meta['etag'])
                        except Exception:
                            pass
                else:
                    logger.warning("模板文件更新到OSS失败")
            except Exception as e:
                logger.warning(f"模板文件更新到OSS异常: {str(e)}")

            logger.info(f"Excel文件已生成: {excel_file}")
            logger.info(f"共导出 {len(excel_data)} 条记录到 {target_sheet} sheet，并同步写入模板")
            logger.info(f"其中关联交易 {len(related_transactions)} 条，合并抵消数据 {len(consolidation_data)} 条")
            
            return excel_file
            
        except Exception as e:
            logger.error(f"导出Excel失败: {str(e)}")
            raise
