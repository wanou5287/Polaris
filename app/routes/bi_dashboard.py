from __future__ import annotations

import copy
import hashlib
import hmac
import io
import json
import os
import secrets
import socket
import threading
import time
from base64 import b64decode, urlsafe_b64decode, urlsafe_b64encode
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from statistics import median
from typing import Any, Dict, List, Sequence, Tuple
from urllib.parse import quote
from zoneinfo import ZoneInfo

import requests
import yaml
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from app.core.logger import logger as app_logger
from app.services.dashboard_share_service import share_dashboard_widget
from app.services.forecast_alert_service import (
    ensure_forecast_alert_schema,
    list_ai_forecasts,
    list_forecast_profiles,
    list_inventory_alerts,
    list_manual_forecasts as list_forecast_manual_rows,
    list_promotion_events,
    recalculate_forecasts_and_alerts,
    save_manual_forecast,
    save_promotion_events,
)
from scripts.yonyou_inventory_sync import (
    AppConfig,
    InventorySyncService,
    JobConfig,
    build_logger,
    default_dates_from_job,
    refresh_inventory_cleaning,
    ensure_inventory_processing_schema,
    ensure_sales_processing_schema,
    load_config,
    quote_mysql_url,
    save_return_unpack_attendance,
)


router = APIRouter()
engine = None
schema_ready = False
project_root = Path(__file__).resolve().parents[2]
yonyou_sync_config_path = project_root / "config" / "yonyou_inventory_sync.yaml"
dashboard_users_config_path = project_root / "config" / "bi_dashboard_users.local.yaml"
DASHBOARD_SESSION_COOKIE = "polaris_session"
DASHBOARD_SESSION_MAX_AGE = 60 * 60 * 24 * 14
DASHBOARD_DEFAULT_PATH = "/financial/bi-dashboard"
DASHBOARD_EDITOR_PATH = "/financial/bi-dashboard/editor"
AUDIT_LOG_ENTRY_PATH = "/financial/bi-dashboard/audit-logs"
AUDIT_LOG_API_PATH = "/financial/bi-dashboard/api/audit-logs"
MASTER_DATA_ENTRY_PATH = "/financial/bi-dashboard/master-data"
MASTER_DATA_API_PATH = "/financial/bi-dashboard/api/master-data"
METRIC_DICTIONARY_ENTRY_PATH = "/financial/bi-dashboard/metric-dictionary"
METRIC_DICTIONARY_API_PATH = "/financial/bi-dashboard/api/metric-dictionary"
PROCUREMENT_ARRIVAL_ENTRY_PATH = "/financial/bi-dashboard/procurement-arrivals"
PROCUREMENT_ARRIVAL_API_PATH = "/financial/bi-dashboard/api/procurement-arrivals"
INVENTORY_FLOW_ENTRY_PATH = "/financial/bi-dashboard/inventory-flows"
INVENTORY_FLOW_API_PATH = "/financial/bi-dashboard/api/inventory-flows"
INVENTORY_FLOW_RULE_API_PATH = "/financial/bi-dashboard/api/inventory-flows/rules"
INVENTORY_FLOW_TASK_API_PATH = "/financial/bi-dashboard/api/inventory-flows/tasks"
TASK_CENTER_API_PATH = "/financial/bi-dashboard/api/task-center"
TASK_CENTER_ITEM_API_PATH = "/financial/bi-dashboard/api/task-center/items"
RECONCILIATION_CENTER_API_PATH = "/financial/bi-dashboard/api/reconciliation-center"
RECONCILIATION_CASE_API_PATH = "/financial/bi-dashboard/api/reconciliation-center/cases"
DATA_AGENT_ENTRY_PATH = "/financial/bi-dashboard/data-agent"
DATA_AGENT_STATUS_API_PATH = "/financial/bi-dashboard/api/data-agent/status"
DATA_AGENT_CHAT_API_PATH = "/financial/bi-dashboard/api/data-agent/chat"
DATA_AGENT_REPORTS_API_PATH = "/financial/bi-dashboard/api/data-agent/reports"
DATA_AGENT_REPORT_GENERATE_API_PATH = "/financial/bi-dashboard/api/data-agent/reports/generate"
SYNC_SCHEDULE_KEY = "raw_yonyou_sync_default"
SYNC_SCHEDULE_JOB_ID = "bi_raw_yonyou_sync"
DATA_AGENT_WEEKLY_JOB_ID = "bi_data_agent_weekly_report"
DATA_AGENT_MONTHLY_JOB_ID = "bi_data_agent_monthly_report"
SYNC_SCHEDULER_TIMEZONE = "Asia/Shanghai"
sync_scheduler: BackgroundScheduler | None = None
sync_scheduler_lock = threading.RLock()
sync_run_lock = threading.Lock()
PREFERRED_SALES_VIEW_NAME = "销售/退货看板"
PREFERRED_SALES_VIEW_DESCRIPTION = "基于销售清洗表预置的销售与退货经营看板"
PREFERRED_INVENTORY_VIEW_NAME = "库存清洗看板"
PREFERRED_INVENTORY_VIEW_DESCRIPTION = "基于库存清洗表预置的库存结构与明细看板"
DATA_AGENT_GITHUB_URL = "https://github.com/3600818203/DataAgent"
DATA_AGENT_REPO_DIR = project_root / "vendor" / "DataAgent"
DATA_AGENT_NAME = "???? Agent"
DATA_AGENT_API_PORT = 18080
DATA_AGENT_UI_PORT = 18501
DATA_AGENT_API_URL = f"http://127.0.0.1:{DATA_AGENT_API_PORT}"
DATA_AGENT_UI_URL = f"http://127.0.0.1:{DATA_AGENT_UI_PORT}"

WIDGET_TYPES: Dict[str, str] = {
    "metric": "指标卡",
    "bar": "柱状图",
    "stacked_bar": "堆积柱状图",
    "stacked_hbar": "堆积条形图",
    "line": "折线图",
    "pie": "饼图",
    "table": "表格",
    "ranking": "排行榜",
    "text": "文本",
}
AGGREGATION_ORDER = ("sum", "avg", "max", "min", "median", "count")
AGGREGATION_LABELS: Dict[str, str] = {
    "sum": "求和",
    "avg": "平均值",
    "max": "最大值",
    "min": "最小值",
    "median": "中位数",
    "count": "计数",
}
AGGREGATIONS = {"sum", "avg", "max", "min", "median", "count"}
FILTER_OPERATOR_ORDER = ("eq", "ne", "gt", "gte", "lt", "lte", "like", "in", "between")
FILTER_OPERATOR_LABELS: Dict[str, str] = {
    "eq": "等于",
    "ne": "不等于",
    "gt": "大于",
    "gte": "大于等于",
    "lt": "小于",
    "lte": "小于等于",
    "like": "包含",
    "in": "属于",
    "between": "区间",
}
FILTER_OPERATORS = {"eq", "ne", "gt", "gte", "lt", "lte", "like", "in", "between"}
SORT_DIRECTIONS = {"asc", "desc"}
LAYOUT_SPANS = {1, 2}
LAYOUT_HEIGHTS = {"compact", "normal", "tall"}
GRID_COLUMNS = 24
GRID_MIN_WIDTH = 2
GRID_MAX_WIDTH = 24
GRID_MIN_HEIGHT = 3
GRID_MAX_HEIGHT = 12

DATASETS: Dict[str, Dict[str, Any]] = {
    "inventory": {
        "label": "现存量明细",
        "table": "bi_inventory_snapshot_daily",
        "date_col": "snapshot_date",
        "fields": {
            "snapshot_date": {"label": "日期", "type": "date", "groupable": True, "filterable": True, "sortable": True},
            "stock_org_name": {"label": "库存组织", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "warehouse_name": {"label": "仓库", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "material_code": {"label": "物料编码", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "material_name": {"label": "物料名称", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "sku_code": {"label": "SKU编码", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "sku_name": {"label": "SKU名称", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "batch_no": {"label": "批号", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "unit_name": {"label": "单位", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "current_qty": {"label": "当前库存", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "available_qty": {"label": "可用库存", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "plan_available_qty": {"label": "计划可用库存", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "incoming_notice_qty": {"label": "在途通知数量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
        },
    },
    "inventory_cleaning": {
        "label": "库存清洗明细",
        "table": "bi_inventory_snapshot_daily_cleaning",
        "date_col": "snapshot_date",
        "fields": {
            "snapshot_date": {"label": "日期", "type": "date", "groupable": True, "filterable": True, "sortable": True},
            "warehouse_name_clean": {"label": "仓库", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "material_code": {"label": "物料编码", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "material_name": {"label": "物料名称", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "stock_status_name": {"label": "物料状态", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "qty": {"label": "数量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "source_row_count": {"label": "原始聚合行数", "type": "number", "numeric": True, "filterable": True, "sortable": True},
        },
    },
    "inventory_turnover": {
        "label": "库存周转分析",
        "table": "",
        "date_col": "month_date",
        "virtual": True,
        "fields": {
            "month_date": {"label": "月份", "type": "date", "filterable": True, "sortable": True},
            "month": {"label": "月份", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "monthly_sales_qty": {"label": "当月销量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "avg_inventory_qty": {"label": "月均库存", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "inventory_turnover_days": {"label": "库存周转天数", "type": "number", "numeric": True, "filterable": True, "sortable": True},
        },
    },
    "sales": {
        "label": "销售出库原始明细",
        "table": "bi_material_sales_daily",
        "date_col": "biz_date",
        "fields": {
            "biz_date": {"label": "日期", "type": "date", "groupable": True, "filterable": True, "sortable": True},
            "source_code": {"label": "单据编号", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "source_lineno": {"label": "行号", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "source_bustype_name": {"label": "业务类型", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "source_vouchdate": {"label": "单据时间", "type": "date", "groupable": True, "filterable": True, "sortable": True},
            "stock_org_name": {"label": "库存组织", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "sales_org_name": {"label": "销售组织", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "warehouse_name": {"label": "仓库", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "customer_name": {"label": "客户", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "material_code": {"label": "物料编码", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "material_name": {"label": "物料名称", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "sku_code": {"label": "SKU编码", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "sku_name": {"label": "SKU名称", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "unit_name": {"label": "单位", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "qty": {"label": "原始数量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "price_qty": {"label": "计价数量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "sub_qty": {"label": "辅数量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "ori_sum": {"label": "原币价税合计", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "nat_sum": {"label": "本币价税合计", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "ori_money": {"label": "原币金额", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "nat_money": {"label": "本币金额", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "ori_tax": {"label": "原币税额", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "nat_tax": {"label": "本币税额", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "tax_rate": {"label": "税率", "type": "number", "numeric": True, "filterable": True, "sortable": True},
        },
    },
    "sales_cleaning": {
        "label": "销售清洗明细",
        "table": "bi_material_sales_daily_cleaning",
        "date_col": "biz_date",
        "fields": {
            "biz_date": {"label": "日期", "type": "date", "groupable": True, "filterable": True, "sortable": True},
            "material_code": {"label": "物料编码", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "material_name": {"label": "物料名称", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "sales_out_xiaoshan": {"label": "销售出库（萧山云仓）", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "sales_out_yuhang": {"label": "销售出库（余杭云仓）", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "transit_intercept_xiaoshan": {"label": "在途拦截（萧山云仓）", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "transit_intercept_yuhang": {"label": "在途拦截（余杭云仓）", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "sales_return_warehouse": {"label": "销售退货（销退仓）", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "return_unpack_attendance": {"label": "退货拆包出勤人数", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "return_unpack_efficiency": {"label": "退货拆包人效", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "total_return_qty": {"label": "当日总退货数量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "total_sales_qty": {"label": "当日总销量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
        },
    },
    "refurb_production": {
        "label": "翻新生产明细",
        "table": "bi_refurb_production_daily",
        "date_col": "biz_date",
        "fields": {
            "biz_date": {"label": "日期", "type": "date", "groupable": True, "filterable": True, "sortable": True},
            "refurb_category": {"label": "翻新种类", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "material_name": {"label": "物料名称", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "feeding_qty": {"label": "领料数量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "total_work_hours": {"label": "总耗费工时", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "plan_qty": {"label": "计划数量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "quality_defect_qty": {"label": "品质-检出不合格品", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "production_good_qty": {"label": "生产-产出良品数量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "production_bad_qty": {"label": "生产-产出不良品数量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "final_good_qty": {"label": "最终合格数量", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "non_refurbishable_rate": {"label": "不可翻新率", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "quality_reject_rate": {"label": "出货合格率-品质", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "plan_achievement_rate": {"label": "计划达成率", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "refurb_efficiency": {"label": "翻新人效", "type": "number", "numeric": True, "filterable": True, "sortable": True},
        },
    },
    "sales_forecast": {
        "label": "销售/生产预测明细",
        "table": "bi_sales_forecast_ai_daily",
        "date_col": "forecast_date",
        "fields": {
            "forecast_date": {"label": "预测日期", "type": "date", "groupable": True, "filterable": True, "sortable": True},
            "material_name": {"label": "物料名称", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "demand_type": {"label": "需求类型", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "material_role": {"label": "物料角色", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "base_qty": {"label": "基础需求", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "ai_qty": {"label": "AI预测值", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "manual_qty": {"label": "手动预测值", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "final_qty": {"label": "最终预测值", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "weekday_factor": {"label": "季节性系数", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "trend_factor": {"label": "趋势系数", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "promo_factor": {"label": "促销系数", "type": "number", "numeric": True, "filterable": True, "sortable": True},
        },
    },
    "inventory_alert": {
        "label": "安全库存预警",
        "table": "bi_inventory_alert_log",
        "date_col": "snapshot_date",
        "fields": {
            "snapshot_date": {"label": "快照日期", "type": "date", "groupable": True, "filterable": True, "sortable": True},
            "material_name": {"label": "物料名称", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "demand_type": {"label": "需求类型", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "material_role": {"label": "物料角色", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "current_stock_qty": {"label": "当前良品库存", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "forecast_14d_qty": {"label": "未来14天需求", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "coverage_days": {"label": "库存覆盖天数", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "threshold_days": {"label": "预警阈值天数", "type": "number", "numeric": True, "filterable": True, "sortable": True},
            "alert_level": {"label": "预警等级", "type": "string", "groupable": True, "filterable": True, "sortable": True},
            "pushed_to_dingtalk": {"label": "已推送钉钉", "type": "number", "numeric": True, "filterable": True, "sortable": True},
        },
    },
}


def load_dashboard_auth_config() -> Dict[str, Any]:
    if not dashboard_users_config_path.exists():
        return {}
    try:
        raw = yaml.safe_load(dashboard_users_config_path.read_text(encoding="utf-8")) or {}
    except Exception:
        return {}
    return raw if isinstance(raw, dict) else {}


def load_dashboard_users() -> Dict[str, str]:
    fallback = {
        os.getenv("BI_DASH_USERNAME", "bi_admin"): os.getenv("BI_DASH_PASSWORD", "ChangeMe123!"),
    }
    raw = load_dashboard_auth_config()
    users: Dict[str, str] = {}
    for item in raw.get("users", []):
        if not isinstance(item, dict):
            continue
        username = str(item.get("username") or "").strip()
        password = str(item.get("password") or "")
        if username and password:
            users[username] = password
    return users or fallback


def dashboard_session_secret() -> str:
    raw = load_dashboard_auth_config()
    configured = str(((raw.get("settings") or {}).get("session_secret")) or os.getenv("BI_DASH_SESSION_SECRET") or "").strip()
    if configured:
        return configured
    seed = f"{project_root}|{os.getenv('BI_DASH_USERNAME', 'bi_admin')}|polaris-session"
    return hashlib.sha256(seed.encode("utf-8")).hexdigest()


def authenticate_dashboard_user(username: str, password: str) -> str | None:
    expected = load_dashboard_users().get(username)
    if expected and secrets.compare_digest(password, expected):
        return username
    return None


def create_dashboard_session(username: str, max_age: int = DASHBOARD_SESSION_MAX_AGE) -> str:
    expires_at = int(time.time()) + max_age
    payload = f"{username}|{expires_at}"
    encoded = urlsafe_b64encode(payload.encode("utf-8")).decode("ascii").rstrip("=")
    signature = hmac.new(
        dashboard_session_secret().encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return f"{encoded}.{signature}"


def parse_dashboard_session(token: str | None) -> str | None:
    if not token or "." not in token:
        return None
    encoded, signature = token.rsplit(".", 1)
    try:
        padding = "=" * (-len(encoded) % 4)
        payload = urlsafe_b64decode(f"{encoded}{padding}".encode("ascii")).decode("utf-8")
        username, expires_at_text = payload.split("|", 1)
        expires_at = int(expires_at_text)
    except Exception:
        return None
    expected_signature = hmac.new(
        dashboard_session_secret().encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    if not secrets.compare_digest(signature, expected_signature):
        return None
    if expires_at < int(time.time()):
        return None
    if username not in load_dashboard_users():
        return None
    return username


def parse_basic_auth(request: Request) -> str | None:
    auth_header = request.headers.get("authorization") or ""
    if not auth_header.startswith("Basic "):
        return None
    try:
        decoded = b64decode(auth_header.split(" ", 1)[1]).decode("utf-8")
        username, password = decoded.split(":", 1)
    except Exception:
        return None
    return authenticate_dashboard_user(username, password)


def current_dashboard_user(request: Request) -> str | None:
    session_user = parse_dashboard_session(request.cookies.get(DASHBOARD_SESSION_COOKIE))
    if session_user:
        return session_user
    return parse_basic_auth(request)


def require_auth(request: Request) -> str:
    username = current_dashboard_user(request)
    if username:
        return username
    raise HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="未登录或会话已失效",
        headers={"WWW-Authenticate": "Basic"},
    )


def sanitize_next_path(raw_value: str | None) -> str:
    candidate = str(raw_value or "").strip()
    if not candidate.startswith("/") or candidate.startswith("//"):
        return DASHBOARD_DEFAULT_PATH
    return candidate


def render_template(template_name: str, replacements: Dict[str, str] | None = None) -> str:
    template_path = Path(__file__).resolve().parents[1] / "templates" / template_name
    content = template_path.read_text(encoding="utf-8")
    for key, value in (replacements or {}).items():
        content = content.replace(key, value)
    return content


def dashboard_logo_wordmark_svg() -> str:
    return """
<svg width="248" height="72" viewBox="0 0 248 72" fill="none" xmlns="http://www.w3.org/2000/svg" aria-label="鍖楁瀬鏄?>
  <defs>
    <linearGradient id="polarisWordmarkBlue" x1="7" y1="8" x2="60" y2="62" gradientUnits="userSpaceOnUse">
      <stop stop-color="#1677FF"/>
      <stop offset="1" stop-color="#63B3FF"/>
    </linearGradient>
  </defs>
  <g>
    <rect x="2" y="2" width="68" height="68" rx="20" fill="white" stroke="#E6ECF5" stroke-width="1.5"/>
    <circle cx="36" cy="36" r="17" fill="url(#polarisWordmarkBlue)" opacity="0.12"/>
    <path d="M36 15L40.9 27.1L53.9 28L43.9 36.3L47.1 49.1L36 42.1L24.9 49.1L28.1 36.3L18.1 28L31.1 27.1L36 15Z" fill="url(#polarisWordmarkBlue)"/>
  </g>
  <g transform="translate(88 13)">
    <text x="0" y="12" font-family="SF Pro Display, PingFang SC, Microsoft YaHei, sans-serif" font-size="13" font-weight="500" fill="#6E6E73" letter-spacing="0.48">Polaris</text>
    <text x="0" y="39" font-family="PingFang SC, Microsoft YaHei, sans-serif" font-size="30" font-weight="700" fill="#1D1D1F" letter-spacing="-0.7">鍖楁瀬鏄?/text>
    <text x="0" y="57" font-family="PingFang SC, Microsoft YaHei, sans-serif" font-size="13" font-weight="600" fill="#1677FF">缁忚惀涓績</text>
  </g>
</svg>
""".strip()


def dashboard_logo_badge_svg() -> str:
    return """
<svg width="260" height="260" viewBox="0 0 260 260" fill="none" xmlns="http://www.w3.org/2000/svg" aria-label="鍖楁瀬鏄?>
  <defs>
    <linearGradient id="polarisBadgeBlue" x1="54" y1="46" x2="210" y2="214" gradientUnits="userSpaceOnUse">
      <stop stop-color="#1677FF"/>
      <stop offset="1" stop-color="#63B3FF"/>
    </linearGradient>
  </defs>
  <rect x="10" y="10" width="240" height="240" rx="56" fill="white" stroke="#E6ECF5" stroke-width="1.5"/>
  <circle cx="130" cy="98" r="42" fill="url(#polarisBadgeBlue)" opacity="0.12"/>
  <path d="M130 48L141.4 76.1L171.6 78.1L148.4 97.4L155.8 127.2L130 111L104.2 127.2L111.6 97.4L88.4 78.1L118.6 76.1L130 48Z" fill="url(#polarisBadgeBlue)"/>
  <text x="130" y="172" text-anchor="middle" font-family="PingFang SC, Microsoft YaHei, sans-serif" font-size="52" font-weight="700" fill="#111827" letter-spacing="-1.2">鍖楁瀬鏄?/text>
  <text x="130" y="204" text-anchor="middle" font-family="SF Pro Display, PingFang SC, Microsoft YaHei, sans-serif" font-size="20" font-weight="500" fill="#6E6E73" letter-spacing="0.5">Polaris</text>
</svg>
""".strip()


def dashboard_logo_badge_small_svg() -> str:
    return """
<svg width="48" height="48" viewBox="0 0 48 48" fill="none" xmlns="http://www.w3.org/2000/svg" aria-label="鍖楁瀬鏄?>
  <defs>
    <linearGradient id="polarisBadgeBlueSmall" x1="8" y1="6" x2="41" y2="42" gradientUnits="userSpaceOnUse">
      <stop stop-color="#1677FF"/>
      <stop offset="1" stop-color="#63B3FF"/>
    </linearGradient>
  </defs>
  <rect x="1.25" y="1.25" width="45.5" height="45.5" rx="13" fill="white" stroke="#E6ECF5"/>
  <circle cx="24" cy="24" r="9.2" fill="url(#polarisBadgeBlueSmall)" opacity="0.14"/>
  <path d="M24 8.5L27 16.3L35.3 16.9L28.9 22.2L31 30.4L24 25.9L17 30.4L19.1 22.2L12.7 16.9L21 16.3L24 8.5Z" fill="url(#polarisBadgeBlueSmall)"/>
</svg>
""".strip()


def dashboard_shell_head() -> str:
    return """
<style>
  .polaris-app-shell {
    min-height: 100vh;
    display: grid;
    grid-template-columns: 84px minmax(0, 1fr);
    gap: 20px;
    padding: 20px;
  }
  .polaris-app-main {
    min-width: 0;
    position: relative;
    z-index: 1;
  }
  .polaris-side-panel {
    position: sticky;
    top: 16px;
    height: calc(100vh - 32px);
    padding: 14px 10px 12px;
    border-radius: 28px;
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 14px;
    background: rgba(255, 255, 255, 0.9);
    border: 1px solid rgba(15, 23, 42, 0.08);
    box-shadow: 0 10px 28px rgba(15, 23, 42, 0.05);
    backdrop-filter: blur(18px) saturate(150%);
    -webkit-backdrop-filter: blur(18px) saturate(150%);
  }
  .polaris-side-brand {
    width: 54px;
    height: 54px;
    display: grid;
    place-items: center;
    flex: none;
  }
  .polaris-side-brand svg {
    width: 50px;
    height: 50px;
    display: block;
  }
  .polaris-side-nav {
    width: 100%;
    display: flex;
    flex-direction: column;
    gap: 10px;
    flex: 1;
  }
  .polaris-side-item,
  .polaris-side-subitem,
  .polaris-side-logout {
    border: 0;
    cursor: pointer;
    font: inherit;
    transition: background .2s ease, color .2s ease, box-shadow .2s ease, transform .2s ease, border-color .2s ease;
  }
  .polaris-side-item {
    width: 100%;
    padding: 12px 6px;
    border-radius: 20px;
    background: transparent;
    border: 1px solid transparent;
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 6px;
    color: #6b7280;
  }
  .polaris-side-item:hover,
  .polaris-side-item.is-active,
  .polaris-side-group.is-open > .polaris-side-item,
  .polaris-side-group.is-active > .polaris-side-item {
    background: rgba(15, 23, 42, 0.04);
    border-color: rgba(15, 23, 42, 0.06);
    color: #111827;
    box-shadow: 0 8px 20px rgba(15, 23, 42, 0.04);
  }
  .polaris-side-icon {
    width: 34px;
    height: 34px;
    border-radius: 12px;
    display: grid;
    place-items: center;
    background: rgba(250, 250, 250, 0.96);
    border: 1px solid rgba(15, 23, 42, 0.08);
  }
  .polaris-side-item.is-active .polaris-side-icon,
  .polaris-side-group.is-open > .polaris-side-item .polaris-side-icon,
  .polaris-side-group.is-active > .polaris-side-item .polaris-side-icon {
    background: rgba(255, 255, 255, 0.98);
    border-color: rgba(15, 23, 42, 0.08);
  }
  .polaris-side-icon svg {
    width: 16px;
    height: 16px;
    stroke: currentColor;
    fill: none;
    stroke-width: 1.9;
    stroke-linecap: round;
    stroke-linejoin: round;
  }
  .polaris-side-item .polaris-side-label {
    font-size: 11px;
    line-height: 1;
    font-weight: 700;
    letter-spacing: 0.01em;
  }
  .polaris-side-group {
    position: relative;
    width: 100%;
  }
  .polaris-side-item[data-nav-toggle]::after {
    content: "";
    width: 7px;
    height: 7px;
    border-right: 1.8px solid currentColor;
    border-bottom: 1.8px solid currentColor;
    transform: rotate(45deg);
    margin-top: 2px;
    opacity: 0.55;
  }
  .polaris-side-submenu {
    position: absolute;
    left: calc(100% + 14px);
    top: 4px;
    width: 224px;
    padding: 14px;
    border-radius: 22px;
    background: rgba(255, 255, 255, 0.98);
    border: 1px solid rgba(15, 23, 42, 0.08);
    box-shadow: 0 14px 30px rgba(15, 23, 42, 0.08);
    display: grid;
    gap: 8px;
    opacity: 0;
    pointer-events: none;
    transform: translateX(-8px);
    transition: opacity .2s ease, transform .2s ease;
    z-index: 40;
  }
  .polaris-side-group:hover .polaris-side-submenu,
  .polaris-side-group.is-open .polaris-side-submenu,
  .polaris-side-group.is-active .polaris-side-submenu {
    opacity: 1;
    pointer-events: auto;
    transform: translateX(0);
  }
  .polaris-side-submenu-title {
    margin: 0 0 4px;
    color: #0f172a;
    font-size: 13px;
    font-weight: 700;
    letter-spacing: -0.01em;
  }
  .polaris-side-subitem {
    width: 100%;
    padding: 10px 12px;
    border-radius: 14px;
    text-align: left;
    background: rgba(249, 250, 251, 0.9);
    border: 1px solid transparent;
    color: #4b5563;
    font-size: 13px;
    font-weight: 600;
  }
  .polaris-side-subitem:hover,
  .polaris-side-subitem.is-active {
    background: rgba(15, 23, 42, 0.04);
    border-color: rgba(15, 23, 42, 0.08);
    color: #111827;
  }
  .polaris-side-logout {
    width: 100%;
    padding: 10px 4px;
    border-radius: 18px;
    background: rgba(255, 255, 255, 0.96);
    border: 1px solid rgba(15, 23, 42, 0.08);
    color: #4b5563;
    font-size: 11px;
    font-weight: 700;
  }
  .polaris-side-logout:hover {
    background: rgba(224, 79, 95, 0.08);
    border-color: rgba(224, 79, 95, 0.14);
    color: #c2414d;
  }
  .polaris-page-transition {
    position: fixed;
    inset: 0;
    z-index: 120;
    display: grid;
    place-items: center;
    background:
      radial-gradient(circle at 20% 20%, rgba(219, 234, 254, 0.7), transparent 26%),
      linear-gradient(180deg, rgba(250, 250, 250, 0.96), rgba(244, 244, 245, 0.98));
    opacity: 1;
    visibility: visible;
    transition: opacity .28s ease, visibility .28s ease;
  }
  .polaris-page-transition.is-ready {
    opacity: 0;
    visibility: hidden;
    pointer-events: none;
  }
  .polaris-page-transition.is-active {
    opacity: 1;
    visibility: visible;
    pointer-events: auto;
  }
  .polaris-page-transition-card {
    min-width: 220px;
    padding: 22px 24px;
    border-radius: 28px;
    background: rgba(255, 255, 255, 0.92);
    border: 1px solid rgba(15, 23, 42, 0.08);
    box-shadow: 0 18px 40px rgba(15, 23, 42, 0.08);
    display: grid;
    justify-items: center;
    gap: 12px;
  }
  .polaris-page-transition-card svg {
    width: 48px;
    height: 48px;
    display: block;
  }
  .polaris-page-transition-card strong {
    color: #0f172a;
    font-size: 16px;
    letter-spacing: -0.02em;
  }
  .polaris-page-transition-card span {
    color: #667085;
    font-size: 12px;
  }
  .polaris-page-transition-bar {
    width: 140px;
    height: 5px;
    border-radius: 999px;
    overflow: hidden;
    background: rgba(15, 23, 42, 0.08);
  }
  .polaris-page-transition-bar::before {
    content: "";
    display: block;
    width: 44%;
    height: 100%;
    border-radius: inherit;
    background: linear-gradient(90deg, #e0f2fe, #bfdbfe 36%, #93c5fd 70%, #dbeafe);
    animation: polaris-buffer-move 1.05s ease-in-out infinite;
  }
  @keyframes polaris-buffer-move {
    0% { transform: translateX(-115%); }
    100% { transform: translateX(330%); }
  }
  @media (max-width: 980px) {
    .polaris-app-shell {
      grid-template-columns: 74px minmax(0, 1fr);
      gap: 14px;
      padding: 14px;
    }
    .polaris-side-panel {
      top: 12px;
      height: calc(100vh - 24px);
      padding-inline: 6px;
      border-radius: 24px;
    }
    .polaris-side-icon {
      width: 32px;
      height: 32px;
      border-radius: 11px;
    }
    .polaris-side-submenu {
      left: calc(100% + 10px);
      width: 206px;
    }
  }
</style>
""".strip()


def dashboard_sidebar_html(active_key: str) -> str:
    report_group_active = active_key in {"attendance", "inventory", "refurb", "forecast"}
    report_group_class = "polaris-side-group is-active is-open" if report_group_active else "polaris-side-group"
    editor_active = "is-active" if active_key == "editor" else ""
    dashboard_active = "is-active" if active_key == "dashboard" else ""
    audit_active = "is-active" if active_key == "audit-logs" else ""
    analysis_active = "is-active" if active_key == "data-agent" else ""
    master_data_active = "is-active" if active_key == "master-data" else ""
    governance_active = "is-active" if active_key == "metric-dictionary" else ""
    sync_active = "is-active" if active_key == "sync" else ""

    def submenu_item(label: str, path: str, key: str) -> str:
        active = " is-active" if active_key == key else ""
        return f'<button class="polaris-side-subitem{active}" type="button" data-nav-target="{path}">{label}</button>'

    return f"""
<aside class="polaris-side-panel">
  <div class="polaris-side-brand">{dashboard_logo_badge_small_svg()}</div>
  <nav class="polaris-side-nav" aria-label="鍖楁瀬鏄熷鑸?>
    <button class="polaris-side-item {dashboard_active}" type="button" data-nav-target="{DASHBOARD_DEFAULT_PATH}">
      <span class="polaris-side-icon">
        <svg viewBox="0 0 18 18" aria-hidden="true"><path d="M2.8 8.4L9 3.1l6.2 5.3"/><path d="M4.6 7.7v7h8.8v-7"/><path d="M7.3 14.7v-4.2h3.4v4.2"/></svg>
      </span>
      <span class="polaris-side-label">缁忚惀</span>
    </button>
    <button class="polaris-side-item {editor_active}" type="button" data-nav-target="{DASHBOARD_EDITOR_PATH}">
      <span class="polaris-side-icon">
        <svg viewBox="0 0 18 18" aria-hidden="true"><rect x="3" y="3" width="12" height="12" rx="2"/><path d="M6 6h6M6 9h6M6 12h3.4"/></svg>
      </span>
      <span class="polaris-side-label">鐪嬫澘</span>
    </button>
    <button class="polaris-side-item {analysis_active}" type="button" data-nav-target="{DATA_AGENT_ENTRY_PATH}">
      <span class="polaris-side-icon">
        <svg viewBox="0 0 18 18" aria-hidden="true"><path d="M6.1 4.4h5.8"/><path d="M9 4.4v1.8"/><rect x="4.1" y="6.2" width="9.8" height="7.3" rx="2"/><circle cx="7.1" cy="9.8" r="0.7"/><circle cx="10.9" cy="9.8" r="0.7"/><path d="M7.5 12.1h3"/></svg>
      </span>
      <span class="polaris-side-label">鍒嗘瀽</span>
    </button>
    <div class="{report_group_class}">
      <button class="polaris-side-item {'is-active' if report_group_active else ''}" type="button" data-nav-toggle="reports" aria-expanded="{'true' if report_group_active else 'false'}">
        <span class="polaris-side-icon">
          <svg viewBox="0 0 18 18" aria-hidden="true"><path d="M4.2 3.6h6.6l3 3v7.6a1.8 1.8 0 0 1-1.8 1.8H4.2a1.8 1.8 0 0 1-1.8-1.8v-8.8a1.8 1.8 0 0 1 1.8-1.8Z"/><path d="M10.8 3.8v3h3"/><path d="M5.8 9.2h6.3M5.8 12h4.4"/></svg>
        </span>
        <span class="polaris-side-label">鎶ヨ〃</span>
      </button>
      <div class="polaris-side-submenu" role="menu" aria-label="鎶ヨ〃鑿滃崟">
        <p class="polaris-side-submenu-title">鎶ヨ〃涓庝笟鍔￠〉</p>
        {submenu_item("鍑哄嫟璁板綍", "/financial/bi-dashboard/attendance", "attendance")}
        {submenu_item("搴撳瓨鏄犲皠", "/financial/bi-dashboard/inventory-mappings", "inventory")}
        {submenu_item("缈绘柊鐢熶骇", "/financial/bi-dashboard/refurb-production", "refurb")}
        {submenu_item("棰勬祴棰勮", "/financial/bi-dashboard/forecast-alerts", "forecast")}
      </div>
    </div>
    <button class="polaris-side-item {sync_active}" type="button" data-nav-target="/financial/bi-dashboard/sync-schedule">
      <span class="polaris-side-icon">
        <svg viewBox="0 0 18 18" aria-hidden="true"><circle cx="9" cy="9" r="5.7"/><path d="M9 5.9v3.3l2.3 1.6"/><path d="M9 1.8v1.3M9 14.9v1.3M16.2 9h-1.3M3.1 9H1.8"/></svg>
      </span>
      <span class="polaris-side-label">璋冨害</span>
    </button>
    <button class="polaris-side-item {governance_active}" type="button" data-nav-target="{METRIC_DICTIONARY_ENTRY_PATH}">
      <span class="polaris-side-icon">
        <svg viewBox="0 0 18 18" aria-hidden="true"><path d="M4.1 4.2h9.8"/><path d="M4.1 9h9.8"/><path d="M4.1 13.8h6.2"/><circle cx="12.8" cy="13.8" r="1.5"/></svg>
      </span>
      <span class="polaris-side-label">娌荤悊</span>
    </button>
    <button class="polaris-side-item {master_data_active}" type="button" data-nav-target="{MASTER_DATA_ENTRY_PATH}">
      <span class="polaris-side-icon">
        <svg viewBox="0 0 18 18" aria-hidden="true"><rect x="3" y="3.2" width="12" height="11.6" rx="2"/><path d="M6 6.6h6M6 9h6M6 11.4h3.2"/></svg>
      </span>
      <span class="polaris-side-label">涓绘暟鎹?/span>
    </button>
    <button class="polaris-side-item {audit_active}" type="button" data-nav-target="{AUDIT_LOG_ENTRY_PATH}">
      <span class="polaris-side-icon">
        <svg viewBox="0 0 18 18" aria-hidden="true"><path d="M4 4.2h10"/><path d="M4 8.2h10"/><path d="M4 12.2h6"/><circle cx="12.8" cy="12.2" r="1.5"/><path d="M12.8 10v-2"/></svg>
      </span>
      <span class="polaris-side-label">瀹¤</span>
    </button>
  </nav>
  <button class="polaris-side-logout" type="button" data-nav-target="/financial/bi-dashboard/logout">閫€鍑?/button>
</aside>
""".strip()


def dashboard_transition_overlay_markup() -> str:
    return f"""
<div class="polaris-page-transition" id="polarisPageTransition" aria-hidden="true">
  <div class="polaris-page-transition-card">
    {dashboard_logo_badge_small_svg()}
    <strong>鍖楁瀬鏄熺粡钀ヤ腑蹇?/strong>
    <span>椤甸潰鍒囨崲涓紝璇风◢鍊?..</span>
    <div class="polaris-page-transition-bar"></div>
  </div>
</div>
""".strip()


def dashboard_transition_script() -> str:
    return """
<script>
  (() => {
    const overlay = document.getElementById("polarisPageTransition");
    if (!overlay) return;

    const closeOverlay = () => {
      overlay.classList.add("is-ready");
      overlay.classList.remove("is-active");
    };

    const openOverlay = () => {
      overlay.classList.remove("is-ready");
      overlay.classList.add("is-active");
    };

    window.polarisNavigate = (url, options = {}) => {
      if (!url) return;
      const target = String(url);
      if (!target) return;
      openOverlay();
      window.setTimeout(() => {
        window.location.href = target;
      }, Number(options.delay || 240));
    };

    document.querySelectorAll("[data-nav-target]").forEach((node) => {
      node.addEventListener("click", () => {
        const target = node.getAttribute("data-nav-target");
        if (target) window.polarisNavigate(target);
      });
    });

    document.querySelectorAll("[data-nav-toggle]").forEach((node) => {
      node.addEventListener("click", (event) => {
        event.stopPropagation();
        const group = node.closest(".polaris-side-group");
        if (!group) return;
        const next = !group.classList.contains("is-open");
        document.querySelectorAll(".polaris-side-group.is-open").forEach((item) => {
          if (item !== group && !item.classList.contains("is-active")) item.classList.remove("is-open");
          const toggle = item.querySelector("[data-nav-toggle]");
          if (toggle && item !== group && !item.classList.contains("is-active")) toggle.setAttribute("aria-expanded", "false");
        });
        group.classList.toggle("is-open", next);
        node.setAttribute("aria-expanded", next ? "true" : "false");
      });
    });

    document.addEventListener("click", () => {
      document.querySelectorAll(".polaris-side-group.is-open").forEach((group) => {
        if (group.classList.contains("is-active")) return;
        group.classList.remove("is-open");
        const toggle = group.querySelector("[data-nav-toggle]");
        if (toggle) toggle.setAttribute("aria-expanded", "false");
      });
    });

    window.addEventListener("pageshow", closeOverlay);
    window.addEventListener("load", () => window.setTimeout(closeOverlay, 170));
  })();
</script>
""".strip()


def dashboard_page_context(active_key: str, replacements: Dict[str, str] | None = None) -> Dict[str, str]:
    context = {
        "__BI_SHELL_HEAD__": dashboard_shell_head(),
        "__BI_SIDE_NAV__": dashboard_sidebar_html(active_key),
        "__BI_TRANSITION_OVERLAY__": dashboard_transition_overlay_markup(),
        "__BI_PAGE_TRANSITION_SCRIPT__": dashboard_transition_script(),
    }
    if replacements:
        context.update(replacements)
    return context


def is_local_port_open(host: str, port: int, timeout: float = 0.35) -> bool:
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except OSError:
        return False


def ensure_data_agent_local_config() -> Path:
    repo_dir = DATA_AGENT_REPO_DIR
    runtime_workspace = repo_dir / "runtime" / "workspace"
    runtime_csv = repo_dir / "runtime" / "csv"
    runtime_workspace.mkdir(parents=True, exist_ok=True)
    runtime_csv.mkdir(parents=True, exist_ok=True)
    config_path = repo_dir / "conf.yaml"
    if config_path.exists():
        return config_path

    config_content = f"""# 鏈湴鐢卞寳鏋佹槦鑷姩鐢熸垚鐨勬暟鎹垎鏋怉gent閰嶇疆
app:
  locale: "zh-CN"
  max_steps: 6
  max_retry_count: 3
  max_replan_count: 10
  plan_temperature: 0.8
  query_limit: 100000
  workspace_directory:
    linux: "/data/data_agent"
    windows: "{str(runtime_workspace).replace(chr(92), '/')}"
  csv_data_directory:
    linux: "/data/csv_files"
    windows: "{str(runtime_csv).replace(chr(92), '/')}"

llm:
  react_agent: &default_llm
    base_url: "https://api.openai.com/v1"
    model: "gpt-4o-mini"
    api_key: "$OPENAI_API_KEY"
  report_agent: *default_llm
  small_talk_agent: *default_llm
  intent_recognition_agent: *default_llm
  plan_agent: *default_llm
  analysis_agent: *default_llm
  extract: *default_llm

database:
  mysql:
    host: "127.0.0.1"
    port: 3306
    user: "bi_client"
    password: "$POLARIS_DATA_AGENT_DB_PASSWORD"
    database: "bi_center"

agents:
  capabilities:
    sale_agent:
      capabilities:
        - "鍒嗘瀽閿€鍞€佸簱瀛樸€侀€€璐с€佺炕鏂扮敓浜т笌缁忚惀缁撴灉"
    analysis_agent:
      capabilities:
        - "瀵瑰巻鍙叉暟鎹仛璁＄畻銆佸綊绾炽€佹尝鍔ㄥ垽鏂拰寮傚父鍒嗘瀽"
  data_sources:
    sale_agent:
      csv: []
      tables: []

ragflow:
  base_url: "http://127.0.0.1:9380"
  api_key: "$RAGFLOW_API_KEY"
  datasets: {{}}
"""
    config_path.write_text(config_content, encoding="utf-8")
    return config_path


def is_data_agent_api_online() -> bool:
    try:
        response = requests.get(f"{DATA_AGENT_API_URL}/openapi.json", timeout=1.5)
        if not response.ok:
            return False
        payload = response.json()
        title = str(payload.get("info", {}).get("title") or "")
        return "Data Agent" in title
    except Exception:
        return False


def is_data_agent_ui_online() -> bool:
    try:
        response = requests.get(DATA_AGENT_UI_URL, timeout=1.5)
        if not response.ok:
            return False
        body = response.text[:2000].lower()
        return "streamlit" in body or "data agent" in body
    except Exception:
        return False


def data_agent_env_snapshot() -> Dict[str, bool]:
    return {
        "openai_api_key": bool(os.getenv("OPENAI_API_KEY")),
        "ragflow_api_key": bool(os.getenv("RAGFLOW_API_KEY")),
        "db_password": bool(os.getenv("POLARIS_DATA_AGENT_DB_PASSWORD")),
    }


def data_agent_status_payload() -> Dict[str, Any]:
    repo_dir = DATA_AGENT_REPO_DIR
    config_path = ensure_data_agent_local_config()
    files = {
        "readme": repo_dir / "README_CN.md",
        "backend": repo_dir / "server.py",
        "frontend": repo_dir / "streamlit_app.py",
        "example_config": repo_dir / "conf.example.yaml",
        "config": config_path,
        "requirements": repo_dir / "requirements.txt",
    }
    env_ready = data_agent_env_snapshot()
    return {
        "module_name": DATA_AGENT_NAME,
        "display_name": "鏁版嵁鍒嗘瀽Agent",
        "github_url": DATA_AGENT_GITHUB_URL,
        "repo_path": str(repo_dir),
        "repo_present": repo_dir.exists(),
        "config_ready": files["config"].exists(),
        "files": {name: path.exists() for name, path in files.items()},
        "env_ready": env_ready,
        "api_url": DATA_AGENT_API_URL,
        "ui_url": DATA_AGENT_UI_URL,
        "api_online": is_data_agent_api_online(),
        "ui_online": is_data_agent_ui_online(),
        "startup_steps": [
            "cd /d \"vendor\\DataAgent\"",
            "pip install -r requirements.txt",
            "python server.py --host 127.0.0.1 --port 18080",
            "streamlit run streamlit_app.py --server.port 18501 --server.headless true",
        ],
        "capabilities": [
            "多智能体协作规划与执行",
            "基于 CSV / MySQL / Doris 的数据分析",
            "流式对话、代码执行与分析报告生成",
            "支持 MCP 与 RAGFlow 扩展",
        ],
        "integration_note": "当前阶段已升级为数据分析 Agent：北极星负责统一入口、状态探测、问答代理、报告生成与 AI 洞察接入，底层仍复用 DataAgent 作为独立智能体服务。",
    }

def json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def json_loads(value: Any, fallback: Any) -> Any:
    if isinstance(value, (dict, list)):
        return value
    if value in (None, ""):
        return fallback
    try:
        return json.loads(value)
    except Exception:
        return fallback


def audit_preview_values(rows: Sequence[Dict[str, Any]], field_name: str, limit: int = 5) -> List[str]:
    values: List[str] = []
    seen: set[str] = set()
    for item in rows:
        raw_value = str(item.get(field_name) or "").strip()
        if not raw_value or raw_value in seen:
            continue
        seen.add(raw_value)
        values.append(raw_value)
        if len(values) >= limit:
            break
    return values


def to_plain(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def sanitize_audit_detail(value: Any, depth: int = 0) -> Any:
    if depth >= 4:
        return "..."
    if isinstance(value, dict):
        items: Dict[str, Any] = {}
        for index, (key, item) in enumerate(value.items()):
            if index >= 18:
                items["__truncated__"] = f"+{len(value) - 18} more"
                break
            normalized_key = str(key)
            lowered = normalized_key.lower()
            if any(token in lowered for token in ("password", "token", "secret", "cookie", "authorization", "api_key")):
                items[normalized_key] = "[REDACTED]"
                continue
            items[normalized_key] = sanitize_audit_detail(item, depth + 1)
        return items
    if isinstance(value, (list, tuple, set)):
        values = list(value)
        result = [sanitize_audit_detail(item, depth + 1) for item in values[:8]]
        if len(values) > 8:
            result.append(f"... +{len(values) - 8} more")
        return result
    plain_value = to_plain(value)
    if isinstance(plain_value, str):
        text_value = plain_value.strip()
        return text_value if len(text_value) <= 240 else f"{text_value[:237]}..."
    return plain_value


def record_dashboard_audit(
    *,
    module_key: str,
    module_name: str,
    action_key: str,
    action_name: str,
    target_type: str = "",
    target_id: Any = "",
    target_name: str = "",
    result_status: str = "success",
    detail_summary: str = "",
    detail: Any = None,
    triggered_by: str = "",
    source_path: str = "",
    source_method: str = "",
    affected_count: int = 0,
) -> None:
    try:
        current_engine = get_engine()
        with current_engine.begin() as conn:
            conn.execute(
                text(
                    """
                    INSERT INTO bi_audit_log(
                        module_key, module_name, action_key, action_name,
                        target_type, target_id, target_name, result_status,
                        detail_summary, detail_json, source_path, source_method,
                        triggered_by, affected_count
                    ) VALUES (
                        :module_key, :module_name, :action_key, :action_name,
                        :target_type, :target_id, :target_name, :result_status,
                        :detail_summary, :detail_json, :source_path, :source_method,
                        :triggered_by, :affected_count
                    )
                    """
                ),
                {
                    "module_key": module_key,
                    "module_name": module_name,
                    "action_key": action_key,
                    "action_name": action_name,
                    "target_type": str(target_type or "")[:64],
                    "target_id": str(target_id or "")[:128],
                    "target_name": str(target_name or "")[:255],
                    "result_status": str(result_status or "success")[:32],
                    "detail_summary": str(detail_summary or "")[:255],
                    "detail_json": json_dumps(sanitize_audit_detail(detail or {})),
                    "source_path": str(source_path or "")[:255],
                    "source_method": str(source_method or "")[:16],
                    "triggered_by": str(triggered_by or "")[:64] or None,
                    "affected_count": max(0, int(affected_count or 0)),
                },
            )
    except Exception as exc:
        app_logger.warning("Dashboard audit write failed: %s", exc)


def to_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(str(value))
    except Exception:
        return None


def parse_date_or_none(raw_value: str | None) -> date | None:
    if not raw_value:
        return None
    try:
        return datetime.strptime(raw_value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"日期格式无效：{raw_value}") from exc


def parse_decimal_or_raise(raw_value: Any, field_name: str) -> Decimal:
    if raw_value in (None, ""):
        raise HTTPException(status_code=400, detail=f"{field_name}不能为空")
    try:
        value = Decimal(str(raw_value).strip())
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"{field_name}格式无效") from exc
    if value < 0:
        raise HTTPException(status_code=400, detail=f"{field_name}不能为负数")
    return value


def parse_int_or_default(raw_value: Any, default: int = 0) -> int:
    try:
        return int(str(raw_value).strip())
    except Exception:
        return default


def parse_bool_or_default(raw_value: Any, default: bool = False) -> bool:
    if raw_value in (None, ""):
        return default
    if isinstance(raw_value, bool):
        return raw_value
    return str(raw_value).strip().lower() in {"1", "true", "yes", "y", "on"}


def scheduler_timezone() -> ZoneInfo:
    return ZoneInfo(SYNC_SCHEDULER_TIMEZONE)


def validate_cron_expression(raw_value: Any) -> str:
    cron_expr = str(raw_value or "").strip()
    if not cron_expr:
        raise HTTPException(status_code=400, detail="Cron 表达式不能为空")
    try:
        CronTrigger.from_crontab(cron_expr, timezone=scheduler_timezone())
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Cron 表达式无效，请使用 5 段标准 crontab 格式") from exc
    return cron_expr


def load_sync_base_config() -> AppConfig:
    if not yonyou_sync_config_path.exists():
        raise RuntimeError("缺少 config/yonyou_inventory_sync.yaml，无法启动原始数据同步任务")
    return load_config(yonyou_sync_config_path)


def default_sync_schedule_settings() -> Dict[str, Any]:
    base_config = load_sync_base_config()
    cron_expr = str(base_config.job.cron or "").strip()
    return {
        "schedule_key": SYNC_SCHEDULE_KEY,
        "is_enabled": bool(cron_expr),
        "mode": "all",
        "cron_expr": cron_expr or "59 23 * * *",
        "sales_days_behind": max(0, int(base_config.job.sales_days_behind)),
        "sales_window_days": max(1, int(base_config.job.sales_window_days)),
        "snapshot_days_behind": max(0, int(base_config.job.snapshot_days_behind)),
    }


def normalize_sync_schedule_row(row: Dict[str, Any] | None) -> Dict[str, Any]:
    defaults = default_sync_schedule_settings()
    row = row or {}
    schedule = {
        "id": parse_int_or_default(row.get("id"), 0),
        "schedule_key": str(row.get("schedule_key") or defaults["schedule_key"]),
        "is_enabled": parse_bool_or_default(row.get("is_enabled"), defaults["is_enabled"]),
        "mode": str(row.get("mode") or defaults["mode"]).strip().lower(),
        "cron_expr": str(row.get("cron_expr") or defaults["cron_expr"]).strip(),
        "sales_days_behind": max(0, parse_int_or_default(row.get("sales_days_behind"), defaults["sales_days_behind"])),
        "sales_window_days": max(1, parse_int_or_default(row.get("sales_window_days"), defaults["sales_window_days"])),
        "snapshot_days_behind": max(0, parse_int_or_default(row.get("snapshot_days_behind"), defaults["snapshot_days_behind"])),
        "last_run_started_at": to_plain(row.get("last_run_started_at")),
        "last_run_finished_at": to_plain(row.get("last_run_finished_at")),
        "last_run_status": str(row.get("last_run_status") or "idle"),
        "last_run_message": str(row.get("last_run_message") or ""),
        "last_trigger": str(row.get("last_trigger") or ""),
        "last_result": json_loads(row.get("last_result_json"), {}),
        "updated_by": str(row.get("updated_by") or ""),
        "created_at": to_plain(row.get("created_at")),
        "updated_at": to_plain(row.get("updated_at")),
    }
    if schedule["mode"] not in {"all", "inventory", "sales"}:
        schedule["mode"] = defaults["mode"]
    if not schedule["cron_expr"]:
        schedule["cron_expr"] = defaults["cron_expr"]
    return schedule


def ensure_sync_schedule_seed(conn) -> None:
    existing = conn.execute(
        text(
            """
            SELECT id
            FROM bi_raw_sync_schedule_config
            WHERE schedule_key = :schedule_key
            """
        ),
        {"schedule_key": SYNC_SCHEDULE_KEY},
    ).first()
    if existing:
        return
    defaults = default_sync_schedule_settings()
    conn.execute(
        text(
            """
            INSERT INTO bi_raw_sync_schedule_config(
                schedule_key, is_enabled, mode, cron_expr, sales_days_behind,
                sales_window_days, snapshot_days_behind, last_run_status, last_run_message
            )
            VALUES(
                :schedule_key, :is_enabled, :mode, :cron_expr, :sales_days_behind,
                :sales_window_days, :snapshot_days_behind, 'idle', ''
            )
            """
        ),
        {
            "schedule_key": defaults["schedule_key"],
            "is_enabled": 1 if defaults["is_enabled"] else 0,
            "mode": defaults["mode"],
            "cron_expr": defaults["cron_expr"],
            "sales_days_behind": defaults["sales_days_behind"],
            "sales_window_days": defaults["sales_window_days"],
            "snapshot_days_behind": defaults["snapshot_days_behind"],
        },
    )


def load_sync_schedule(conn) -> Dict[str, Any]:
    ensure_sync_schedule_seed(conn)
    row = conn.execute(
        text(
            """
            SELECT
                id, schedule_key, is_enabled, mode, cron_expr, sales_days_behind,
                sales_window_days, snapshot_days_behind, last_run_started_at,
                last_run_finished_at, last_run_status, last_run_message, last_trigger,
                last_result_json, updated_by, created_at, updated_at
            FROM bi_raw_sync_schedule_config
            WHERE schedule_key = :schedule_key
            """
        ),
        {"schedule_key": SYNC_SCHEDULE_KEY},
    ).mappings().first()
    return normalize_sync_schedule_row(dict(row) if row else None)


def update_sync_schedule_runtime(
    *,
    last_run_started_at: datetime | None = None,
    last_run_finished_at: datetime | None = None,
    last_run_status: str | None = None,
    last_run_message: str | None = None,
    last_trigger: str | None = None,
    last_result: Dict[str, Any] | None = None,
) -> None:
    fields: Dict[str, Any] = {}
    if last_run_started_at is not None:
        fields["last_run_started_at"] = last_run_started_at
    if last_run_finished_at is not None:
        fields["last_run_finished_at"] = last_run_finished_at
    if last_run_status is not None:
        fields["last_run_status"] = last_run_status
    if last_run_message is not None:
        fields["last_run_message"] = last_run_message
    if last_trigger is not None:
        fields["last_trigger"] = last_trigger
    if last_result is not None:
        fields["last_result_json"] = json_dumps(last_result)
    if not fields:
        return
    fields["schedule_key"] = SYNC_SCHEDULE_KEY
    assignment_sql = ", ".join(f"{column} = :{column}" for column in fields if column != "schedule_key")
    current_engine = get_engine()
    with current_engine.begin() as conn:
        conn.execute(
            text(
                f"""
                UPDATE bi_raw_sync_schedule_config
                SET {assignment_sql}
                WHERE schedule_key = :schedule_key
                """
            ),
            fields,
        )


def build_sync_runtime_config(schedule: Dict[str, Any]) -> AppConfig:
    base_config = load_sync_base_config()
    base_config.job = JobConfig(
        cron=schedule["cron_expr"],
        sales_days_behind=schedule["sales_days_behind"],
        sales_window_days=schedule["sales_window_days"],
        snapshot_days_behind=schedule["snapshot_days_behind"],
    )
    return base_config


def execute_raw_sync(trigger: str = "manual") -> Dict[str, Any]:
    started_at = datetime.now()
    if not sync_run_lock.acquire(blocking=False):
        message = "已有原始数据同步任务在执行，本次触发已跳过。"
        update_sync_schedule_runtime(
            last_run_started_at=started_at,
            last_run_finished_at=started_at,
            last_run_status="skipped",
            last_run_message=message,
            last_trigger=trigger,
        )
        if trigger == "manual":
            raise RuntimeError(message)
        app_logger.warning(message)
        return {"status": "skipped", "message": message}

    try:
        ensure_schema()
        current_engine = get_engine()
        with current_engine.connect() as conn:
            schedule = load_sync_schedule(conn)
        update_sync_schedule_runtime(
            last_run_started_at=started_at,
            last_run_status="running",
            last_run_message="原始数据同步任务执行中...",
            last_trigger=trigger,
        )
        runtime_config = build_sync_runtime_config(schedule)
        job_logger = build_logger(runtime_config.logging)
        service = InventorySyncService(runtime_config, job_logger)
        snapshot_date, sales_start_date, sales_end_date = default_dates_from_job(runtime_config.job)
        run_result = service.run_once(
            mode=schedule["mode"],
            snapshot_date=snapshot_date,
            sales_start_date=sales_start_date,
            sales_end_date=sales_end_date,
            dry_run=False,
        )
        forecast_result = recalculate_forecasts_and_alerts(
            current_engine,
            updated_by=trigger,
            send_notifications=True,
        )
        finished_at = datetime.now()
        result_payload = {
            "mode": schedule["mode"],
            "snapshot_date": snapshot_date.isoformat(),
            "sales_start_date": sales_start_date.isoformat(),
            "sales_end_date": sales_end_date.isoformat(),
            **run_result,
            "forecast_refresh": forecast_result,
        }
        message = (
            f"执行完成：库存原始 {run_result.get('inventory_raw', 0)} 行，"
            f"销售原始 {run_result.get('sales_raw', 0)} 行，"
            f"预警 {forecast_result.get('alert_rows', 0)} 条。"
        )
        update_sync_schedule_runtime(
            last_run_finished_at=finished_at,
            last_run_status="success",
            last_run_message=message,
            last_trigger=trigger,
            last_result=result_payload,
        )
        return {"status": "success", "message": message, "result": result_payload}
    except Exception as exc:
        finished_at = datetime.now()
        message = f"{type(exc).__name__}: {exc}"
        update_sync_schedule_runtime(
            last_run_finished_at=finished_at,
            last_run_status="failed",
            last_run_message=message,
            last_trigger=trigger,
        )
        app_logger.exception("Raw sync job failed: %s", exc)
        raise
    finally:
        sync_run_lock.release()


def get_sync_scheduler() -> BackgroundScheduler:
    global sync_scheduler
    with sync_scheduler_lock:
        if sync_scheduler is None:
            sync_scheduler = BackgroundScheduler(
                timezone=scheduler_timezone(),
                job_defaults={"coalesce": True, "max_instances": 1, "misfire_grace_time": 3600},
            )
        return sync_scheduler


def sync_scheduler_snapshot() -> Dict[str, Any]:
    scheduler = sync_scheduler
    job = scheduler.get_job(SYNC_SCHEDULE_JOB_ID) if scheduler and scheduler.running else None
    weekly_job = scheduler.get_job(DATA_AGENT_WEEKLY_JOB_ID) if scheduler and scheduler.running else None
    monthly_job = scheduler.get_job(DATA_AGENT_MONTHLY_JOB_ID) if scheduler and scheduler.running else None
    return {
        "scheduler_running": bool(scheduler and scheduler.running),
        "is_running": sync_run_lock.locked(),
        "next_run_at": to_plain(job.next_run_time) if job and job.next_run_time else None,
        "weekly_report_next_run_at": to_plain(weekly_job.next_run_time) if weekly_job and weekly_job.next_run_time else None,
        "monthly_report_next_run_at": to_plain(monthly_job.next_run_time) if monthly_job and monthly_job.next_run_time else None,
        "timezone": SYNC_SCHEDULER_TIMEZONE,
    }


def refresh_sync_scheduler() -> Dict[str, Any]:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        schedule = load_sync_schedule(conn)

    scheduler = get_sync_scheduler()
    with sync_scheduler_lock:
        if not scheduler.running:
            scheduler.start()
        existing_job = scheduler.get_job(SYNC_SCHEDULE_JOB_ID)
        if existing_job is not None:
            scheduler.remove_job(SYNC_SCHEDULE_JOB_ID)
        if schedule["is_enabled"]:
            trigger = CronTrigger.from_crontab(schedule["cron_expr"], timezone=scheduler_timezone())
            scheduler.add_job(
                lambda: execute_raw_sync("scheduled"),
                trigger=trigger,
                id=SYNC_SCHEDULE_JOB_ID,
                replace_existing=True,
            )
        for job_id in (DATA_AGENT_WEEKLY_JOB_ID, DATA_AGENT_MONTHLY_JOB_ID):
            if scheduler.get_job(job_id) is not None:
                scheduler.remove_job(job_id)
        scheduler.add_job(
            run_weekly_data_agent_report,
            trigger=CronTrigger(day_of_week="mon", hour=9, minute=5, timezone=scheduler_timezone()),
            id=DATA_AGENT_WEEKLY_JOB_ID,
            replace_existing=True,
        )
        scheduler.add_job(
            run_monthly_data_agent_report,
            trigger=CronTrigger(day=1, hour=9, minute=10, timezone=scheduler_timezone()),
            id=DATA_AGENT_MONTHLY_JOB_ID,
            replace_existing=True,
        )
    runtime = sync_scheduler_snapshot()
    return {**schedule, **runtime}


def start_sync_scheduler() -> None:
    try:
        schedule = refresh_sync_scheduler()
        app_logger.info(
            "Raw sync scheduler initialized. enabled=%s cron=%s next_run_at=%s",
            schedule["is_enabled"],
            schedule["cron_expr"],
            schedule.get("next_run_at"),
        )
    except Exception as exc:  # pragma: no cover - startup guard
        app_logger.warning("Raw sync scheduler init failed: %s", exc)


def stop_sync_scheduler() -> None:
    global sync_scheduler
    with sync_scheduler_lock:
        if sync_scheduler and sync_scheduler.running:
            sync_scheduler.shutdown(wait=False)
        sync_scheduler = None


def serialize_sync_schedule(schedule: Dict[str, Any]) -> Dict[str, Any]:
    return {**schedule, **sync_scheduler_snapshot()}


def validate_sync_schedule_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    mode = str(payload.get("mode") or "all").strip().lower()
    if mode not in {"all", "inventory", "sales"}:
        raise HTTPException(status_code=400, detail="鍚屾妯″紡鏃犳晥")
    is_enabled = parse_bool_or_default(payload.get("is_enabled"), False)
    cron_expr = str(payload.get("cron_expr") or "").strip() or "59 23 * * *"
    if is_enabled:
        cron_expr = validate_cron_expression(cron_expr)
    normalized = {
        "is_enabled": is_enabled,
        "mode": mode,
        "cron_expr": cron_expr,
        "sales_days_behind": max(0, parse_int_or_default(payload.get("sales_days_behind"), 1)),
        "sales_window_days": max(1, min(31, parse_int_or_default(payload.get("sales_window_days"), 1))),
        "snapshot_days_behind": max(0, parse_int_or_default(payload.get("snapshot_days_behind"), 0)),
    }
    return normalized


def parse_numeric_or_zero(raw_value: Any) -> Decimal:
    if raw_value in (None, "", "-"):
        return Decimal("0")
    if isinstance(raw_value, Decimal):
        return raw_value
    normalized = str(raw_value).strip().replace(",", "")
    if normalized.endswith("%"):
        normalized = normalized[:-1]
    try:
        return Decimal(normalized)
    except Exception:
        return Decimal("0")


def parse_refurb_date(raw_value: Any, field_name: str = "鏃ユ湡") -> date:
    if isinstance(raw_value, datetime):
        return raw_value.date()
    if isinstance(raw_value, date):
        return raw_value
    text_value = str(raw_value or "").strip()
    if not text_value:
        raise HTTPException(status_code=400, detail=f"{field_name}不能为空")
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text_value, pattern).date()
        except ValueError:
            continue
    raise HTTPException(status_code=400, detail=f"{field_name}格式无效：{text_value}")


def calculate_refurb_metrics(payload: Dict[str, Any]) -> Dict[str, Decimal]:
    feeding_qty = parse_numeric_or_zero(payload.get("feeding_qty"))
    total_work_hours = parse_numeric_or_zero(payload.get("total_work_hours"))
    plan_qty = parse_numeric_or_zero(payload.get("plan_qty"))
    quality_defect_qty = parse_numeric_or_zero(payload.get("quality_defect_qty"))
    production_good_qty = parse_numeric_or_zero(payload.get("production_good_qty"))
    production_bad_qty = parse_numeric_or_zero(payload.get("production_bad_qty"))
    final_good_qty = production_good_qty - quality_defect_qty
    non_refurbishable_rate = Decimal("0")
    if feeding_qty != 0:
        non_refurbishable_rate = Decimal("1") - (final_good_qty / feeding_qty)
    quality_reject_rate = Decimal("0")
    if production_good_qty != 0:
        quality_reject_rate = quality_defect_qty / production_good_qty
    plan_achievement_rate = Decimal("0")
    if plan_qty != 0:
        plan_achievement_rate = final_good_qty / plan_qty
    refurb_efficiency = Decimal("0")
    if total_work_hours != 0:
        refurb_efficiency = final_good_qty / (total_work_hours / Decimal("8"))
    return {
        "feeding_qty": feeding_qty,
        "total_work_hours": total_work_hours,
        "plan_qty": plan_qty,
        "quality_defect_qty": quality_defect_qty,
        "production_good_qty": production_good_qty,
        "production_bad_qty": production_bad_qty,
        "final_good_qty": final_good_qty,
        "non_refurbishable_rate": non_refurbishable_rate,
        "quality_reject_rate": quality_reject_rate,
        "plan_achievement_rate": plan_achievement_rate,
        "refurb_efficiency": refurb_efficiency,
    }


def normalize_refurb_production_payload(payload: Dict[str, Any] | None) -> Dict[str, Any]:
    raw = payload or {}
    biz_date = parse_refurb_date(raw.get("biz_date"))
    refurb_category = str(raw.get("refurb_category") or "").strip()
    material_name = str(raw.get("material_name") or "").strip()
    if not refurb_category:
        raise HTTPException(status_code=400, detail="缈绘柊绉嶇被涓嶈兘涓虹┖")
    if not material_name:
        raise HTTPException(status_code=400, detail="鐗╂枡鍚嶇О涓嶈兘涓虹┖")
    metrics = calculate_refurb_metrics(raw)
    return {
        "biz_date": biz_date,
        "refurb_category": refurb_category,
        "material_name": material_name,
        **metrics,
    }


def parse_refurb_excel_rows(upload_file: UploadFile, content: bytes) -> List[Dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Excel 鏂囦欢鏃犳硶瑙ｆ瀽锛岃涓婁紶 .xlsx 鏂囦欢") from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 鏂囦欢娌℃湁鍙鍏ョ殑鏁版嵁")

    header_row_index = None
    for index, row in enumerate(rows):
        normalized = [str(cell or "").strip() for cell in row]
        if "鏃ユ湡" in normalized and "缈绘柊绉嶇被" in normalized and "鐗╂枡鍚嶇О" in normalized:
            header_row_index = index
            break
    if header_row_index is None:
        raise HTTPException(status_code=400, detail="Excel 表头未识别，请使用参考模板导入")

    header = [str(cell or "").strip() for cell in rows[header_row_index]]
    header_map = {name: idx for idx, name in enumerate(header) if name}
    required_headers = {
        "鏃ユ湡": "biz_date",
        "缈绘柊绉嶇被": "refurb_category",
        "鐗╂枡鍚嶇О": "material_name",
        "棰嗘枡鏁伴噺": "feeding_qty",
        "鎬昏€楄垂宸ユ椂": "total_work_hours",
        "璁″垝鏁伴噺": "plan_qty",
        "品质-检出不合格品": "quality_defect_qty",
        "鐢熶骇-浜у嚭鑹搧鏁伴噺": "production_good_qty",
        "生产-产出不良品数量": "production_bad_qty",
    }
    missing_headers = [name for name in required_headers if name not in header_map]
    if missing_headers:
        raise HTTPException(status_code=400, detail=f"Excel 缺少必要列：{'、'.join(missing_headers)}")

    normalized_rows: List[Dict[str, Any]] = []
    for row_number, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        candidate = {
            target: row[header_map[source]] if header_map[source] < len(row) else None
            for source, target in required_headers.items()
        }
        if not any(str(value or "").strip() for value in candidate.values()):
            continue
        try:
            normalized_rows.append(normalize_refurb_production_payload(candidate))
        except HTTPException as exc:
            raise HTTPException(status_code=400, detail=f"第 {row_number} 行导入失败：{exc.detail}") from exc
    if not normalized_rows:
        raise HTTPException(status_code=400, detail=f"{upload_file.filename or 'Excel 鏂囦欢'} 涓病鏈夊彲瀵煎叆鐨勬暟鎹")
    return normalized_rows


def save_refurb_production_rows(current_engine, rows: Sequence[Dict[str, Any]], updated_by: str) -> int:
    if not rows:
        return 0
    payload = []
    for row in rows:
        payload.append(
            {
                "biz_date": row["biz_date"],
                "refurb_category": row["refurb_category"],
                "material_name": row["material_name"],
                "feeding_qty": row["feeding_qty"],
                "total_work_hours": row["total_work_hours"],
                "plan_qty": row["plan_qty"],
                "quality_defect_qty": row["quality_defect_qty"],
                "production_good_qty": row["production_good_qty"],
                "production_bad_qty": row["production_bad_qty"],
                "final_good_qty": row["final_good_qty"],
                "non_refurbishable_rate": row["non_refurbishable_rate"],
                "quality_reject_rate": row["quality_reject_rate"],
                "plan_achievement_rate": row["plan_achievement_rate"],
                "refurb_efficiency": row["refurb_efficiency"],
                "updated_by": updated_by,
            }
        )
    with current_engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO bi_refurb_production_daily(
                    biz_date, refurb_category, material_name, feeding_qty, total_work_hours,
                    plan_qty, quality_defect_qty, production_good_qty, production_bad_qty,
                    final_good_qty, non_refurbishable_rate, quality_reject_rate,
                    plan_achievement_rate, refurb_efficiency, updated_by
                )
                VALUES(
                    :biz_date, :refurb_category, :material_name, :feeding_qty, :total_work_hours,
                    :plan_qty, :quality_defect_qty, :production_good_qty, :production_bad_qty,
                    :final_good_qty, :non_refurbishable_rate, :quality_reject_rate,
                    :plan_achievement_rate, :refurb_efficiency, :updated_by
                )
                ON DUPLICATE KEY UPDATE
                    feeding_qty = VALUES(feeding_qty),
                    total_work_hours = VALUES(total_work_hours),
                    plan_qty = VALUES(plan_qty),
                    quality_defect_qty = VALUES(quality_defect_qty),
                    production_good_qty = VALUES(production_good_qty),
                    production_bad_qty = VALUES(production_bad_qty),
                    final_good_qty = VALUES(final_good_qty),
                    non_refurbishable_rate = VALUES(non_refurbishable_rate),
                    quality_reject_rate = VALUES(quality_reject_rate),
                    plan_achievement_rate = VALUES(plan_achievement_rate),
                    refurb_efficiency = VALUES(refurb_efficiency),
                    updated_by = VALUES(updated_by)
                """
            ),
            payload,
        )
    return len(payload)


def refurb_production_summaries(conn, start_date: date | None = None, end_date: date | None = None, limit: int = 200) -> List[Dict[str, Any]]:
    query = """
        SELECT
            id, biz_date, refurb_category, material_name, feeding_qty, total_work_hours,
            plan_qty, quality_defect_qty, production_good_qty, production_bad_qty,
            final_good_qty, non_refurbishable_rate, quality_reject_rate,
            plan_achievement_rate, refurb_efficiency, updated_by, created_at, updated_at
        FROM bi_refurb_production_daily
        WHERE 1 = 1
    """
    params: Dict[str, Any] = {"limit": max(1, min(int(limit or 200), 1000))}
    if start_date is not None:
        query += " AND biz_date >= :start_date"
        params["start_date"] = start_date
    if end_date is not None:
        query += " AND biz_date <= :end_date"
        params["end_date"] = end_date
    query += " ORDER BY biz_date DESC, refurb_category ASC, material_name ASC LIMIT :limit"
    rows = conn.execute(text(query), params).mappings().all()
    result: List[Dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        for key in (
            "feeding_qty",
            "total_work_hours",
            "plan_qty",
            "quality_defect_qty",
            "production_good_qty",
            "production_bad_qty",
            "final_good_qty",
            "non_refurbishable_rate",
            "quality_reject_rate",
            "plan_achievement_rate",
            "refurb_efficiency",
        ):
            item[key] = to_plain(item.get(key))
        item["id"] = parse_int_or_default(item.get("id"), 0)
        item["biz_date"] = to_plain(item.get("biz_date"))
        item["created_at"] = to_plain(item.get("created_at"))
        item["updated_at"] = to_plain(item.get("updated_at"))
        result.append(item)
    return result


def normalize_forecast_manual_payload(payload: Dict[str, Any] | None) -> Dict[str, Any]:
    raw = payload or {}
    forecast_date = parse_refurb_date(raw.get("forecast_date"), "棰勬祴鏃ユ湡")
    material_name = str(raw.get("material_name") or "").strip()
    demand_type = str(raw.get("demand_type") or "").strip().lower()
    if not material_name:
        raise HTTPException(status_code=400, detail="鐗╂枡鍚嶇О涓嶈兘涓虹┖")
    if demand_type not in {"sales", "refurb"}:
        raise HTTPException(status_code=400, detail="闇€姹傜被鍨嬩粎鏀寔 sales 鎴?refurb")
    manual_qty = parse_numeric_or_zero(raw.get("manual_qty"))
    if manual_qty < 0:
        raise HTTPException(status_code=400, detail="鎵嬪姩棰勬祴鍊间笉鑳戒负璐熸暟")
    return {
        "forecast_date": forecast_date,
        "material_name": material_name,
        "demand_type": demand_type,
        "manual_qty": manual_qty,
        "notes": str(raw.get("notes") or "").strip(),
    }


def normalize_promotion_event_payload(rows: Sequence[Dict[str, Any]] | None) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    for index, row in enumerate(rows or [], start=1):
        event_name = str((row or {}).get("event_name") or "").strip()
        month_day_start = str((row or {}).get("month_day_start") or "").strip()
        month_day_end = str((row or {}).get("month_day_end") or "").strip()
        if not event_name or not month_day_start or not month_day_end:
            raise HTTPException(status_code=400, detail=f"第 {index} 条促销事件缺少名称或日期范围")
        try:
            datetime.strptime(f"2000-{month_day_start}", "%Y-%m-%d")
            datetime.strptime(f"2000-{month_day_end}", "%Y-%m-%d")
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"第 {index} 条促销事件日期格式无效，应为 MM-DD") from exc
        uplift_factor = parse_numeric_or_zero((row or {}).get("uplift_factor"))
        if uplift_factor <= 0:
            raise HTTPException(status_code=400, detail=f"第 {index} 条促销事件系数必须大于 0")
        normalized.append(
            {
                "id": parse_int_or_default((row or {}).get("id"), 0),
                "event_name": event_name,
                "month_day_start": month_day_start,
                "month_day_end": month_day_end,
                "uplift_factor": uplift_factor,
                "is_enabled": parse_bool_or_default((row or {}).get("is_enabled"), True),
            }
        )
    return normalized


def normalize_inventory_mapping_payload(
    rows: Sequence[Dict[str, Any]] | None,
    *,
    mapping_type: str,
) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    seen: set[str] = set()
    for index, item in enumerate(rows or []):
        if not isinstance(item, dict):
            continue
        normalized_item = {
            "id": parse_int_or_default(item.get("id"), 0),
            "sort_order": max(0, parse_int_or_default(item.get("sort_order"), (index + 1) * 10)),
            "is_enabled": 1 if bool(item.get("is_enabled", True)) else 0,
        }
        if mapping_type == "warehouse":
            source_name = str(item.get("source_warehouse_name") or "").strip()
            clean_name = str(item.get("warehouse_name_clean") or "").strip()
            if not source_name and not clean_name:
                continue
            if not source_name or not clean_name:
                raise HTTPException(status_code=400, detail="仓库映射的原始仓库和清洗后仓库都不能为空")
            if source_name in seen:
                raise HTTPException(status_code=400, detail=f"仓库映射重复：{source_name}")
            seen.add(source_name)
            normalized_item.update(
                {
                    "source_warehouse_name": source_name,
                    "warehouse_name_clean": clean_name,
                }
            )
        elif mapping_type == "status":
            stock_status_id = str(item.get("stock_status_id") or "").strip()
            stock_status_name = str(item.get("stock_status_name") or "").strip()
            if not stock_status_id and not stock_status_name:
                continue
            if not stock_status_id or not stock_status_name:
                raise HTTPException(status_code=400, detail="库存状态映射的状态 ID 和状态名称都不能为空")
            if stock_status_id in seen:
                raise HTTPException(status_code=400, detail=f"搴撳瓨鐘舵€佹槧灏勯噸澶嶏細{stock_status_id}")
            seen.add(stock_status_id)
            normalized_item.update(
                {
                    "stock_status_id": stock_status_id,
                    "stock_status_name": stock_status_name,
                }
            )
        else:
            raise ValueError(mapping_type)
        normalized.append(normalized_item)
    return normalized


def default_metric_dictionary_rows() -> List[Dict[str, Any]]:
    effective_date = date.today().replace(month=1, day=1)
    return [
        {
            "metric_key": "inventory_sellable_qty",
            "metric_name": "可售库存",
            "business_domain": "库存基础",
            "owner_role": "计划 / 仓配",
            "definition_text": "当前可直接用于销售或可计入口径供给的库存数量。",
            "formula_text": "按库存清洗表中计入可售或可供给口径的状态数量求和。",
            "source_table": "inventory_cleaning",
            "source_fields": "snapshot_date, material_code, warehouse_name_clean, stock_status_name, qty",
            "dimension_notes": "可按仓库 / SKU / 状态 / 日期统计",
            "version_tag": "v1",
            "effective_date": effective_date,
            "sort_order": 10,
            "is_enabled": 1,
            "created_by": "system",
            "updated_by": "system",
        },
        {
            "metric_key": "inventory_coverage_days",
            "metric_name": "库存覆盖天数",
            "business_domain": "计划预警",
            "owner_role": "计划",
            "definition_text": "当前可供库存按近期消耗节奏或预测需求可支撑的天数。",
            "formula_text": "可供库存 / 日均需求。当日均需求为 0 时按空值处理。",
            "source_table": "inventory_cleaning + bi_forecast_daily_snapshot",
            "source_fields": "current_stock_qty, forecast_qty, threshold_days",
            "dimension_notes": "可按物料 / 需求类型 / 日期统计",
            "version_tag": "v1",
            "effective_date": effective_date,
            "sort_order": 20,
            "is_enabled": 1,
            "created_by": "system",
            "updated_by": "system",
        },
        {
            "metric_key": "sales_income",
            "metric_name": "销售收入",
            "business_domain": "经营驾驶舱",
            "owner_role": "供应链 / 财务",
            "definition_text": "周期内按经营口径确认的销售收入。",
            "formula_text": "按销售清洗表在选定周期内对收入金额求和。",
            "source_table": "sales_cleaning",
            "source_fields": "biz_date, amount, platform_name, sku_code",
            "dimension_notes": "可按平台 / 店铺 / SKU / 日期统计",
            "version_tag": "v1",
            "effective_date": effective_date,
            "sort_order": 30,
            "is_enabled": 1,
            "created_by": "system",
            "updated_by": "system",
        },
        {
            "metric_key": "sales_return_amount",
            "metric_name": "退货金额",
            "business_domain": "退货分析",
            "owner_role": "售后 / 财务",
            "definition_text": "周期内发生的退货及其对应金额。",
            "formula_text": "按退货口径对退货单或反向销售记录金额求和。",
            "source_table": "sales_cleaning",
            "source_fields": "return_qty, return_amount, warehouse_name_clean",
            "dimension_notes": "可按平台 / SKU / 退货原因 / 日期统计",
            "version_tag": "v1",
            "effective_date": effective_date,
            "sort_order": 40,
            "is_enabled": 1,
            "created_by": "system",
            "updated_by": "system",
        },
        {
            "metric_key": "refurb_qualified_rate",
            "metric_name": "翻新合格率",
            "business_domain": "翻新生产",
            "owner_role": "翻新 / 品质",
            "definition_text": "翻新生产在给定周期内的合格率，用于反映产线质量表现。",
            "formula_text": "final_good_qty / production_good_qty，当 production_good_qty 为 0 时返回 0。",
            "source_table": "bi_refurb_production_daily",
            "source_fields": "production_good_qty, final_good_qty, quality_defect_qty",
            "dimension_notes": "可按日期 / 翻新类别 / 物料统计",
            "version_tag": "v1",
            "effective_date": effective_date,
            "sort_order": 50,
            "is_enabled": 1,
            "created_by": "system",
            "updated_by": "system",
        },
        {
            "metric_key": "forecast_bias_rate",
            "metric_name": "预测偏差率",
            "business_domain": "计划预警",
            "owner_role": "计划",
            "definition_text": "系统预测值与实际结果之间的偏离程度，用于衡量预测可信度。",
            "formula_text": "|预测值 - 实际值| / max(实际值, 1)。",
            "source_table": "bi_forecast_manual / bi_forecast_daily_snapshot / sales_cleaning",
            "source_fields": "manual_qty, ai_qty, actual_sales_qty",
            "dimension_notes": "可按物料 / 需求类型 / 月份统计",
            "version_tag": "v1",
            "effective_date": effective_date,
            "sort_order": 60,
            "is_enabled": 1,
            "created_by": "system",
            "updated_by": "system",
        },
        {
            "metric_key": "inventory_alert_count",
            "metric_name": "安全库存预警数",
            "business_domain": "计划预警",
            "owner_role": "计划 / 采购",
            "definition_text": "当前周期内触发安全库存风险的预警条数。",
            "formula_text": "按阈值规则对每个物料进行判断，命中则计数。",
            "source_table": "bi_forecast_daily_snapshot / bi_forecast_material_profile",
            "source_fields": "current_stock_qty, forecast_qty, threshold_days, alert_level",
            "dimension_notes": "可按物料 / 预警等级 / 日期统计",
            "version_tag": "v1",
            "effective_date": effective_date,
            "sort_order": 70,
            "is_enabled": 1,
            "created_by": "system",
            "updated_by": "system",
        },
        {
            "metric_key": "data_agent_report_coverage",
            "metric_name": "AI 报告覆盖率",
            "business_domain": "经营驾驶舱",
            "owner_role": "管理层 / 数据",
            "definition_text": "周报、月报对关键经营主题的覆盖情况，用于衡量 AI 报告的管理使用价值。",
            "formula_text": "已生成报告数 / 应生成报告数，按周期统计。",
            "source_table": "bi_data_agent_report",
            "source_fields": "report_type, period_start, period_end, status",
            "dimension_notes": "可按报告类型 / 周期统计",
            "version_tag": "v1",
            "effective_date": effective_date,
            "sort_order": 80,
            "is_enabled": 1,
            "created_by": "system",
            "updated_by": "system",
        },
    ]


def ensure_metric_dictionary_seed(conn) -> None:
    count = int(conn.execute(text("SELECT COUNT(*) FROM bi_metric_dictionary")).scalar() or 0)
    if count > 0:
        return
    conn.execute(
        text(
            """
            INSERT INTO bi_metric_dictionary(
                metric_key, metric_name, business_domain, owner_role,
                definition_text, formula_text, source_table, source_fields,
                dimension_notes, version_tag, effective_date, sort_order,
                is_enabled, created_by, updated_by
            ) VALUES (
                :metric_key, :metric_name, :business_domain, :owner_role,
                :definition_text, :formula_text, :source_table, :source_fields,
                :dimension_notes, :version_tag, :effective_date, :sort_order,
                :is_enabled, :created_by, :updated_by
            )
            """
        ),
        default_metric_dictionary_rows(),
    )


def normalize_metric_dictionary_payload(rows: Sequence[Dict[str, Any]] | None) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    seen_keys: set[str] = set()
    for index, item in enumerate(rows or []):
        if not isinstance(item, dict):
            continue
        metric_key = str(item.get("metric_key") or "").strip()
        metric_name = str(item.get("metric_name") or "").strip()
        business_domain = str(item.get("business_domain") or "").strip()
        owner_role = str(item.get("owner_role") or "").strip()
        definition_text = str(item.get("definition_text") or "").strip()
        formula_text = str(item.get("formula_text") or "").strip()
        source_table = str(item.get("source_table") or "").strip()
        source_fields = str(item.get("source_fields") or "").strip()
        dimension_notes = str(item.get("dimension_notes") or "").strip()
        version_tag = str(item.get("version_tag") or "v1").strip()
        effective_date_raw = str(item.get("effective_date") or "").strip()
        if not any(
            [
                metric_key,
                metric_name,
                business_domain,
                owner_role,
                definition_text,
                formula_text,
                source_table,
                source_fields,
                dimension_notes,
            ]
        ):
            continue
        if not metric_key or not metric_name:
            raise HTTPException(status_code=400, detail=f"第 {index + 1} 条指标编码和名称不能为空")
        if metric_key in seen_keys:
            raise HTTPException(status_code=400, detail=f"指标编码重复：{metric_key}")
        seen_keys.add(metric_key)
        effective_date = None
        if effective_date_raw:
            try:
                effective_date = datetime.strptime(effective_date_raw, "%Y-%m-%d").date()
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=f"指标 {metric_key} 的生效日期格式应为 YYYY-MM-DD") from exc
        normalized.append(
            {
                "id": parse_int_or_default(item.get("id"), 0),
                "metric_key": metric_key,
                "metric_name": metric_name,
                "business_domain": business_domain or "经营驾驶舱",
                "owner_role": owner_role,
                "definition_text": definition_text,
                "formula_text": formula_text,
                "source_table": source_table,
                "source_fields": source_fields,
                "dimension_notes": dimension_notes,
                "version_tag": version_tag or "v1",
                "effective_date": effective_date,
                "sort_order": max(0, parse_int_or_default(item.get("sort_order"), (index + 1) * 10)),
                "is_enabled": 1 if bool(item.get("is_enabled", True)) else 0,
            }
        )
    return normalized


def table_exists(conn, table_name: str) -> bool:
    return bool(conn.execute(text("SHOW TABLES LIKE :table_name"), {"table_name": table_name}).fetchone())


def ensure_table_columns(conn, table_name: str, column_definitions: Dict[str, str]) -> None:
    existing = {row[0] for row in conn.execute(text(f"SHOW COLUMNS FROM {table_name}")).fetchall()}
    for column_name, definition_sql in column_definitions.items():
        if column_name in existing:
            continue
        conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition_sql}"))


def normalize_master_warehouse_payload(rows: Sequence[Dict[str, Any]] | None) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    seen_source_names: set[str] = set()
    for index, item in enumerate(rows or []):
        if not isinstance(item, dict):
            continue
        source_name = str(item.get("source_warehouse_name") or "").strip()
        clean_name = str(item.get("warehouse_name_clean") or "").strip()
        if not source_name and not clean_name:
            continue
        if not source_name or not clean_name:
            raise HTTPException(status_code=400, detail=f"第 {index + 1} 条仓库主数据缺少原始仓库或统一仓库名称")
        if source_name in seen_source_names:
            raise HTTPException(status_code=400, detail=f"浠撳簱涓绘暟鎹噸澶嶏細{source_name}")
        seen_source_names.add(source_name)
        normalized.append(
            {
                "id": parse_int_or_default(item.get("id"), 0),
                "source_warehouse_name": source_name,
                "warehouse_name_clean": clean_name,
                "warehouse_code": str(item.get("warehouse_code") or clean_name).strip(),
                "warehouse_type": str(item.get("warehouse_type") or "").strip(),
                "platform_owner": str(item.get("platform_owner") or "").strip(),
                "city": str(item.get("city") or "").strip(),
                "is_sellable_warehouse": 1 if parse_bool_or_default(item.get("is_sellable_warehouse"), False) else 0,
                "is_reverse_warehouse": 1 if parse_bool_or_default(item.get("is_reverse_warehouse"), False) else 0,
                "sort_order": max(0, parse_int_or_default(item.get("sort_order"), (index + 1) * 10)),
                "is_enabled": 1 if parse_bool_or_default(item.get("is_enabled"), True) else 0,
            }
        )
    return normalized


def normalize_master_status_payload(rows: Sequence[Dict[str, Any]] | None) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    seen_status_ids: set[str] = set()
    for index, item in enumerate(rows or []):
        if not isinstance(item, dict):
            continue
        status_id = str(item.get("stock_status_id") or "").strip()
        status_name = str(item.get("stock_status_name") or "").strip()
        if not status_id and not status_name:
            continue
        if not status_id or not status_name:
            raise HTTPException(status_code=400, detail=f"第 {index + 1} 条库存状态缺少状态 ID 或状态名称")
        if status_id in seen_status_ids:
            raise HTTPException(status_code=400, detail=f"搴撳瓨鐘舵€侀噸澶嶏細{status_id}")
        seen_status_ids.add(status_id)
        normalized.append(
            {
                "id": parse_int_or_default(item.get("id"), 0),
                "stock_status_id": status_id,
                "stock_status_name": status_name,
                "status_group": str(item.get("status_group") or "").strip(),
                "can_sell": 1 if parse_bool_or_default(item.get("can_sell"), False) else 0,
                "can_forecast_supply": 1 if parse_bool_or_default(item.get("can_forecast_supply"), False) else 0,
                "need_quality_check": 1 if parse_bool_or_default(item.get("need_quality_check"), False) else 0,
                "next_default_status": str(item.get("next_default_status") or "").strip(),
                "sort_order": max(0, parse_int_or_default(item.get("sort_order"), (index + 1) * 10)),
                "is_enabled": 1 if parse_bool_or_default(item.get("is_enabled"), True) else 0,
            }
        )
    return normalized


def normalize_sku_master_payload(rows: Sequence[Dict[str, Any]] | None) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    seen_codes: set[str] = set()
    for index, item in enumerate(rows or []):
        if not isinstance(item, dict):
            continue
        sku_code = str(item.get("sku_code") or "").strip()
        sku_name = str(item.get("sku_name") or "").strip()
        if not sku_code and not sku_name:
            continue
        if not sku_code or not sku_name:
            raise HTTPException(status_code=400, detail=f"第 {index + 1} 条 SKU 缺少编码或名称")
        if sku_code in seen_codes:
            raise HTTPException(status_code=400, detail=f"SKU 编码重复：{sku_code}")
        seen_codes.add(sku_code)
        normalized.append(
            {
                "id": parse_int_or_default(item.get("id"), 0),
                "sku_code": sku_code,
                "sku_name": sku_name,
                "sku_type": str(item.get("sku_type") or "").strip(),
                "product_line": str(item.get("product_line") or "").strip(),
                "model": str(item.get("model") or "").strip(),
                "spec_version": str(item.get("spec_version") or "").strip(),
                "lifecycle_status": str(item.get("lifecycle_status") or "").strip(),
                "owner_dept": str(item.get("owner_dept") or "").strip(),
                "sort_order": max(0, parse_int_or_default(item.get("sort_order"), (index + 1) * 10)),
                "is_active": 1 if parse_bool_or_default(item.get("is_active"), True) else 0,
            }
        )
    return normalized


def normalize_channel_shop_payload(rows: Sequence[Dict[str, Any]] | None) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    seen_pairs: set[tuple[str, str]] = set()
    for index, item in enumerate(rows or []):
        if not isinstance(item, dict):
            continue
        channel_code = str(item.get("channel_code") or "").strip()
        channel_name = str(item.get("channel_name") or "").strip()
        shop_name = str(item.get("shop_name") or "").strip()
        if not channel_code and not channel_name and not shop_name:
            continue
        if not channel_code or not channel_name:
            raise HTTPException(status_code=400, detail=f"第 {index + 1} 条渠道店铺缺少渠道编码或渠道名称")
        pair_key = (channel_code, shop_name)
        if pair_key in seen_pairs:
            raise HTTPException(status_code=400, detail=f"渠道店铺重复：{channel_code} / {shop_name}")
        seen_pairs.add(pair_key)
        normalized.append(
            {
                "id": parse_int_or_default(item.get("id"), 0),
                "channel_code": channel_code,
                "channel_name": channel_name,
                "shop_name": shop_name,
                "platform_name": str(item.get("platform_name") or "").strip(),
                "owner_dept": str(item.get("owner_dept") or "").strip(),
                "sort_order": max(0, parse_int_or_default(item.get("sort_order"), (index + 1) * 10)),
                "is_active": 1 if parse_bool_or_default(item.get("is_active"), True) else 0,
            }
        )
    return normalized


def procurement_arrival_status_options() -> List[Dict[str, str]]:
    return [
        {"value": "draft", "label": "草稿"},
        {"value": "ready", "label": "待执行"},
        {"value": "completed", "label": "已到货"},
        {"value": "exception", "label": "异常"},
    ]


def procurement_document_status_options() -> List[Dict[str, str]]:
    return [
        {"value": "pending", "label": "待编排"},
        {"value": "generated", "label": "已生成"},
        {"value": "synced", "label": "已回写"},
        {"value": "failed", "label": "失败待补救"},
    ]


def generate_procurement_arrival_no() -> str:
    return f"CGDH{datetime.now().strftime('%Y%m%d%H%M%S')}{int(time.time() * 1000) % 1000:03d}"


def default_procurement_arrival_rows() -> List[Dict[str, Any]]:
    today = date.today()
    return [
        {
            "arrival_no": f"CGDH{today.strftime('%Y%m%d')}001",
            "purchase_order_no": f"PO{today.strftime('%Y%m')}001",
            "supplier_name": "深圳智学电子",
            "warehouse_code": "SZ-FG",
            "warehouse_name": "深圳成品仓",
            "channel_code": "ALL",
            "channel_name": "公共补货",
            "sku_code": "TAB-A12-128G",
            "sku_name": "学习平板 A12 128G",
            "expected_qty": 600,
            "arrived_qty": 560,
            "qualified_qty": 552,
            "exception_qty": 8,
            "unit": "台",
            "arrival_date": today - timedelta(days=1),
            "status": "ready",
            "document_status": "generated",
            "exception_reason": "",
            "remark": "首批到货，待系统回写入库结果",
            "source_system": "manual",
            "created_by": "system",
            "updated_by": "system",
            "sort_order": 10,
        },
        {
            "arrival_no": f"CGDH{today.strftime('%Y%m%d')}002",
            "purchase_order_no": f"PO{today.strftime('%Y%m')}002",
            "supplier_name": "东莞启程包装",
            "warehouse_code": "SZ-RV",
            "warehouse_name": "深圳逆向仓",
            "channel_code": "JD",
            "channel_name": "京东自营",
            "sku_code": "ACC-CASE-A12",
            "sku_name": "学习平板 A12 原装保护套",
            "expected_qty": 400,
            "arrived_qty": 400,
            "qualified_qty": 400,
            "exception_qty": 0,
            "unit": "件",
            "arrival_date": today - timedelta(days=2),
            "status": "completed",
            "document_status": "synced",
            "exception_reason": "",
            "remark": "已完成到货和单据回写",
            "source_system": "manual",
            "created_by": "system",
            "updated_by": "system",
            "sort_order": 20,
        },
        {
            "arrival_no": f"CGDH{today.strftime('%Y%m%d')}003",
            "purchase_order_no": f"PO{today.strftime('%Y%m')}003",
            "supplier_name": "上海优测屏显",
            "warehouse_code": "SH-QC",
            "warehouse_name": "上海质检仓",
            "channel_code": "TMALL",
            "channel_name": "天猫旗舰店",
            "sku_code": "TAB-PRO-256G",
            "sku_name": "学习平板 Pro 256G",
            "expected_qty": 260,
            "arrived_qty": 260,
            "qualified_qty": 248,
            "exception_qty": 12,
            "unit": "台",
            "arrival_date": today,
            "status": "exception",
            "document_status": "failed",
            "exception_reason": "屏幕瑕疵 12 台，待供应商补发",
            "remark": "异常已上报，等待补偿和重新编排",
            "source_system": "manual",
            "created_by": "system",
            "updated_by": "system",
            "sort_order": 30,
        },
        {
            "arrival_no": f"CGDH{today.strftime('%Y%m%d')}004",
            "purchase_order_no": f"PO{today.strftime('%Y%m')}004",
            "supplier_name": "深圳智学电子",
            "warehouse_code": "WH-EAST",
            "warehouse_name": "武汉东区仓",
            "channel_code": "DY",
            "channel_name": "抖音商城",
            "sku_code": "TAB-LITE-64G",
            "sku_name": "学习平板 Lite 64G",
            "expected_qty": 320,
            "arrived_qty": 0,
            "qualified_qty": 0,
            "exception_qty": 0,
            "unit": "台",
            "arrival_date": today + timedelta(days=1),
            "status": "draft",
            "document_status": "pending",
            "exception_reason": "",
            "remark": "供应商已确认发车，待收货",
            "source_system": "manual",
            "created_by": "system",
            "updated_by": "system",
            "sort_order": 40,
        },
    ]


def ensure_procurement_arrival_seed(conn) -> None:
    count = int(conn.execute(text("SELECT COUNT(*) FROM bi_procurement_arrival")).scalar() or 0)
    if count > 0:
        return
    conn.execute(
        text(
            """
            INSERT INTO bi_procurement_arrival(
                arrival_no, purchase_order_no, supplier_name, warehouse_code, warehouse_name,
                channel_code, channel_name, sku_code, sku_name, expected_qty, arrived_qty,
                qualified_qty, exception_qty, unit, arrival_date, status, document_status,
                exception_reason, remark, source_system, created_by, updated_by, sort_order
            ) VALUES (
                :arrival_no, :purchase_order_no, :supplier_name, :warehouse_code, :warehouse_name,
                :channel_code, :channel_name, :sku_code, :sku_name, :expected_qty, :arrived_qty,
                :qualified_qty, :exception_qty, :unit, :arrival_date, :status, :document_status,
                :exception_reason, :remark, :source_system, :created_by, :updated_by, :sort_order
            )
            """
        ),
        default_procurement_arrival_rows(),
    )


def normalize_procurement_arrival_payload(payload: Dict[str, Any] | None) -> Dict[str, Any]:
    item = payload if isinstance(payload, dict) else {}
    purchase_order_no = str(item.get("purchase_order_no") or "").strip()
    supplier_name = str(item.get("supplier_name") or "").strip()
    warehouse_code = str(item.get("warehouse_code") or "").strip()
    warehouse_name = str(item.get("warehouse_name") or "").strip()
    sku_code = str(item.get("sku_code") or "").strip()
    sku_name = str(item.get("sku_name") or "").strip()
    arrival_date = parse_date_or_none(str(item.get("arrival_date") or "").strip())
    if not purchase_order_no:
        raise HTTPException(status_code=400, detail="采购单号不能为空")
    if not supplier_name:
        raise HTTPException(status_code=400, detail="供应商不能为空")
    if not warehouse_code:
        raise HTTPException(status_code=400, detail="到货仓不能为空")
    if not sku_code or not sku_name:
        raise HTTPException(status_code=400, detail="SKU 编码和 SKU 名称不能为空")
    if arrival_date is None:
        raise HTTPException(status_code=400, detail="到货日期不能为空，且格式必须为 YYYY-MM-DD")

    expected_qty = max(0.0, to_number(item.get("expected_qty")) or 0.0)
    arrived_qty = max(0.0, to_number(item.get("arrived_qty")) or 0.0)
    qualified_qty = max(0.0, to_number(item.get("qualified_qty")) or 0.0)
    exception_qty = max(0.0, to_number(item.get("exception_qty")) or 0.0)

    if qualified_qty > arrived_qty:
        qualified_qty = arrived_qty
    exception_cap = max(arrived_qty - qualified_qty, 0.0)
    if exception_qty == 0.0 and arrived_qty > qualified_qty:
        exception_qty = exception_cap
    elif exception_qty > exception_cap:
        exception_qty = exception_cap

    status = str(item.get("status") or "draft").strip().lower() or "draft"
    document_status = str(item.get("document_status") or "pending").strip().lower() or "pending"
    valid_statuses = {option["value"] for option in procurement_arrival_status_options()}
    valid_document_statuses = {option["value"] for option in procurement_document_status_options()}
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail="閲囪喘鍒拌揣鐘舵€佷笉鍚堟硶")
    if document_status not in valid_document_statuses:
        raise HTTPException(status_code=400, detail="鍗曟嵁缂栨帓鐘舵€佷笉鍚堟硶")

    return {
        "id": parse_int_or_default(item.get("id"), 0),
        "arrival_no": str(item.get("arrival_no") or "").strip() or generate_procurement_arrival_no(),
        "purchase_order_no": purchase_order_no,
        "supplier_name": supplier_name,
        "warehouse_code": warehouse_code,
        "warehouse_name": warehouse_name or warehouse_code,
        "channel_code": str(item.get("channel_code") or "").strip(),
        "channel_name": str(item.get("channel_name") or "").strip(),
        "sku_code": sku_code,
        "sku_name": sku_name,
        "expected_qty": round(expected_qty, 2),
        "arrived_qty": round(arrived_qty, 2),
        "qualified_qty": round(qualified_qty, 2),
        "exception_qty": round(exception_qty, 2),
        "unit": str(item.get("unit") or "台").strip() or "台",
        "arrival_date": arrival_date,
        "status": status,
        "document_status": document_status,
        "exception_reason": str(item.get("exception_reason") or "").strip(),
        "remark": str(item.get("remark") or "").strip(),
        "source_system": str(item.get("source_system") or "manual").strip() or "manual",
        "sort_order": max(0, parse_int_or_default(item.get("sort_order"), 100)),
    }


def serialize_procurement_arrival_row(row: Dict[str, Any]) -> Dict[str, Any]:
    item = {key: to_plain(value) for key, value in row.items()}
    expected_qty = float(item.get("expected_qty") or 0.0)
    arrived_qty = float(item.get("arrived_qty") or 0.0)
    qualified_qty = float(item.get("qualified_qty") or 0.0)
    exception_qty = float(item.get("exception_qty") or 0.0)
    item["expected_qty"] = expected_qty
    item["arrived_qty"] = arrived_qty
    item["qualified_qty"] = qualified_qty
    item["exception_qty"] = exception_qty
    item["pending_qty"] = max(expected_qty - arrived_qty, 0.0)
    item["fulfillment_rate"] = round(arrived_qty / expected_qty, 4) if expected_qty > 0 else 0.0
    item["quality_rate"] = round(qualified_qty / arrived_qty, 4) if arrived_qty > 0 else 0.0
    return item


def inventory_flow_action_options() -> List[Dict[str, str]]:
    return [
        {"value": "status_transition", "label": "状态流转"},
        {"value": "warehouse_transfer", "label": "仓间调拨"},
    ]


def inventory_flow_task_status_options() -> List[Dict[str, str]]:
    return [
        {"value": "draft", "label": "草稿"},
        {"value": "pending", "label": "待执行"},
        {"value": "completed", "label": "已完成"},
        {"value": "blocked", "label": "阻塞"},
        {"value": "cancelled", "label": "已取消"},
    ]


def inventory_flow_priority_options() -> List[Dict[str, str]]:
    return [
        {"value": "high", "label": "高"},
        {"value": "normal", "label": "中"},
        {"value": "low", "label": "低"},
    ]


def inventory_flow_trigger_options() -> List[Dict[str, str]]:
    return [
        {"value": "procurement_arrival", "label": "采购到货"},
        {"value": "manual", "label": "人工建单"},
        {"value": "after_sales", "label": "售后逆向"},
        {"value": "refurb", "label": "翻新生产"},
    ]


def generate_inventory_flow_task_no() -> str:
    return f"KCLZ{datetime.now().strftime('%Y%m%d%H%M%S')}{int(time.time() * 1000) % 1000:03d}"


def default_inventory_flow_rules(conn) -> List[Dict[str, Any]]:
    status_rows = conn.execute(
        text(
            """
            SELECT stock_status_id, stock_status_name
            FROM bi_inventory_status_map
            WHERE is_enabled = 1
            """
        )
    ).mappings().all()
    warehouse_rows = conn.execute(
        text(
            """
            SELECT warehouse_code, warehouse_name_clean
            FROM bi_inventory_warehouse_map
            WHERE is_enabled = 1
            """
        )
    ).mappings().all()
    status_by_name = {
        str(row["stock_status_name"] or "").strip(): {
            "id": str(row["stock_status_id"] or "").strip(),
            "name": str(row["stock_status_name"] or "").strip(),
        }
        for row in status_rows
    }
    warehouse_by_name = {
        str(row["warehouse_name_clean"] or "").strip(): {
            "code": str(row["warehouse_code"] or "").strip(),
            "name": str(row["warehouse_name_clean"] or "").strip(),
        }
        for row in warehouse_rows
    }

    def status(name: str) -> tuple[str, str]:
        matched = status_by_name.get(name) or {"id": "", "name": name}
        return matched["id"], matched["name"]

    def warehouse(name: str) -> tuple[str, str]:
        matched = warehouse_by_name.get(name) or {"code": "", "name": name}
        return matched["code"], matched["name"]

    qc_id, qc_name = status("待检")
    good_id, good_name = status("采购良品")
    bad_id, bad_name = status("采购不良品")
    reverse_bad_id, reverse_bad_name = status("不良品")
    refurb_good_id, refurb_good_name = status("翻新良品")
    sellable_wh_code, sellable_wh_name = warehouse("良品仓")
    bad_wh_code, bad_wh_name = warehouse("涓嶈壇鍝佷粨")
    reverse_wh_code, reverse_wh_name = warehouse("销退仓")
    refurb_wh_code, refurb_wh_name = warehouse("生产仓")

    return [
        {
            "rule_name": "采购到货合格转采购良品",
            "trigger_source": "procurement_arrival",
            "trigger_condition": "qualified",
            "action_type": "status_transition",
            "source_status_id": qc_id,
            "source_status_name": qc_name,
            "target_status_id": good_id,
            "target_status_name": good_name,
            "source_warehouse_code": "",
            "source_warehouse_name": "",
            "target_warehouse_code": sellable_wh_code,
            "target_warehouse_name": sellable_wh_name,
            "priority": "high",
            "auto_create_task": 1,
            "is_enabled": 1,
            "sort_order": 10,
            "note": "采购到货合格数量进入待执行库存流转。",
            "created_by": "system",
            "updated_by": "system",
        },
        {
            "rule_name": "閲囪喘鍒拌揣寮傚父杞噰璐笉鑹搧",
            "trigger_source": "procurement_arrival",
            "trigger_condition": "exception",
            "action_type": "warehouse_transfer",
            "source_status_id": qc_id,
            "source_status_name": qc_name,
            "target_status_id": bad_id,
            "target_status_name": bad_name,
            "source_warehouse_code": "",
            "source_warehouse_name": "",
            "target_warehouse_code": bad_wh_code,
            "target_warehouse_name": bad_wh_name,
            "priority": "high",
            "auto_create_task": 1,
            "is_enabled": 1,
            "sort_order": 20,
            "note": "采购异常数量转入不良品仓，待后续补偿或报损。",
            "created_by": "system",
            "updated_by": "system",
        },
        {
            "rule_name": "销退不良转翻新生产",
            "trigger_source": "manual",
            "trigger_condition": "manual",
            "action_type": "warehouse_transfer",
            "source_status_id": reverse_bad_id,
            "source_status_name": reverse_bad_name,
            "target_status_id": refurb_good_id,
            "target_status_name": refurb_good_name,
            "source_warehouse_code": reverse_wh_code,
            "source_warehouse_name": reverse_wh_name,
            "target_warehouse_code": refurb_wh_code,
            "target_warehouse_name": refurb_wh_name,
            "priority": "normal",
            "auto_create_task": 0,
            "is_enabled": 1,
            "sort_order": 30,
            "note": "售后逆向件转入翻新生产处理。",
            "created_by": "system",
            "updated_by": "system",
        },
    ]


def ensure_inventory_flow_seed(conn) -> None:
    count = int(conn.execute(text("SELECT COUNT(*) FROM bi_inventory_flow_rule")).scalar() or 0)
    if count > 0:
        return
    conn.execute(
        text(
            """
            INSERT INTO bi_inventory_flow_rule(
                rule_name, trigger_source, trigger_condition, action_type,
                source_status_id, source_status_name, target_status_id, target_status_name,
                source_warehouse_code, source_warehouse_name, target_warehouse_code, target_warehouse_name,
                priority, auto_create_task, is_enabled, sort_order, note, created_by, updated_by
            ) VALUES (
                :rule_name, :trigger_source, :trigger_condition, :action_type,
                :source_status_id, :source_status_name, :target_status_id, :target_status_name,
                :source_warehouse_code, :source_warehouse_name, :target_warehouse_code, :target_warehouse_name,
                :priority, :auto_create_task, :is_enabled, :sort_order, :note, :created_by, :updated_by
            )
            """
        ),
        default_inventory_flow_rules(conn),
    )


def normalize_inventory_flow_rule_payload(rows: Sequence[Dict[str, Any]] | None) -> List[Dict[str, Any]]:
    valid_actions = {item["value"] for item in inventory_flow_action_options()}
    valid_priorities = {item["value"] for item in inventory_flow_priority_options()}
    valid_sources = {item["value"] for item in inventory_flow_trigger_options()}
    normalized: List[Dict[str, Any]] = []
    seen_names: set[str] = set()
    for index, item in enumerate(rows or []):
        if not isinstance(item, dict):
            continue
        rule_name = str(item.get("rule_name") or "").strip()
        if not rule_name:
            if any(str(item.get(key) or "").strip() for key in ("trigger_source", "action_type", "source_status_id", "target_status_id")):
                raise HTTPException(status_code=400, detail=f"第 {index + 1} 条库存流转规则缺少规则名称")
            continue
        if rule_name in seen_names:
            raise HTTPException(status_code=400, detail=f"库存流转规则重复：{rule_name}")
        seen_names.add(rule_name)
        trigger_source = str(item.get("trigger_source") or "manual").strip().lower() or "manual"
        action_type = str(item.get("action_type") or "status_transition").strip().lower() or "status_transition"
        priority = str(item.get("priority") or "normal").strip().lower() or "normal"
        if trigger_source not in valid_sources:
            raise HTTPException(status_code=400, detail=f"规则 {rule_name} 的触发来源不合法")
        if action_type not in valid_actions:
            raise HTTPException(status_code=400, detail=f"规则 {rule_name} 的动作类型不合法")
        if priority not in valid_priorities:
            raise HTTPException(status_code=400, detail=f"规则 {rule_name} 的优先级不合法")
        normalized.append(
            {
                "id": parse_int_or_default(item.get("id"), 0),
                "rule_name": rule_name,
                "trigger_source": trigger_source,
                "trigger_condition": str(item.get("trigger_condition") or "manual").strip().lower() or "manual",
                "action_type": action_type,
                "source_status_id": str(item.get("source_status_id") or "").strip(),
                "target_status_id": str(item.get("target_status_id") or "").strip(),
                "source_warehouse_code": str(item.get("source_warehouse_code") or "").strip(),
                "target_warehouse_code": str(item.get("target_warehouse_code") or "").strip(),
                "priority": priority,
                "auto_create_task": 1 if parse_bool_or_default(item.get("auto_create_task"), False) else 0,
                "is_enabled": 1 if parse_bool_or_default(item.get("is_enabled"), True) else 0,
                "sort_order": max(0, parse_int_or_default(item.get("sort_order"), (index + 1) * 10)),
                "note": str(item.get("note") or "").strip(),
            }
        )
    return normalized


def normalize_inventory_flow_task_payload(payload: Dict[str, Any] | None) -> Dict[str, Any]:
    item = payload if isinstance(payload, dict) else {}
    valid_actions = {entry["value"] for entry in inventory_flow_action_options()}
    valid_statuses = {entry["value"] for entry in inventory_flow_task_status_options()}
    valid_priorities = {entry["value"] for entry in inventory_flow_priority_options()}
    action_type = str(item.get("action_type") or "status_transition").strip().lower() or "status_transition"
    task_status = str(item.get("task_status") or "draft").strip().lower() or "draft"
    priority = str(item.get("priority") or "normal").strip().lower() or "normal"
    if action_type not in valid_actions:
        raise HTTPException(status_code=400, detail="库存流转任务动作类型不合法")
    if task_status not in valid_statuses:
        raise HTTPException(status_code=400, detail="搴撳瓨娴佽浆浠诲姟鐘舵€佷笉鍚堟硶")
    if priority not in valid_priorities:
        raise HTTPException(status_code=400, detail="搴撳瓨娴佽浆浠诲姟浼樺厛绾т笉鍚堟硶")

    sku_code = str(item.get("sku_code") or "").strip()
    sku_name = str(item.get("sku_name") or "").strip()
    if not sku_code or not sku_name:
        raise HTTPException(status_code=400, detail="库存流转任务需要 SKU 编码和名称")

    request_qty = max(0.0, to_number(item.get("request_qty")) or 0.0)
    confirmed_qty = max(0.0, to_number(item.get("confirmed_qty")) or 0.0)
    if request_qty <= 0:
        raise HTTPException(status_code=400, detail="库存流转任务数量必须大于 0")
    if confirmed_qty > request_qty:
        confirmed_qty = request_qty

    planned_execute_date = parse_date_or_none(str(item.get("planned_execute_date") or "").strip())
    return {
        "id": parse_int_or_default(item.get("id"), 0),
        "task_no": str(item.get("task_no") or "").strip() or generate_inventory_flow_task_no(),
        "source_record_type": str(item.get("source_record_type") or "manual").strip() or "manual",
        "source_record_id": str(item.get("source_record_id") or "").strip(),
        "source_record_no": str(item.get("source_record_no") or "").strip(),
        "trigger_source": str(item.get("trigger_source") or "manual").strip() or "manual",
        "action_type": action_type,
        "task_status": task_status,
        "priority": priority,
        "sku_code": sku_code,
        "sku_name": sku_name,
        "request_qty": round(request_qty, 2),
        "confirmed_qty": round(confirmed_qty, 2),
        "source_status_id": str(item.get("source_status_id") or "").strip(),
        "target_status_id": str(item.get("target_status_id") or "").strip(),
        "source_warehouse_code": str(item.get("source_warehouse_code") or "").strip(),
        "target_warehouse_code": str(item.get("target_warehouse_code") or "").strip(),
        "planned_execute_date": planned_execute_date,
        "reason_text": str(item.get("reason_text") or "").strip(),
        "note": str(item.get("note") or "").strip(),
    }


def serialize_inventory_flow_rule_row(row: Dict[str, Any]) -> Dict[str, Any]:
    item = {key: to_plain(value) for key, value in row.items()}
    item["auto_create_task"] = bool(item.get("auto_create_task"))
    item["is_enabled"] = bool(item.get("is_enabled"))
    item["sort_order"] = int(item.get("sort_order") or 0)
    item["id"] = int(item.get("id") or 0)
    return item


def serialize_inventory_flow_task_row(row: Dict[str, Any]) -> Dict[str, Any]:
    item = {key: to_plain(value) for key, value in row.items()}
    item["id"] = int(item.get("id") or 0)
    item["request_qty"] = float(item.get("request_qty") or 0.0)
    item["confirmed_qty"] = float(item.get("confirmed_qty") or 0.0)
    item["sort_order"] = int(item.get("sort_order") or 0)
    item["completion_rate"] = round(item["confirmed_qty"] / item["request_qty"], 4) if item["request_qty"] > 0 else 0.0
    return item


def resolve_inventory_master_lookups(conn) -> tuple[Dict[str, str], Dict[str, str]]:
    status_lookup = {
        str(row["stock_status_id"] or "").strip(): str(row["stock_status_name"] or "").strip()
        for row in conn.execute(
            text(
                """
                SELECT stock_status_id, stock_status_name
                FROM bi_inventory_status_map
                WHERE is_enabled = 1
                """
            )
        ).mappings().all()
    }
    warehouse_lookup = {
        str(row["warehouse_code"] or "").strip(): str(row["warehouse_name_clean"] or "").strip()
        for row in conn.execute(
            text(
                """
                SELECT warehouse_code, warehouse_name_clean
                FROM bi_inventory_warehouse_map
                WHERE is_enabled = 1
                """
            )
        ).mappings().all()
    }
    return status_lookup, warehouse_lookup


def sync_procurement_inventory_flow_tasks(conn, procurement_item: Dict[str, Any], updated_by: str) -> Dict[str, Any]:
    task_stats = {"created_count": 0, "updated_count": 0, "cancelled_count": 0, "task_nos": []}
    rules = conn.execute(
        text(
            """
            SELECT
                id, rule_name, trigger_source, trigger_condition, action_type,
                source_status_id, source_status_name, target_status_id, target_status_name,
                source_warehouse_code, source_warehouse_name, target_warehouse_code, target_warehouse_name,
                priority, auto_create_task, is_enabled, sort_order, note
            FROM bi_inventory_flow_rule
            WHERE trigger_source = 'procurement_arrival' AND auto_create_task = 1 AND is_enabled = 1
            ORDER BY sort_order, id
            """
        )
    ).mappings().all()
    active_task_nos: List[str] = []

    if procurement_item["status"] != "draft":
        for rule in rules:
            condition = str(rule["trigger_condition"] or "manual")
            qty = 0.0
            suffix = "GEN"
            task_status = "pending"
            if condition == "qualified" and float(procurement_item.get("qualified_qty") or 0.0) > 0:
                qty = float(procurement_item.get("qualified_qty") or 0.0)
                suffix = "Q"
                task_status = "completed" if procurement_item["status"] == "completed" else "pending"
            elif condition == "exception" and float(procurement_item.get("exception_qty") or 0.0) > 0:
                qty = float(procurement_item.get("exception_qty") or 0.0)
                suffix = "E"
                task_status = "blocked" if procurement_item["status"] == "exception" else "pending"
            if qty <= 0:
                continue

            task_no = f"{procurement_item['arrival_no']}-{suffix}"
            active_task_nos.append(task_no)
            task_stats["task_nos"].append(task_no)
            source_warehouse_code = str(procurement_item.get("warehouse_code") or rule["source_warehouse_code"] or "").strip()
            source_warehouse_name = str(procurement_item.get("warehouse_name") or rule["source_warehouse_name"] or source_warehouse_code).strip()
            target_warehouse_code = str(rule["target_warehouse_code"] or source_warehouse_code or "").strip()
            target_warehouse_name = str(rule["target_warehouse_name"] or source_warehouse_name or target_warehouse_code).strip()
            payload = {
                "task_no": task_no,
                "source_record_type": "procurement_arrival",
                "source_record_id": str(procurement_item["id"]),
                "source_record_no": procurement_item["arrival_no"],
                "trigger_source": "procurement_arrival",
                "action_type": str(rule["action_type"] or "status_transition"),
                "task_status": task_status,
                "priority": str(rule["priority"] or "normal"),
                "sku_code": procurement_item["sku_code"],
                "sku_name": procurement_item["sku_name"],
                "request_qty": qty,
                "confirmed_qty": qty if task_status == "completed" else 0.0,
                "source_status_id": str(rule["source_status_id"] or ""),
                "source_status_name": str(rule["source_status_name"] or ""),
                "target_status_id": str(rule["target_status_id"] or ""),
                "target_status_name": str(rule["target_status_name"] or ""),
                "source_warehouse_code": source_warehouse_code,
                "source_warehouse_name": source_warehouse_name,
                "target_warehouse_code": target_warehouse_code,
                "target_warehouse_name": target_warehouse_name,
                "planned_execute_date": procurement_item["arrival_date"],
                "reason_text": procurement_item["exception_reason"] if condition == "exception" else "",
                "note": f"鐢遍噰璐埌璐?{procurement_item['arrival_no']} 鑷姩瑙﹀彂",
                "updated_by": updated_by,
                "sort_order": 100,
            }
            existing = conn.execute(
                text("SELECT id FROM bi_inventory_flow_task WHERE task_no = :task_no LIMIT 1"),
                {"task_no": task_no},
            ).fetchone()
            if existing:
                payload["id"] = int(existing[0])
                conn.execute(
                    text(
                        """
                        UPDATE bi_inventory_flow_task
                        SET
                            source_record_type = :source_record_type,
                            source_record_id = :source_record_id,
                            source_record_no = :source_record_no,
                            trigger_source = :trigger_source,
                            action_type = :action_type,
                            task_status = :task_status,
                            priority = :priority,
                            sku_code = :sku_code,
                            sku_name = :sku_name,
                            request_qty = :request_qty,
                            confirmed_qty = :confirmed_qty,
                            source_status_id = :source_status_id,
                            source_status_name = :source_status_name,
                            target_status_id = :target_status_id,
                            target_status_name = :target_status_name,
                            source_warehouse_code = :source_warehouse_code,
                            source_warehouse_name = :source_warehouse_name,
                            target_warehouse_code = :target_warehouse_code,
                            target_warehouse_name = :target_warehouse_name,
                            planned_execute_date = :planned_execute_date,
                            reason_text = :reason_text,
                            note = :note,
                            updated_by = :updated_by,
                            sort_order = :sort_order
                        WHERE id = :id
                        """
                    ),
                    payload,
                )
                task_stats["updated_count"] += 1
            else:
                payload["created_by"] = updated_by
                conn.execute(
                    text(
                        """
                        INSERT INTO bi_inventory_flow_task(
                            task_no, source_record_type, source_record_id, source_record_no, trigger_source,
                            action_type, task_status, priority, sku_code, sku_name, request_qty, confirmed_qty,
                            source_status_id, source_status_name, target_status_id, target_status_name,
                            source_warehouse_code, source_warehouse_name, target_warehouse_code, target_warehouse_name,
                            planned_execute_date, reason_text, note, created_by, updated_by, sort_order
                        ) VALUES (
                            :task_no, :source_record_type, :source_record_id, :source_record_no, :trigger_source,
                            :action_type, :task_status, :priority, :sku_code, :sku_name, :request_qty, :confirmed_qty,
                            :source_status_id, :source_status_name, :target_status_id, :target_status_name,
                            :source_warehouse_code, :source_warehouse_name, :target_warehouse_code, :target_warehouse_name,
                            :planned_execute_date, :reason_text, :note, :created_by, :updated_by, :sort_order
                        )
                        """
                    ),
                    payload,
                )
                task_stats["created_count"] += 1

    existing_rows = conn.execute(
        text(
            """
            SELECT id, task_no, task_status
            FROM bi_inventory_flow_task
            WHERE source_record_type = 'procurement_arrival' AND source_record_id = :source_record_id
            """
        ),
        {"source_record_id": str(procurement_item["id"])},
    ).mappings().all()
    for row in existing_rows:
        if row["task_no"] in active_task_nos or row["task_status"] == "cancelled":
            continue
        conn.execute(
            text(
                """
                UPDATE bi_inventory_flow_task
                SET task_status = 'cancelled', updated_by = :updated_by
                WHERE id = :id
                """
            ),
            {"id": int(row["id"]), "updated_by": updated_by},
        )
        task_stats["cancelled_count"] += 1
    return task_stats


def task_center_status_options() -> List[Dict[str, str]]:
    return [
        {"value": "open", "label": "待处理"},
        {"value": "in_progress", "label": "处理中"},
        {"value": "blocked", "label": "阻塞"},
        {"value": "completed", "label": "已完成"},
    ]


def task_center_source_module_options() -> List[Dict[str, str]]:
    return [
        {"value": "procurement", "label": "采购到货"},
        {"value": "inventory_flow", "label": "库存流转"},
    ]


def task_center_category_options() -> List[Dict[str, str]]:
    return [
        {"value": "procurement_followup", "label": "到货跟进"},
        {"value": "exception_followup", "label": "异常补偿"},
        {"value": "inventory_execution", "label": "库存执行"},
        {"value": "inventory_exception", "label": "库存阻塞"},
    ]


def task_center_status_label(value: str) -> str:
    lookup = {item["value"]: item["label"] for item in task_center_status_options()}
    return lookup.get(str(value or "").strip(), str(value or "").strip())


def task_center_source_label(value: str) -> str:
    lookup = {item["value"]: item["label"] for item in task_center_source_module_options()}
    return lookup.get(str(value or "").strip(), str(value or "").strip())


def task_center_category_label(value: str) -> str:
    lookup = {item["value"]: item["label"] for item in task_center_category_options()}
    return lookup.get(str(value or "").strip(), str(value or "").strip())


def inventory_flow_priority_label(value: str) -> str:
    lookup = {item["value"]: item["label"] for item in inventory_flow_priority_options()}
    return lookup.get(str(value or "").strip(), str(value or "").strip())


def task_center_status_sort(value: str) -> int:
    ranks = {"blocked": 0, "open": 1, "in_progress": 2, "completed": 3}
    return ranks.get(str(value or "").strip(), 9)


def merge_task_center_status(existing_status: Any, derived_status: str) -> str:
    current = str(existing_status or "").strip().lower()
    if derived_status in {"blocked", "completed"}:
        return derived_status
    if current in {"in_progress", "blocked", "completed"}:
        return current
    return derived_status


def normalize_task_center_payload(payload: Dict[str, Any] | None) -> Dict[str, Any]:
    item = payload if isinstance(payload, dict) else {}
    task_status = str(item.get("task_status") or "open").strip().lower() or "open"
    priority = str(item.get("priority") or "normal").strip().lower() or "normal"
    valid_statuses = {option["value"] for option in task_center_status_options()}
    valid_priorities = {option["value"] for option in inventory_flow_priority_options()}
    if task_status not in valid_statuses:
        raise HTTPException(status_code=400, detail="任务中心状态不合法")
    if priority not in valid_priorities:
        raise HTTPException(status_code=400, detail="任务中心优先级不合法")
    task_id = parse_int_or_default(item.get("id"), 0)
    if task_id <= 0:
        raise HTTPException(status_code=400, detail="任务中心记录缺少有效 ID")
    return {
        "id": task_id,
        "task_status": task_status,
        "priority": priority,
        "owner_name": str(item.get("owner_name") or "").strip(),
        "owner_role": str(item.get("owner_role") or "").strip(),
        "due_date": parse_date_or_none(str(item.get("due_date") or "").strip()),
        "note": str(item.get("note") or "").strip(),
    }


def serialize_task_center_row(row: Dict[str, Any]) -> Dict[str, Any]:
    item = {key: to_plain(value) for key, value in row.items()}
    item["id"] = int(item.get("id") or 0)
    item["sort_order"] = int(item.get("sort_order") or 0)
    due_date = parse_date_or_none(str(item.get("due_date") or "").strip())
    item["is_overdue"] = bool(
        due_date is not None and due_date < date.today() and str(item.get("task_status") or "") != "completed"
    )
    item["source_module_label"] = task_center_source_label(str(item.get("source_module") or ""))
    item["task_category_label"] = task_center_category_label(str(item.get("task_category") or ""))
    item["task_status_label"] = task_center_status_label(str(item.get("task_status") or ""))
    item["priority_label"] = inventory_flow_priority_label(str(item.get("priority") or ""))
    snapshot = json_loads(item.get("source_snapshot_json"), {})
    item["source_snapshot"] = snapshot if isinstance(snapshot, dict) else {}
    item.pop("source_snapshot_json", None)
    return item


def build_task_center_item_from_procurement(procurement_item: Dict[str, Any]) -> Dict[str, Any]:
    source_status = str(procurement_item.get("status") or "draft").strip()
    detail_status = str(procurement_item.get("document_status") or "pending").strip()
    if source_status == "exception" or detail_status == "failed":
        task_status = "blocked"
        task_category = "exception_followup"
        priority = "high"
    elif source_status == "completed" and detail_status in {"generated", "synced"}:
        task_status = "completed"
        task_category = "procurement_followup"
        priority = "low"
    elif detail_status == "pending":
        task_status = "open"
        task_category = "procurement_followup"
        priority = "high" if source_status != "draft" else "normal"
    else:
        task_status = "open"
        task_category = "procurement_followup"
        priority = "normal"

    return {
        "source_module": "procurement",
        "source_type": "procurement_arrival",
        "source_id": str(procurement_item.get("id") or ""),
        "source_no": str(procurement_item.get("arrival_no") or ""),
        "task_title": f"到货跟进 · {procurement_item.get('arrival_no') or procurement_item.get('purchase_order_no') or '未命名单据'}",
        "task_category": task_category,
        "task_status": task_status,
        "priority": priority,
        "owner_name": "",
        "owner_role": "供应链运营",
        "due_date": parse_date_or_none(str(procurement_item.get("arrival_date") or "").strip()),
        "source_status": source_status,
        "source_detail_status": detail_status,
        "summary_text": (
            f"采购单 {procurement_item.get('purchase_order_no') or '--'} / "
            f"{procurement_item.get('supplier_name') or '--'} / "
            f"实到 {procurement_item.get('arrived_qty') or 0}{procurement_item.get('unit') or ''}"
        ),
        "note": "",
        "sort_order": 20 + task_center_status_sort(task_status) * 10,
        "source_snapshot": {
            "purchase_order_no": procurement_item.get("purchase_order_no"),
            "supplier_name": procurement_item.get("supplier_name"),
            "warehouse_name": procurement_item.get("warehouse_name"),
            "channel_name": procurement_item.get("channel_name"),
            "sku_code": procurement_item.get("sku_code"),
            "sku_name": procurement_item.get("sku_name"),
            "expected_qty": procurement_item.get("expected_qty"),
            "arrived_qty": procurement_item.get("arrived_qty"),
            "qualified_qty": procurement_item.get("qualified_qty"),
            "exception_qty": procurement_item.get("exception_qty"),
            "pending_qty": procurement_item.get("pending_qty"),
            "status": source_status,
            "document_status": detail_status,
            "exception_reason": procurement_item.get("exception_reason"),
            "remark": procurement_item.get("remark"),
        },
    }


def build_task_center_item_from_inventory_task(task_item: Dict[str, Any]) -> Dict[str, Any]:
    source_status = str(task_item.get("task_status") or "draft").strip()
    action_type = str(task_item.get("action_type") or "status_transition").strip()
    confirmed_qty = float(task_item.get("confirmed_qty") or 0.0)
    if source_status == "blocked":
        task_status = "blocked"
        task_category = "inventory_exception"
    elif source_status in {"completed", "cancelled"}:
        task_status = "completed"
        task_category = "inventory_execution"
    elif confirmed_qty > 0:
        task_status = "in_progress"
        task_category = "inventory_execution"
    else:
        task_status = "open"
        task_category = "inventory_execution"

    return {
        "source_module": "inventory_flow",
        "source_type": "inventory_flow_task",
        "source_id": str(task_item.get("id") or ""),
        "source_no": str(task_item.get("task_no") or ""),
        "task_title": f"库存执行 · {task_item.get('task_no') or task_item.get('source_record_no') or '未命名任务'}",
        "task_category": task_category,
        "task_status": task_status,
        "priority": str(task_item.get("priority") or "normal"),
        "owner_name": "",
        "owner_role": "仓配执行",
        "due_date": parse_date_or_none(str(task_item.get("planned_execute_date") or "").strip()),
        "source_status": source_status,
        "source_detail_status": action_type,
        "summary_text": (
            f"{task_item.get('sku_name') or '--'} / "
            f"申请 {task_item.get('request_qty') or 0} / "
            f"已确认 {task_item.get('confirmed_qty') or 0}"
        ),
        "note": "",
        "sort_order": 10 + task_center_status_sort(task_status) * 10,
        "source_snapshot": {
            "source_record_no": task_item.get("source_record_no"),
            "trigger_source": task_item.get("trigger_source"),
            "action_type": action_type,
            "sku_code": task_item.get("sku_code"),
            "sku_name": task_item.get("sku_name"),
            "request_qty": task_item.get("request_qty"),
            "confirmed_qty": task_item.get("confirmed_qty"),
            "completion_rate": task_item.get("completion_rate"),
            "source_status_name": task_item.get("source_status_name"),
            "target_status_name": task_item.get("target_status_name"),
            "source_warehouse_name": task_item.get("source_warehouse_name"),
            "target_warehouse_name": task_item.get("target_warehouse_name"),
            "reason_text": task_item.get("reason_text"),
            "note": task_item.get("note"),
        },
    }


def upsert_task_center_item(conn, payload: Dict[str, Any], updated_by: str) -> Dict[str, Any]:
    existing = conn.execute(
        text(
            """
            SELECT id, task_status, owner_name, owner_role, due_date, note
            FROM bi_task_center_item
            WHERE source_module = :source_module AND source_type = :source_type AND source_id = :source_id
            LIMIT 1
            """
        ),
        {
            "source_module": payload["source_module"],
            "source_type": payload["source_type"],
            "source_id": payload["source_id"],
        },
    ).mappings().first()

    if existing:
        task_status = merge_task_center_status(existing.get("task_status"), str(payload.get("task_status") or "open"))
        record = {
            **payload,
            "id": int(existing["id"]),
            "task_status": task_status,
            "owner_name": str(existing.get("owner_name") or payload.get("owner_name") or "").strip(),
            "owner_role": str(existing.get("owner_role") or payload.get("owner_role") or "").strip(),
            "due_date": existing.get("due_date") or payload.get("due_date"),
            "note": str(existing.get("note") or payload.get("note") or "").strip(),
            "source_snapshot_json": json_dumps(payload.get("source_snapshot") or {}),
            "updated_by": updated_by,
        }
        conn.execute(
            text(
                """
                UPDATE bi_task_center_item
                SET
                    source_no = :source_no,
                    task_title = :task_title,
                    task_category = :task_category,
                    task_status = :task_status,
                    priority = :priority,
                    owner_name = :owner_name,
                    owner_role = :owner_role,
                    due_date = :due_date,
                    source_status = :source_status,
                    source_detail_status = :source_detail_status,
                    summary_text = :summary_text,
                    note = :note,
                    source_snapshot_json = :source_snapshot_json,
                    sort_order = :sort_order,
                    updated_by = :updated_by
                WHERE id = :id
                """
            ),
            record,
        )
        return {"id": int(existing["id"]), "created": False}

    record = {
        **payload,
        "owner_name": str(payload.get("owner_name") or "").strip(),
        "owner_role": str(payload.get("owner_role") or "").strip(),
        "note": str(payload.get("note") or "").strip(),
        "source_snapshot_json": json_dumps(payload.get("source_snapshot") or {}),
        "created_by": updated_by,
        "updated_by": updated_by,
    }
    result = conn.execute(
        text(
            """
            INSERT INTO bi_task_center_item(
                source_module, source_type, source_id, source_no, task_title, task_category,
                task_status, priority, owner_name, owner_role, due_date, source_status,
                source_detail_status, summary_text, note, source_snapshot_json,
                sort_order, created_by, updated_by
            ) VALUES (
                :source_module, :source_type, :source_id, :source_no, :task_title, :task_category,
                :task_status, :priority, :owner_name, :owner_role, :due_date, :source_status,
                :source_detail_status, :summary_text, :note, :source_snapshot_json,
                :sort_order, :created_by, :updated_by
            )
            """
        ),
        record,
    )
    return {"id": int(result.lastrowid or 0), "created": True}


def sync_task_center_snapshot(conn, updated_by: str = "system") -> Dict[str, int]:
    procurement_rows = conn.execute(
        text(
            """
            SELECT
                id, arrival_no, purchase_order_no, supplier_name, warehouse_code, warehouse_name,
                channel_code, channel_name, sku_code, sku_name, expected_qty, arrived_qty,
                qualified_qty, exception_qty, unit, arrival_date, status, document_status,
                exception_reason, remark, source_system, created_by, updated_by,
                sort_order, created_at, updated_at
            FROM bi_procurement_arrival
            ORDER BY arrival_date DESC, sort_order, id DESC
            """
        )
    ).mappings().all()
    inventory_rows = conn.execute(
        text(
            """
            SELECT
                id, task_no, source_record_type, source_record_id, source_record_no, trigger_source,
                action_type, task_status, priority, sku_code, sku_name, request_qty, confirmed_qty,
                source_status_id, source_status_name, target_status_id, target_status_name,
                source_warehouse_code, source_warehouse_name, target_warehouse_code, target_warehouse_name,
                planned_execute_date, reason_text, note, created_by, updated_by, sort_order, created_at, updated_at
            FROM bi_inventory_flow_task
            ORDER BY updated_at DESC, id DESC
            """
        )
    ).mappings().all()

    stats = {"procurement_synced": 0, "inventory_synced": 0}
    for row in procurement_rows:
        upsert_task_center_item(conn, build_task_center_item_from_procurement(serialize_procurement_arrival_row(row)), updated_by)
        stats["procurement_synced"] += 1
    for row in inventory_rows:
        upsert_task_center_item(conn, build_task_center_item_from_inventory_task(serialize_inventory_flow_task_row(row)), updated_by)
        stats["inventory_synced"] += 1
    return stats


def reconciliation_case_status_options() -> List[Dict[str, str]]:
    return [
        {"value": "open", "label": "待处理"},
        {"value": "compensating", "label": "补偿中"},
        {"value": "resolved", "label": "已解决"},
        {"value": "ignored", "label": "已忽略"},
    ]


def reconciliation_case_type_options() -> List[Dict[str, str]]:
    return [
        {"value": "document_sync", "label": "单据回写异常"},
        {"value": "inventory_task_missing", "label": "自动任务缺失"},
        {"value": "inventory_task_lag", "label": "任务闭环滞后"},
        {"value": "inventory_task_blocked", "label": "流转任务阻塞"},
        {"value": "inventory_task_overdue", "label": "流转任务逾期"},
    ]


def reconciliation_compensation_action_options() -> List[Dict[str, str]]:
    return [
        {"value": "none", "label": "仅保存编排"},
        {"value": "retry_document_sync", "label": "重试单据编排"},
        {"value": "resync_inventory_tasks", "label": "重建库存任务"},
        {"value": "reopen_inventory_task", "label": "解除任务阻塞"},
        {"value": "mark_resolved", "label": "直接标记解决"},
        {"value": "ignore_case", "label": "忽略当前案例"},
    ]


def reconciliation_case_status_label(value: str) -> str:
    lookup = {item["value"]: item["label"] for item in reconciliation_case_status_options()}
    return lookup.get(str(value or "").strip(), str(value or "").strip())


def reconciliation_case_type_label(value: str) -> str:
    lookup = {item["value"]: item["label"] for item in reconciliation_case_type_options()}
    return lookup.get(str(value or "").strip(), str(value or "").strip())


def reconciliation_compensation_action_label(value: str) -> str:
    lookup = {item["value"]: item["label"] for item in reconciliation_compensation_action_options()}
    return lookup.get(str(value or "").strip(), str(value or "").strip())


def reconciliation_case_status_sort(value: str) -> int:
    ranks = {"open": 0, "compensating": 1, "resolved": 2, "ignored": 3}
    return ranks.get(str(value or "").strip(), 9)


def merge_reconciliation_case_status(existing_status: Any, derived_status: str) -> str:
    current = str(existing_status or "").strip().lower()
    if derived_status == "resolved":
        return "ignored" if current == "ignored" else "resolved"
    if current == "ignored":
        return "ignored"
    if current == "compensating":
        return "compensating"
    return derived_status


def normalize_reconciliation_case_payload(payload: Dict[str, Any] | None) -> Dict[str, Any]:
    item = payload if isinstance(payload, dict) else {}
    case_id = parse_int_or_default(item.get("id"), 0)
    if case_id <= 0:
        raise HTTPException(status_code=400, detail="对账案例缺少有效 ID")
    case_status = str(item.get("case_status") or "open").strip().lower() or "open"
    valid_statuses = {entry["value"] for entry in reconciliation_case_status_options()}
    if case_status not in valid_statuses:
        raise HTTPException(status_code=400, detail="对账案例状态不合法")
    compensation_action = str(item.get("compensation_action") or "").strip().lower()
    valid_actions = {entry["value"] for entry in reconciliation_compensation_action_options()}
    if compensation_action and compensation_action not in valid_actions:
        raise HTTPException(status_code=400, detail="补偿动作不合法")
    return {
        "id": case_id,
        "case_status": case_status,
        "owner_name": str(item.get("owner_name") or "").strip(),
        "owner_role": str(item.get("owner_role") or "").strip(),
        "due_date": parse_date_or_none(str(item.get("due_date") or "").strip()),
        "compensation_action": "" if compensation_action in {"", "none"} else compensation_action,
        "compensation_note": str(item.get("compensation_note") or "").strip(),
    }


def serialize_reconciliation_case_row(row: Dict[str, Any]) -> Dict[str, Any]:
    item = {key: to_plain(value) for key, value in row.items()}
    item["id"] = int(item.get("id") or 0)
    item["sort_order"] = int(item.get("sort_order") or 0)
    due_date = parse_date_or_none(str(item.get("due_date") or "").strip())
    item["is_overdue"] = bool(
        due_date is not None and due_date < date.today() and str(item.get("case_status") or "") not in {"resolved", "ignored"}
    )
    item["source_module_label"] = task_center_source_label(str(item.get("source_module") or ""))
    item["case_type_label"] = reconciliation_case_type_label(str(item.get("case_type") or ""))
    item["case_status_label"] = reconciliation_case_status_label(str(item.get("case_status") or ""))
    item["severity_label"] = inventory_flow_priority_label(str(item.get("severity") or ""))
    item["last_compensation_action_label"] = reconciliation_compensation_action_label(str(item.get("last_compensation_action") or ""))
    expected_snapshot = json_loads(item.get("expected_snapshot_json"), {})
    actual_snapshot = json_loads(item.get("actual_snapshot_json"), {})
    item["expected_snapshot"] = expected_snapshot if isinstance(expected_snapshot, dict) else {}
    item["actual_snapshot"] = actual_snapshot if isinstance(actual_snapshot, dict) else {}
    item.pop("expected_snapshot_json", None)
    item.pop("actual_snapshot_json", None)
    return item


def fetch_procurement_item_by_id(conn, row_id: str | int) -> Dict[str, Any] | None:
    row = conn.execute(
        text(
            """
            SELECT
                id, arrival_no, purchase_order_no, supplier_name, warehouse_code, warehouse_name,
                channel_code, channel_name, sku_code, sku_name, expected_qty, arrived_qty,
                qualified_qty, exception_qty, unit, arrival_date, status, document_status,
                exception_reason, remark, source_system, created_by, updated_by,
                sort_order, created_at, updated_at
            FROM bi_procurement_arrival
            WHERE id = :id
            LIMIT 1
            """
        ),
        {"id": parse_int_or_default(row_id, 0)},
    ).mappings().first()
    return serialize_procurement_arrival_row(row) if row else None


def fetch_inventory_flow_task_by_id(conn, row_id: str | int) -> Dict[str, Any] | None:
    row = conn.execute(
        text(
            """
            SELECT
                id, task_no, source_record_type, source_record_id, source_record_no, trigger_source,
                action_type, task_status, priority, sku_code, sku_name, request_qty, confirmed_qty,
                source_status_id, source_status_name, target_status_id, target_status_name,
                source_warehouse_code, source_warehouse_name, target_warehouse_code, target_warehouse_name,
                planned_execute_date, reason_text, note, created_by, updated_by, sort_order, created_at, updated_at
            FROM bi_inventory_flow_task
            WHERE id = :id
            LIMIT 1
            """
        ),
        {"id": parse_int_or_default(row_id, 0)},
    ).mappings().first()
    return serialize_inventory_flow_task_row(row) if row else None


def build_reconciliation_cases_from_procurement(
    procurement_item: Dict[str, Any],
    related_tasks: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    source_status = str(procurement_item.get("status") or "draft").strip()
    detail_status = str(procurement_item.get("document_status") or "pending").strip()
    if source_status == "draft":
        return []

    arrival_no = str(procurement_item.get("arrival_no") or "")
    expected_task_nos: List[str] = []
    if float(procurement_item.get("qualified_qty") or 0.0) > 0:
        expected_task_nos.append(f"{arrival_no}-Q")
    if float(procurement_item.get("exception_qty") or 0.0) > 0:
        expected_task_nos.append(f"{arrival_no}-E")

    existing_task_map = {
        str(task.get("task_no") or ""): task
        for task in related_tasks
        if str(task.get("task_no") or "").strip()
    }
    related_task_summaries = [
        {
            "task_no": task.get("task_no"),
            "task_status": task.get("task_status"),
            "planned_execute_date": task.get("planned_execute_date"),
            "request_qty": task.get("request_qty"),
        }
        for task in related_tasks[:6]
    ]

    cases: List[Dict[str, Any]] = []
    if detail_status in {"pending", "failed"}:
        cases.append(
            {
                "source_module": "procurement",
                "source_type": "procurement_arrival",
                "source_id": str(procurement_item.get("id") or ""),
                "source_no": arrival_no,
                "case_type": "document_sync",
                "case_title": f"单据回写对账 · {arrival_no or procurement_item.get('purchase_order_no') or '未命名单据'}",
                "case_status": "open",
                "severity": "high" if detail_status == "failed" else "normal",
                "diff_summary": f"到货单仍处于 {detail_status}，需要补偿编排或回写。",
                "owner_name": "",
                "owner_role": "供应链运营",
                "due_date": parse_date_or_none(str(procurement_item.get("arrival_date") or "").strip()),
                "expected_snapshot": {
                    "expected_document_status": "generated / synced",
                    "arrival_no": arrival_no,
                    "purchase_order_no": procurement_item.get("purchase_order_no"),
                    "qualified_qty": procurement_item.get("qualified_qty"),
                    "exception_qty": procurement_item.get("exception_qty"),
                },
                "actual_snapshot": {
                    "status": source_status,
                    "document_status": detail_status,
                    "exception_reason": procurement_item.get("exception_reason"),
                    "remark": procurement_item.get("remark"),
                    "related_tasks": related_task_summaries,
                },
                "last_compensation_action": "",
                "compensation_note": "",
                "sort_order": 10 + reconciliation_case_status_sort("open") * 10,
            }
        )

    missing_task_nos = [task_no for task_no in expected_task_nos if task_no not in existing_task_map]
    if missing_task_nos:
        cases.append(
            {
                "source_module": "procurement",
                "source_type": "procurement_arrival",
                "source_id": str(procurement_item.get("id") or ""),
                "source_no": arrival_no,
                "case_type": "inventory_task_missing",
                "case_title": f"自动任务缺失 · {arrival_no or procurement_item.get('purchase_order_no') or '未命名单据'}",
                "case_status": "open",
                "severity": "high",
                "diff_summary": f"预期任务 {len(expected_task_nos)} 条，当前缺失 {len(missing_task_nos)} 条。",
                "owner_name": "",
                "owner_role": "供应链运营",
                "due_date": parse_date_or_none(str(procurement_item.get("arrival_date") or "").strip()),
                "expected_snapshot": {
                    "expected_task_nos": expected_task_nos,
                    "qualified_qty": procurement_item.get("qualified_qty"),
                    "exception_qty": procurement_item.get("exception_qty"),
                },
                "actual_snapshot": {
                    "missing_task_nos": missing_task_nos,
                    "existing_task_nos": sorted(existing_task_map.keys()),
                    "related_tasks": related_task_summaries,
                },
                "last_compensation_action": "",
                "compensation_note": "",
                "sort_order": 20 + reconciliation_case_status_sort("open") * 10,
            }
        )

    lagging_tasks = [
        task
        for task in related_tasks
        if str(task.get("task_status") or "").strip() not in {"completed", "cancelled"}
    ]
    if source_status == "completed" and lagging_tasks:
        cases.append(
            {
                "source_module": "procurement",
                "source_type": "procurement_arrival",
                "source_id": str(procurement_item.get("id") or ""),
                "source_no": arrival_no,
                "case_type": "inventory_task_lag",
                "case_title": f"任务闭环滞后 · {arrival_no or procurement_item.get('purchase_order_no') or '未命名单据'}",
                "case_status": "open",
                "severity": "high" if any(str(task.get("task_status") or "") == "blocked" for task in lagging_tasks) else "normal",
                "diff_summary": f"到货已完成，但仍有 {len(lagging_tasks)} 条库存任务未闭环。",
                "owner_name": "",
                "owner_role": "仓配执行",
                "due_date": parse_date_or_none(str(procurement_item.get("arrival_date") or "").strip()),
                "expected_snapshot": {
                    "expected_task_status": "completed",
                    "expected_task_nos": expected_task_nos,
                },
                "actual_snapshot": {
                    "lagging_tasks": [
                        {
                            "task_no": task.get("task_no"),
                            "task_status": task.get("task_status"),
                            "planned_execute_date": task.get("planned_execute_date"),
                        }
                        for task in lagging_tasks[:8]
                    ],
                    "document_status": detail_status,
                },
                "last_compensation_action": "",
                "compensation_note": "",
                "sort_order": 30 + reconciliation_case_status_sort("open") * 10,
            }
        )
    return cases


def build_reconciliation_cases_from_inventory_task(task_item: Dict[str, Any]) -> List[Dict[str, Any]]:
    task_status = str(task_item.get("task_status") or "draft").strip()
    planned_execute_date = parse_date_or_none(str(task_item.get("planned_execute_date") or "").strip())
    task_no = str(task_item.get("task_no") or "")
    cases: List[Dict[str, Any]] = []

    if task_status == "blocked":
        cases.append(
            {
                "source_module": "inventory_flow",
                "source_type": "inventory_flow_task",
                "source_id": str(task_item.get("id") or ""),
                "source_no": task_no,
                "case_type": "inventory_task_blocked",
                "case_title": f"流转任务阻塞 · {task_no or task_item.get('source_record_no') or '未命名任务'}",
                "case_status": "open",
                "severity": "high" if str(task_item.get("priority") or "") == "high" else "normal",
                "diff_summary": str(task_item.get("reason_text") or "").strip() or "库存流转任务已阻塞，需要解除或改派。",
                "owner_name": "",
                "owner_role": "仓配执行",
                "due_date": planned_execute_date,
                "expected_snapshot": {
                    "expected_task_status": "pending / completed",
                    "action_type": task_item.get("action_type"),
                    "planned_execute_date": task_item.get("planned_execute_date"),
                },
                "actual_snapshot": {
                    "task_status": task_status,
                    "source_record_no": task_item.get("source_record_no"),
                    "sku_name": task_item.get("sku_name"),
                    "reason_text": task_item.get("reason_text"),
                    "note": task_item.get("note"),
                },
                "last_compensation_action": "",
                "compensation_note": "",
                "sort_order": 40 + reconciliation_case_status_sort("open") * 10,
            }
        )

    if task_status in {"draft", "pending"} and planned_execute_date is not None and planned_execute_date < date.today():
        cases.append(
            {
                "source_module": "inventory_flow",
                "source_type": "inventory_flow_task",
                "source_id": str(task_item.get("id") or ""),
                "source_no": task_no,
                "case_type": "inventory_task_overdue",
                "case_title": f"流转任务逾期 · {task_no or task_item.get('source_record_no') or '未命名任务'}",
                "case_status": "open",
                "severity": "high" if str(task_item.get("priority") or "") == "high" else "normal",
                "diff_summary": f"计划日期 {task_item.get('planned_execute_date') or '--'} 已超期，但任务仍未完成。",
                "owner_name": "",
                "owner_role": "仓配执行",
                "due_date": planned_execute_date,
                "expected_snapshot": {
                    "expected_task_status": "completed",
                    "planned_execute_date": task_item.get("planned_execute_date"),
                },
                "actual_snapshot": {
                    "task_status": task_status,
                    "priority": task_item.get("priority"),
                    "source_record_no": task_item.get("source_record_no"),
                    "sku_name": task_item.get("sku_name"),
                    "request_qty": task_item.get("request_qty"),
                    "confirmed_qty": task_item.get("confirmed_qty"),
                },
                "last_compensation_action": "",
                "compensation_note": "",
                "sort_order": 50 + reconciliation_case_status_sort("open") * 10,
            }
        )

    return cases


def upsert_reconciliation_case(conn, payload: Dict[str, Any], updated_by: str) -> Dict[str, Any]:
    existing = conn.execute(
        text(
            """
            SELECT
                id, case_status, owner_name, owner_role, due_date,
                last_compensation_action, compensation_note, compensated_at, compensated_by
            FROM bi_reconciliation_case
            WHERE source_module = :source_module AND source_type = :source_type
              AND source_id = :source_id AND case_type = :case_type
            LIMIT 1
            """
        ),
        {
            "source_module": payload["source_module"],
            "source_type": payload["source_type"],
            "source_id": payload["source_id"],
            "case_type": payload["case_type"],
        },
    ).mappings().first()

    if existing:
        record = {
            **payload,
            "id": int(existing["id"]),
            "case_status": merge_reconciliation_case_status(existing.get("case_status"), str(payload.get("case_status") or "open")),
            "owner_name": str(existing.get("owner_name") or payload.get("owner_name") or "").strip(),
            "owner_role": str(existing.get("owner_role") or payload.get("owner_role") or "").strip(),
            "due_date": existing.get("due_date") or payload.get("due_date"),
            "last_compensation_action": str(existing.get("last_compensation_action") or payload.get("last_compensation_action") or "").strip(),
            "compensation_note": str(existing.get("compensation_note") or payload.get("compensation_note") or "").strip(),
            "compensated_at": existing.get("compensated_at"),
            "compensated_by": existing.get("compensated_by"),
            "expected_snapshot_json": json_dumps(payload.get("expected_snapshot") or {}),
            "actual_snapshot_json": json_dumps(payload.get("actual_snapshot") or {}),
            "updated_by": updated_by,
        }
        conn.execute(
            text(
                """
                UPDATE bi_reconciliation_case
                SET
                    source_no = :source_no,
                    case_title = :case_title,
                    case_status = :case_status,
                    severity = :severity,
                    diff_summary = :diff_summary,
                    owner_name = :owner_name,
                    owner_role = :owner_role,
                    due_date = :due_date,
                    expected_snapshot_json = :expected_snapshot_json,
                    actual_snapshot_json = :actual_snapshot_json,
                    last_compensation_action = :last_compensation_action,
                    compensation_note = :compensation_note,
                    compensated_at = :compensated_at,
                    compensated_by = :compensated_by,
                    sort_order = :sort_order,
                    updated_by = :updated_by
                WHERE id = :id
                """
            ),
            record,
        )
        return {"id": int(existing["id"]), "created": False}

    record = {
        **payload,
        "owner_name": str(payload.get("owner_name") or "").strip(),
        "owner_role": str(payload.get("owner_role") or "").strip(),
        "last_compensation_action": str(payload.get("last_compensation_action") or "").strip(),
        "compensation_note": str(payload.get("compensation_note") or "").strip(),
        "compensated_at": payload.get("compensated_at"),
        "compensated_by": str(payload.get("compensated_by") or "").strip() or None,
        "expected_snapshot_json": json_dumps(payload.get("expected_snapshot") or {}),
        "actual_snapshot_json": json_dumps(payload.get("actual_snapshot") or {}),
        "created_by": updated_by,
        "updated_by": updated_by,
    }
    result = conn.execute(
        text(
            """
            INSERT INTO bi_reconciliation_case(
                source_module, source_type, source_id, source_no, case_type, case_title,
                case_status, severity, diff_summary, owner_name, owner_role, due_date,
                expected_snapshot_json, actual_snapshot_json, last_compensation_action,
                compensation_note, compensated_at, compensated_by,
                sort_order, created_by, updated_by
            ) VALUES (
                :source_module, :source_type, :source_id, :source_no, :case_type, :case_title,
                :case_status, :severity, :diff_summary, :owner_name, :owner_role, :due_date,
                :expected_snapshot_json, :actual_snapshot_json, :last_compensation_action,
                :compensation_note, :compensated_at, :compensated_by,
                :sort_order, :created_by, :updated_by
            )
            """
        ),
        record,
    )
    return {"id": int(result.lastrowid or 0), "created": True}


def sync_reconciliation_snapshot(conn, updated_by: str = "system") -> Dict[str, int]:
    procurement_rows = conn.execute(
        text(
            """
            SELECT
                id, arrival_no, purchase_order_no, supplier_name, warehouse_code, warehouse_name,
                channel_code, channel_name, sku_code, sku_name, expected_qty, arrived_qty,
                qualified_qty, exception_qty, unit, arrival_date, status, document_status,
                exception_reason, remark, source_system, created_by, updated_by,
                sort_order, created_at, updated_at
            FROM bi_procurement_arrival
            ORDER BY arrival_date DESC, sort_order, id DESC
            """
        )
    ).mappings().all()
    inventory_rows = conn.execute(
        text(
            """
            SELECT
                id, task_no, source_record_type, source_record_id, source_record_no, trigger_source,
                action_type, task_status, priority, sku_code, sku_name, request_qty, confirmed_qty,
                source_status_id, source_status_name, target_status_id, target_status_name,
                source_warehouse_code, source_warehouse_name, target_warehouse_code, target_warehouse_name,
                planned_execute_date, reason_text, note, created_by, updated_by, sort_order, created_at, updated_at
            FROM bi_inventory_flow_task
            ORDER BY updated_at DESC, id DESC
            """
        )
    ).mappings().all()

    inventory_tasks = [serialize_inventory_flow_task_row(row) for row in inventory_rows]
    tasks_by_procurement: Dict[str, List[Dict[str, Any]]] = {}
    for task in inventory_tasks:
        if str(task.get("source_record_type") or "") != "procurement_arrival":
            continue
        source_id = str(task.get("source_record_id") or "").strip()
        tasks_by_procurement.setdefault(source_id, []).append(task)

    active_keys: set[Tuple[str, str, str, str]] = set()
    stats = {"case_count": 0, "resolved_count": 0}
    for row in procurement_rows:
        procurement_item = serialize_procurement_arrival_row(row)
        for case in build_reconciliation_cases_from_procurement(procurement_item, tasks_by_procurement.get(str(procurement_item["id"]), [])):
            upsert_reconciliation_case(conn, case, updated_by)
            active_keys.add((case["source_module"], case["source_type"], case["source_id"], case["case_type"]))
            stats["case_count"] += 1
    for task in inventory_tasks:
        for case in build_reconciliation_cases_from_inventory_task(task):
            upsert_reconciliation_case(conn, case, updated_by)
            active_keys.add((case["source_module"], case["source_type"], case["source_id"], case["case_type"]))
            stats["case_count"] += 1

    existing_rows = conn.execute(
        text(
            """
            SELECT id, source_module, source_type, source_id, case_type, case_status
            FROM bi_reconciliation_case
            """
        )
    ).mappings().all()
    for row in existing_rows:
        key = (
            str(row["source_module"] or ""),
            str(row["source_type"] or ""),
            str(row["source_id"] or ""),
            str(row["case_type"] or ""),
        )
        if key in active_keys or str(row["case_status"] or "") in {"resolved", "ignored"}:
            continue
        conn.execute(
            text(
                """
                UPDATE bi_reconciliation_case
                SET case_status = 'resolved', updated_by = :updated_by
                WHERE id = :id
                """
            ),
            {"id": int(row["id"]), "updated_by": updated_by},
        )
        stats["resolved_count"] += 1
    return stats


def ensure_master_data_seed(conn) -> None:
    ensure_table_columns(
        conn,
        "bi_inventory_warehouse_map",
        {
            "warehouse_code": "VARCHAR(64) NOT NULL DEFAULT ''",
            "warehouse_type": "VARCHAR(32) NOT NULL DEFAULT ''",
            "platform_owner": "VARCHAR(32) NOT NULL DEFAULT ''",
            "city": "VARCHAR(64) NOT NULL DEFAULT ''",
            "is_sellable_warehouse": "TINYINT(1) NOT NULL DEFAULT 0",
            "is_reverse_warehouse": "TINYINT(1) NOT NULL DEFAULT 0",
        },
    )
    ensure_table_columns(
        conn,
        "bi_inventory_status_map",
        {
            "status_group": "VARCHAR(32) NOT NULL DEFAULT ''",
            "can_sell": "TINYINT(1) NOT NULL DEFAULT 0",
            "can_forecast_supply": "TINYINT(1) NOT NULL DEFAULT 0",
            "need_quality_check": "TINYINT(1) NOT NULL DEFAULT 0",
            "next_default_status": "VARCHAR(64) NOT NULL DEFAULT ''",
        },
    )

    conn.execute(
        text(
            """
            UPDATE bi_inventory_warehouse_map
            SET
                warehouse_code = CASE
                    WHEN warehouse_code = '' THEN warehouse_name_clean
                    ELSE warehouse_code
                END,
                warehouse_type = CASE
                    WHEN warehouse_type <> '' THEN warehouse_type
                    WHEN warehouse_name_clean LIKE '%閿€閫€%' OR warehouse_name_clean LIKE '%閫€璐?' THEN 'reverse'
                    WHEN warehouse_name_clean LIKE '%寰呮%' THEN 'qc'
                    WHEN warehouse_name_clean LIKE '%缈绘柊%' THEN 'refurb'
                    WHEN warehouse_name_clean LIKE '%鑹搧%' OR warehouse_name_clean LIKE '%浜戜粨%' THEN 'sellable'
                    ELSE 'general'
                END,
                platform_owner = CASE
                    WHEN platform_owner = '' THEN '鐢ㄥ弸'
                    ELSE platform_owner
                END,
                city = CASE
                    WHEN city <> '' THEN city
                    WHEN source_warehouse_name LIKE '%钀у北%' OR warehouse_name_clean LIKE '%钀у北%' THEN '鏉窞路钀у北'
                    WHEN source_warehouse_name LIKE '%浣欐澀%' OR warehouse_name_clean LIKE '%浣欐澀%' THEN '鏉窞路浣欐澀'
                    ELSE ''
                END,
                is_sellable_warehouse = CASE
                    WHEN is_sellable_warehouse = 1 THEN 1
                    WHEN warehouse_name_clean LIKE '%鑹搧%' OR warehouse_name_clean LIKE '%鍙敭%' THEN 1
                    ELSE 0
                END,
                is_reverse_warehouse = CASE
                    WHEN is_reverse_warehouse = 1 THEN 1
                    WHEN warehouse_name_clean LIKE '%閿€閫€%' OR warehouse_name_clean LIKE '%閫€璐?' THEN 1
                    ELSE 0
                END
            """
        )
    )
    conn.execute(
        text(
            """
            UPDATE bi_inventory_status_map
            SET
                status_group = CASE
                    WHEN status_group <> '' THEN status_group
                    WHEN stock_status_name LIKE '%寰呮%' THEN 'qc'
                    WHEN stock_status_name LIKE '%缈绘柊%' THEN 'refurb'
                    WHEN stock_status_name LIKE '%鍐荤粨%' OR stock_status_name LIKE '%涓嶈壇%' THEN 'exception'
                    WHEN stock_status_name LIKE '%鑹搧%' OR stock_status_name LIKE '%鍙敭%' THEN 'sellable'
                    ELSE 'general'
                END,
                can_sell = CASE
                    WHEN can_sell = 1 THEN 1
                    WHEN stock_status_name LIKE '%鑹搧%' OR stock_status_name LIKE '%鍙敭%' THEN 1
                    ELSE 0
                END,
                can_forecast_supply = CASE
                    WHEN can_forecast_supply = 1 THEN 1
                    WHEN stock_status_name LIKE '%鑹搧%' OR stock_status_name LIKE '%鍙敭%' THEN 1
                    ELSE 0
                END,
                need_quality_check = CASE
                    WHEN need_quality_check = 1 THEN 1
                    WHEN stock_status_name LIKE '%寰呮%' THEN 1
                    ELSE 0
                END,
                next_default_status = CASE
                    WHEN next_default_status <> '' THEN next_default_status
                    WHEN stock_status_name LIKE '%寰呮%' THEN '鑹搧'
                    WHEN stock_status_name LIKE '%缈绘柊瀹屾垚%' THEN '鑹搧'
                    ELSE ''
                END
            """
        )
    )

    if int(conn.execute(text("SELECT COUNT(*) FROM bi_sku_master")).scalar() or 0) == 0:
        if table_exists(conn, "bi_inventory_snapshot_daily"):
            conn.execute(
                text(
                    """
                    INSERT IGNORE INTO bi_sku_master(
                        sku_code, sku_name, sku_type, product_line, model, spec_version,
                        lifecycle_status, owner_dept, sort_order, is_active, created_by, updated_by
                    )
                    SELECT DISTINCT
                        COALESCE(NULLIF(sku_code, ''), NULLIF(material_code, '')) AS sku_code,
                        COALESCE(NULLIF(sku_name, ''), NULLIF(material_name, ''), COALESCE(NULLIF(sku_code, ''), NULLIF(material_code, ''))) AS sku_name,
                        '',
                        '',
                        '',
                        '',
                        'active',
                        '',
                        100,
                        1,
                        'system',
                        'system'
                    FROM bi_inventory_snapshot_daily
                    WHERE COALESCE(NULLIF(sku_code, ''), NULLIF(material_code, '')) IS NOT NULL
                      AND COALESCE(NULLIF(sku_code, ''), NULLIF(material_code, '')) <> ''
                    """
                )
            )
        if table_exists(conn, "bi_material_sales_daily"):
            conn.execute(
                text(
                    """
                    INSERT IGNORE INTO bi_sku_master(
                        sku_code, sku_name, sku_type, product_line, model, spec_version,
                        lifecycle_status, owner_dept, sort_order, is_active, created_by, updated_by
                    )
                    SELECT DISTINCT
                        COALESCE(NULLIF(sku_code, ''), NULLIF(material_code, '')) AS sku_code,
                        COALESCE(NULLIF(sku_name, ''), NULLIF(material_name, ''), COALESCE(NULLIF(sku_code, ''), NULLIF(material_code, ''))) AS sku_name,
                        '',
                        '',
                        '',
                        '',
                        'active',
                        '',
                        100,
                        1,
                        'system',
                        'system'
                    FROM bi_material_sales_daily
                    WHERE COALESCE(NULLIF(sku_code, ''), NULLIF(material_code, '')) IS NOT NULL
                      AND COALESCE(NULLIF(sku_code, ''), NULLIF(material_code, '')) <> ''
                    """
                )
            )

    if int(conn.execute(text("SELECT COUNT(*) FROM bi_channel_shop_master")).scalar() or 0) == 0 and table_exists(conn, "bi_material_sales_daily"):
        conn.execute(
            text(
                """
                INSERT IGNORE INTO bi_channel_shop_master(
                    channel_code, channel_name, shop_name, platform_name,
                    owner_dept, sort_order, is_active, created_by, updated_by
                )
                SELECT DISTINCT
                    CONCAT('ch_', SUBSTRING(MD5(COALESCE(NULLIF(sales_org_name, ''), 'unknown')), 1, 12)) AS channel_code,
                    COALESCE(NULLIF(sales_org_name, ''), '鏈綊绫绘笭閬?) AS channel_name,
                    COALESCE(NULLIF(customer_name, ''), '') AS shop_name,
                    '鐢ㄥ弸閿€鍞?,
                    '',
                    100,
                    1,
                    'system',
                    'system'
                FROM bi_material_sales_daily
                WHERE COALESCE(NULLIF(sales_org_name, ''), NULLIF(customer_name, '')) IS NOT NULL
                  AND COALESCE(NULLIF(sales_org_name, ''), NULLIF(customer_name, '')) <> ''
                """
            )
        )

def metric_label(dataset: str, field: str, agg: str) -> str:
    if agg == "count" and field == "*":
        return AGGREGATION_LABELS["count"]
    field_label = DATASETS[dataset]["fields"].get(field, {}).get("label", field)
    return f"{field_label}{AGGREGATION_LABELS.get(agg, agg)}"


def ordered_keys(values: Sequence[str] | set[str], preferred_order: Sequence[str]) -> List[str]:
    value_set = set(values)
    return [item for item in preferred_order if item in value_set]


def default_dimensions(widget_type: str, dataset: str | None = None) -> List[str]:
    if widget_type in {"metric", "text"}:
        return []
    if dataset == "inventory_turnover":
        return ["month"]
    return ["material_name"]


def supports_series_field(widget_type: str) -> bool:
    return widget_type in {"bar", "stacked_bar", "stacked_hbar", "line"}


def default_series_field(dataset: str, widget_type: str) -> str:
    if not supports_series_field(widget_type):
        return ""
    if dataset == "inventory_cleaning" and widget_type in {"stacked_bar", "stacked_hbar"}:
        return "stock_status_name" if "stock_status_name" in DATASETS[dataset]["fields"] else ""
    return ""


def default_metrics(dataset: str, widget_type: str) -> List[Dict[str, Any]]:
    if dataset == "inventory_turnover":
        return [{"field": "inventory_turnover_days", "agg": "avg", "label": "库存周转天数"}]
    if dataset == "sales":
        return [{"field": "qty", "agg": "sum", "label": "原始数量"}]
    if dataset == "inventory_cleaning":
        return [{"field": "qty", "agg": "sum", "label": "数量"}]
    if dataset == "sales_cleaning":
        if widget_type in {"stacked_bar", "stacked_hbar"}:
            return [
                {"field": "sales_out_xiaoshan", "agg": "sum", "label": "销售出库（萧山云仓）"},
                {"field": "sales_out_yuhang", "agg": "sum", "label": "销售出库（余杭云仓）"},
            ]
        return [{"field": "total_sales_qty", "agg": "sum", "label": "当日总销量"}]
    if widget_type in {"stacked_bar", "stacked_hbar"}:
        return [
            {"field": "current_qty", "agg": "sum", "label": "当前库存"},
            {"field": "available_qty", "agg": "sum", "label": "可用库存"},
        ]
    return [{"field": "current_qty", "agg": "sum", "label": "当前库存"}]


def default_layout() -> Dict[str, Any]:
    return {"x": 0, "y": 0, "w": 12, "h": 5, "span": 1, "height": "normal"}


def height_to_rows(height: str) -> int:
    return {"compact": 4, "normal": 5, "tall": 7}.get(height, 5)


def rows_to_height(rows: int) -> str:
    if rows >= 7:
        return "tall"
    if rows <= 4:
        return "compact"
    return "normal"


def default_widget_config(widget_type: str, dataset: str) -> Dict[str, Any]:
    return {
        "dataset": dataset,
        "dimensions": default_dimensions(widget_type, dataset),
        "series_field": default_series_field(dataset, widget_type),
        "metrics": default_metrics(dataset, widget_type),
        "date_filter": {"mode": "follow_page", "date": "", "start_date": "", "end_date": ""},
        "filters": [],
        "sort": [{"field": "metric_0", "direction": "desc"}],
        "limit": 20,
        "text_content": "请输入说明文本",
    }


def preset_sales_view_widgets() -> List[Dict[str, Any]]:
    dataset = "sales_cleaning"
    return [
        {
            "title": "当日总销量",
            "widget_type": "metric",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": [],
                "metrics": [{"field": "total_sales_qty", "agg": "sum", "label": "当日总销量"}],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 1,
            },
            "layout": {"x": 0, "y": 0, "w": 6, "h": 4},
            "sort_order": 0,
        },
        {
            "title": "当日总退货数量",
            "widget_type": "metric",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": [],
                "metrics": [{"field": "total_return_qty", "agg": "sum", "label": "当日总退货数量"}],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 1,
            },
            "layout": {"x": 6, "y": 0, "w": 6, "h": 4},
            "sort_order": 10,
        },
        {
            "title": "退货拆包出勤人数",
            "widget_type": "metric",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": [],
                "metrics": [{"field": "return_unpack_attendance", "agg": "max", "label": "退货拆包出勤人数"}],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 1,
            },
            "layout": {"x": 12, "y": 0, "w": 6, "h": 4},
            "sort_order": 20,
        },
        {
            "title": "退货拆包人效",
            "widget_type": "metric",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": [],
                "metrics": [{"field": "return_unpack_efficiency", "agg": "sum", "label": "退货拆包人效"}],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 1,
            },
            "layout": {"x": 18, "y": 0, "w": 6, "h": 4},
            "sort_order": 30,
        },
        {
            "title": "当日销售出库分仓 TOP10",
            "widget_type": "stacked_bar",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": ["material_name"],
                "metrics": [
                    {"field": "sales_out_xiaoshan", "agg": "sum", "label": "销售出库（萧山云仓）"},
                    {"field": "sales_out_yuhang", "agg": "sum", "label": "销售出库（余杭云仓）"},
                ],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 10,
            },
            "layout": {"x": 0, "y": 4, "w": 12, "h": 6},
            "sort_order": 100,
        },
        {
            "title": "当日在途拦截分仓 TOP10",
            "widget_type": "stacked_bar",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": ["material_name"],
                "metrics": [
                    {"field": "transit_intercept_xiaoshan", "agg": "sum", "label": "在途拦截（萧山云仓）"},
                    {"field": "transit_intercept_yuhang", "agg": "sum", "label": "在途拦截（余杭云仓）"},
                ],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 10,
            },
            "layout": {"x": 12, "y": 4, "w": 12, "h": 6},
            "sort_order": 110,
        },
        {
            "title": "销退仓退货 TOP10",
            "widget_type": "bar",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": ["material_name"],
                "metrics": [{"field": "sales_return_warehouse", "agg": "sum", "label": "销售退货（销退仓）"}],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 10,
            },
            "layout": {"x": 0, "y": 10, "w": 12, "h": 6},
            "sort_order": 200,
        },
        {
            "title": "当日总销量排行榜",
            "widget_type": "ranking",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": ["material_name"],
                "metrics": [{"field": "total_sales_qty", "agg": "sum", "label": "当日总销量"}],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 10,
            },
            "layout": {"x": 12, "y": 10, "w": 12, "h": 6},
            "sort_order": 210,
        },
        {
            "title": "销售与退货明细",
            "widget_type": "table",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": ["material_code", "material_name"],
                "metrics": [
                    {"field": "sales_out_xiaoshan", "agg": "sum", "label": "销售出库（萧山云仓）"},
                    {"field": "sales_out_yuhang", "agg": "sum", "label": "销售出库（余杭云仓）"},
                    {"field": "sales_return_warehouse", "agg": "sum", "label": "销售退货（销退仓）"},
                    {"field": "total_return_qty", "agg": "sum", "label": "当日总退货数量"},
                    {"field": "total_sales_qty", "agg": "sum", "label": "当日总销量"},
                ],
                "filters": [],
                "sort": [{"field": "metric_4", "direction": "desc"}],
                "limit": 20,
            },
            "layout": {"x": 0, "y": 16, "w": 24, "h": 8},
            "sort_order": 300,
        },
    ]


def ensure_preset_sales_view(conn) -> None:
    row = conn.execute(
        text("SELECT id FROM bi_dashboard_view WHERE name = :name LIMIT 1"),
        {"name": PREFERRED_SALES_VIEW_NAME},
    ).mappings().first()
    if row:
        view_id = int(row["id"])
    else:
        result = conn.execute(
            text(
                """
                INSERT INTO bi_dashboard_view(name, description, global_filters_json)
                VALUES (:name, :description, :global_filters_json)
                """
            ),
            {
                "name": PREFERRED_SALES_VIEW_NAME,
                "description": PREFERRED_SALES_VIEW_DESCRIPTION,
                "global_filters_json": "[]",
            },
        )
        view_id = int(result.lastrowid)

    widget_count = int(
        conn.execute(
            text("SELECT COUNT(*) FROM bi_dashboard_widget WHERE view_id = :view_id"),
            {"view_id": view_id},
        ).scalar()
        or 0
    )
    if widget_count > 0:
        return

    insert_preset_widgets(conn, view_id, preset_sales_view_widgets())


def insert_preset_widgets(conn, view_id: int, items: Sequence[Dict[str, Any]]) -> None:
    for item in items:
        config = normalize_widget_config(item["widget_type"], item["config"])
        layout = normalize_layout(item["layout"])
        conn.execute(
            text(
                """
                INSERT INTO bi_dashboard_widget(
                    view_id, title, widget_type, dataset, config_json, layout_json, sort_order, analysis_text
                ) VALUES (
                    :view_id, :title, :widget_type, :dataset, :config_json, :layout_json, :sort_order, :analysis_text
                )
                """
            ),
            {
                "view_id": view_id,
                "title": item["title"],
                "widget_type": item["widget_type"],
                "dataset": item["dataset"],
                "config_json": json_dumps(config),
                "layout_json": json_dumps(layout),
                "sort_order": int(item["sort_order"]),
                "analysis_text": "",
            },
        )


def insert_missing_preset_widgets(conn, view_id: int, items: Sequence[Dict[str, Any]]) -> None:
    existing_titles = {
        str(row["title"])
        for row in conn.execute(
            text("SELECT title FROM bi_dashboard_widget WHERE view_id = :view_id"),
            {"view_id": view_id},
        ).mappings().all()
    }
    missing_items = [item for item in items if str(item["title"]) not in existing_titles]
    if missing_items:
        insert_preset_widgets(conn, view_id, missing_items)


def preset_inventory_view_widgets() -> List[Dict[str, Any]]:
    dataset = "inventory_cleaning"
    return [
        {
            "title": "当日总库存",
            "widget_type": "metric",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": [],
                "metrics": [{"field": "qty", "agg": "sum", "label": "当日总库存"}],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 1,
            },
            "layout": {"x": 0, "y": 0, "w": 6, "h": 4},
            "sort_order": 0,
        },
        {
            "title": "良品仓库存",
            "widget_type": "metric",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": [],
                "metrics": [{"field": "qty", "agg": "sum", "label": "良品仓库存"}],
                "filters": [{"field": "warehouse_name_clean", "op": "eq", "value": "良品仓"}],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 1,
            },
            "layout": {"x": 6, "y": 0, "w": 6, "h": 4},
            "sort_order": 10,
        },
        {
            "title": "不良品仓库存",
            "widget_type": "metric",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": [],
                "metrics": [{"field": "qty", "agg": "sum", "label": "不良品仓库存"}],
                "filters": [{"field": "warehouse_name_clean", "op": "eq", "value": "不良品仓"}],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 1,
            },
            "layout": {"x": 12, "y": 0, "w": 6, "h": 4},
            "sort_order": 20,
        },
        {
            "title": "销退仓库存",
            "widget_type": "metric",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": [],
                "metrics": [{"field": "qty", "agg": "sum", "label": "销退仓库存"}],
                "filters": [{"field": "warehouse_name_clean", "op": "eq", "value": "销退仓"}],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 1,
            },
            "layout": {"x": 18, "y": 0, "w": 6, "h": 4},
            "sort_order": 30,
        },
        {
            "title": "库存趋势",
            "widget_type": "line",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": ["snapshot_date"],
                "metrics": [{"field": "qty", "agg": "sum", "label": "库存数量"}],
                "filters": [],
                "sort": [{"field": "snapshot_date", "direction": "asc"}],
                "limit": 31,
            },
            "layout": {"x": 0, "y": 4, "w": 12, "h": 6},
            "sort_order": 100,
        },
        {
            "title": "库存状态分布",
            "widget_type": "pie",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": ["stock_status_name"],
                "metrics": [{"field": "qty", "agg": "sum", "label": "库存数量"}],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 12,
            },
            "layout": {"x": 12, "y": 4, "w": 12, "h": 6},
            "sort_order": 110,
        },
        {
            "title": "仓库库存分布",
            "widget_type": "bar",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": ["warehouse_name_clean"],
                "metrics": [{"field": "qty", "agg": "sum", "label": "库存数量"}],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 10,
            },
            "layout": {"x": 0, "y": 10, "w": 12, "h": 6},
            "sort_order": 200,
        },
        {
            "title": "高库存物料 TOP15",
            "widget_type": "ranking",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": ["material_name"],
                "metrics": [{"field": "qty", "agg": "sum", "label": "库存数量"}],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 15,
            },
            "layout": {"x": 12, "y": 10, "w": 12, "h": 6},
            "sort_order": 210,
        },
        {
            "title": "翻新物料库存分布",
            "widget_type": "stacked_hbar",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": ["material_name"],
                "series_field": "stock_status_name",
                "metrics": [{"field": "qty", "agg": "sum", "label": "库存数量"}],
                "filters": [{"field": "stock_status_name", "op": "in", "value": ["翻新良品", "翻新不良品", "不良品", "采购良品"]}],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 12,
            },
            "layout": {"x": 0, "y": 16, "w": 24, "h": 7},
            "sort_order": 250,
        },
        {
            "title": "库存清洗明细",
            "widget_type": "table",
            "dataset": dataset,
            "config": {
                "dataset": dataset,
                "dimensions": ["warehouse_name_clean", "material_code", "material_name", "stock_status_name"],
                "metrics": [{"field": "qty", "agg": "sum", "label": "库存数量"}],
                "filters": [],
                "sort": [{"field": "metric_0", "direction": "desc"}],
                "limit": 30,
            },
            "layout": {"x": 0, "y": 23, "w": 24, "h": 8},
            "sort_order": 300,
        },
    ]


def ensure_preset_inventory_view(conn) -> None:
    row = conn.execute(
        text("SELECT id FROM bi_dashboard_view WHERE name = :name LIMIT 1"),
        {"name": PREFERRED_INVENTORY_VIEW_NAME},
    ).mappings().first()
    if row:
        view_id = int(row["id"])
    else:
        result = conn.execute(
            text(
                """
                INSERT INTO bi_dashboard_view(name, description, global_filters_json)
                VALUES (:name, :description, :global_filters_json)
                """
            ),
            {
                "name": PREFERRED_INVENTORY_VIEW_NAME,
                "description": PREFERRED_INVENTORY_VIEW_DESCRIPTION,
                "global_filters_json": "[]",
            },
        )
        view_id = int(result.lastrowid)

    insert_missing_preset_widgets(conn, view_id, preset_inventory_view_widgets())


def normalize_layout(raw_layout: Dict[str, Any] | None) -> Dict[str, Any]:
    layout = dict(raw_layout or {})
    try:
        x = int(layout.get("x", 0) or 0)
    except Exception:
        x = 0
    try:
        y = int(layout.get("y", 0) or 0)
    except Exception:
        y = 0
    try:
        span = int(layout.get("span", 1) or 1)
    except Exception:
        span = 1
    height = str(layout.get("height", "normal") or "normal")
    if span not in LAYOUT_SPANS:
        span = 1
    if height not in LAYOUT_HEIGHTS:
        height = "normal"
    try:
        width = int(layout.get("w", GRID_COLUMNS if span == 2 else GRID_COLUMNS // 2) or (GRID_COLUMNS if span == 2 else GRID_COLUMNS // 2))
    except Exception:
        width = GRID_COLUMNS if span == 2 else GRID_COLUMNS // 2
    try:
        rows = int(layout.get("h", height_to_rows(height)) or height_to_rows(height))
    except Exception:
        rows = height_to_rows(height)
    width = max(GRID_MIN_WIDTH, min(GRID_MAX_WIDTH, width))
    rows = max(GRID_MIN_HEIGHT, min(GRID_MAX_HEIGHT, rows))
    x = max(0, min(GRID_COLUMNS - width, x))
    y = max(0, y)
    span = 2 if width > (GRID_COLUMNS // 2) else 1
    height = rows_to_height(rows)
    return {"x": x, "y": y, "w": width, "h": rows, "span": span, "height": height}


def layout_sort_value(layout: Dict[str, Any]) -> int:
    normalized = normalize_layout(layout)
    return (normalized["y"] * 1000) + (normalized["x"] * 10)


def layouts_collide(left: Dict[str, Any], right: Dict[str, Any]) -> bool:
    return (
        left["x"] < right["x"] + right["w"]
        and left["x"] + left["w"] > right["x"]
        and left["y"] < right["y"] + right["h"]
        and left["y"] + left["h"] > right["y"]
    )


def normalize_template_widgets(raw_widgets: Sequence[Dict[str, Any]] | None) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    for index, item in enumerate(raw_widgets or []):
        if not isinstance(item, dict):
            continue
        normalized.append(
            {
                "title": str(item.get("title") or f"鍗＄墖 {index + 1}"),
                "layout": normalize_layout(item.get("layout")),
            }
        )
    return sorted(normalized, key=lambda item: layout_sort_value(item["layout"]))


def build_layout_template_payload(widgets: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    normalized_widgets = normalize_template_widgets(
        [{"title": widget.get("title"), "layout": widget.get("layout")} for widget in widgets]
    )
    return {"grid_columns": GRID_COLUMNS, "widgets": normalized_widgets}


def layout_template_from_row(row: Dict[str, Any], include_layout: bool = False) -> Dict[str, Any]:
    payload = json_loads(row.get("layout_json"), {})
    template_widgets = normalize_template_widgets(payload.get("widgets") if isinstance(payload, dict) else [])
    max_rows = max((item["layout"]["y"] + item["layout"]["h"] for item in template_widgets), default=0)
    result = {
        "id": int(row["id"]),
        "name": row["name"],
        "description": row["description"],
        "widget_count": len(template_widgets),
        "max_rows": max_rows,
        "created_at": to_plain(row.get("created_at")),
        "updated_at": to_plain(row.get("updated_at")),
    }
    if include_layout:
        result["layout_payload"] = {"grid_columns": GRID_COLUMNS, "widgets": template_widgets}
    return result


def normalize_widget_filters(dataset: str, raw_filters: Sequence[Dict[str, Any]] | None) -> List[Dict[str, Any]]:
    fields = DATASETS[dataset]["fields"]
    normalized: List[Dict[str, Any]] = []
    for item in raw_filters or []:
        if not isinstance(item, dict):
            continue
        field = str(item.get("field") or "")
        operator = str(item.get("op") or "eq").lower()
        value = item.get("value")
        if field not in fields or not fields[field].get("filterable", False):
            continue
        if operator not in FILTER_OPERATORS:
            operator = "eq"
        if operator in {"in", "between"} and isinstance(value, str):
            value = [part.strip() for part in value.split(",") if part.strip()]
        normalized.append({"field": field, "op": operator, "value": value})
    return normalized


def normalize_widget_date_filter(dataset: str, raw_filter: Dict[str, Any] | None) -> Dict[str, Any]:
    raw = raw_filter if isinstance(raw_filter, dict) else {}
    mode = str(raw.get("mode") or "follow_page").lower()
    if mode not in {"follow_page", "single", "range", "all"}:
        mode = "follow_page"

    date_value = parse_date_or_none(str(raw.get("date") or ""))
    start_date = parse_date_or_none(str(raw.get("start_date") or ""))
    end_date = parse_date_or_none(str(raw.get("end_date") or ""))

    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date

    if mode == "single" and date_value is None:
        mode = "follow_page"
    if mode == "range" and start_date is None and end_date is None:
        mode = "follow_page"

    return {
        "mode": mode,
        "date": date_value.isoformat() if date_value else "",
        "start_date": start_date.isoformat() if start_date else "",
        "end_date": end_date.isoformat() if end_date else "",
        "date_col": DATASETS[dataset]["date_col"],
    }


def normalize_global_filters(raw_filters: Sequence[Dict[str, Any]] | None) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    for item in raw_filters or []:
        if not isinstance(item, dict):
            continue
        dataset = str(item.get("dataset") or "")
        if dataset not in DATASETS:
            continue
        widget_filters = normalize_widget_filters(dataset, [item])
        if not widget_filters:
            continue
        normalized.append({"dataset": dataset, **widget_filters[0]})
    return normalized


def normalize_widget_config(widget_type: str, payload: Dict[str, Any] | None) -> Dict[str, Any]:
    raw = dict(payload or {})
    dataset = str(raw.get("dataset") or "inventory")
    if dataset not in DATASETS:
        dataset = "inventory"
    fields = DATASETS[dataset]["fields"]
    config = {
        "dataset": dataset,
        "dimensions": [],
        "series_field": "",
        "metrics": [],
        "date_filter": normalize_widget_date_filter(dataset, raw.get("date_filter")),
        "filters": normalize_widget_filters(dataset, raw.get("filters")),
        "sort": [],
        "limit": 20,
        "text_content": str(raw.get("text_content") or "请输入说明文本"),
    }

    for dim in raw.get("dimensions") or []:
        if dim in fields and fields[dim].get("groupable", False):
            config["dimensions"].append(dim)
    if widget_type not in {"metric", "text"} and not config["dimensions"]:
        config["dimensions"] = default_dimensions(widget_type, dataset)

    series_field = str(raw.get("series_field") or "").strip()
    if (
        supports_series_field(widget_type)
        and series_field in fields
        and fields[series_field].get("groupable", False)
        and series_field not in config["dimensions"]
    ):
        config["series_field"] = series_field
    elif supports_series_field(widget_type):
        config["series_field"] = default_series_field(dataset, widget_type)

    for item in raw.get("metrics") or []:
        if not isinstance(item, dict):
            continue
        field = item.get("field", "*")
        agg = str(item.get("agg") or "sum").lower()
        if agg not in AGGREGATIONS:
            continue
        if agg == "count":
            if field not in ("*", "", None) and field not in fields:
                continue
            field = "*" if field in ("", None) else field
        elif field not in fields or not fields[field].get("numeric", False):
            continue
        config["metrics"].append(
            {
                "field": field,
                "agg": agg,
                "label": str(item.get("label") or metric_label(dataset, field, agg)),
            }
        )
    if widget_type != "text" and not config["metrics"]:
        config["metrics"] = default_metrics(dataset, widget_type)

    for item in raw.get("sort") or []:
        if not isinstance(item, dict):
            continue
        field = str(item.get("field") or "")
        direction = str(item.get("direction") or "asc").lower()
        if field and direction in SORT_DIRECTIONS:
            config["sort"].append({"field": field, "direction": direction})

    try:
        config["limit"] = max(1, min(500, int(raw.get("limit") or 20)))
    except Exception:
        config["limit"] = 20
    return config


def get_engine():
    global engine
    if engine is not None:
        return engine
    raw = yaml.safe_load(yonyou_sync_config_path.read_text(encoding="utf-8")) or {}
    db_url = raw.get("database", {}).get("url", "")
    if not db_url:
        raise RuntimeError("缂哄皯 config/yonyou_inventory_sync.yaml 涓殑 database.url 閰嶇疆")
    engine = create_engine(quote_mysql_url(db_url), pool_pre_ping=True, future=True)
    return engine


def ensure_schema() -> None:
    global schema_ready
    if schema_ready:
        return
    current_engine = get_engine()
    ensure_sales_processing_schema(current_engine)
    ensure_inventory_processing_schema(current_engine)
    ensure_forecast_alert_schema(current_engine)
    with current_engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_dashboard_view (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    name VARCHAR(128) NOT NULL,
                    description VARCHAR(512) NOT NULL DEFAULT '',
                    global_filters_json LONGTEXT NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_dashboard_widget (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    view_id BIGINT NOT NULL,
                    title VARCHAR(128) NOT NULL,
                    widget_type VARCHAR(32) NOT NULL,
                    dataset VARCHAR(32) NOT NULL DEFAULT 'inventory',
                    config_json LONGTEXT NOT NULL,
                    layout_json LONGTEXT NULL,
                    sort_order INT NOT NULL DEFAULT 100,
                    analysis_text LONGTEXT NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    INDEX idx_bi_dash_widget_view_sort (view_id, sort_order, id),
                    CONSTRAINT fk_bi_dash_widget_view FOREIGN KEY (view_id)
                        REFERENCES bi_dashboard_view(id) ON DELETE CASCADE
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_dashboard_layout_template (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    name VARCHAR(128) NOT NULL,
                    description VARCHAR(512) NOT NULL DEFAULT '',
                    layout_json LONGTEXT NOT NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_raw_sync_schedule_config (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    schedule_key VARCHAR(64) NOT NULL,
                    is_enabled TINYINT(1) NOT NULL DEFAULT 0,
                    mode VARCHAR(16) NOT NULL DEFAULT 'all',
                    cron_expr VARCHAR(64) NOT NULL DEFAULT '',
                    sales_days_behind INT NOT NULL DEFAULT 1,
                    sales_window_days INT NOT NULL DEFAULT 1,
                    snapshot_days_behind INT NOT NULL DEFAULT 0,
                    last_run_started_at DATETIME NULL,
                    last_run_finished_at DATETIME NULL,
                    last_run_status VARCHAR(32) NOT NULL DEFAULT 'idle',
                    last_run_message VARCHAR(1024) NOT NULL DEFAULT '',
                    last_trigger VARCHAR(32) NOT NULL DEFAULT '',
                    last_result_json LONGTEXT NULL,
                    updated_by VARCHAR(64) NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_bi_raw_sync_schedule_key (schedule_key)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_refurb_production_daily (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    biz_date DATE NOT NULL,
                    refurb_category VARCHAR(128) NOT NULL,
                    material_name VARCHAR(255) NOT NULL,
                    feeding_qty DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    total_work_hours DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    plan_qty DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    quality_defect_qty DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    production_good_qty DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    production_bad_qty DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    final_good_qty DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    non_refurbishable_rate DECIMAL(18, 6) NOT NULL DEFAULT 0,
                    quality_reject_rate DECIMAL(18, 6) NOT NULL DEFAULT 0,
                    plan_achievement_rate DECIMAL(18, 6) NOT NULL DEFAULT 0,
                    refurb_efficiency DECIMAL(18, 6) NOT NULL DEFAULT 0,
                    updated_by VARCHAR(64) NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_bi_refurb_daily (biz_date, refurb_category, material_name),
                    INDEX idx_bi_refurb_date (biz_date),
                    INDEX idx_bi_refurb_category (refurb_category),
                    INDEX idx_bi_refurb_material (material_name)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_data_agent_report (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    report_type VARCHAR(16) NOT NULL,
                    period_start DATE NOT NULL,
                    period_end DATE NOT NULL,
                    period_label VARCHAR(64) NOT NULL,
                    trigger_mode VARCHAR(16) NOT NULL DEFAULT 'manual',
                    title VARCHAR(255) NOT NULL,
                    summary_json LONGTEXT NULL,
                    report_content LONGTEXT NOT NULL,
                    generated_by VARCHAR(32) NOT NULL DEFAULT 'fallback',
                    status VARCHAR(16) NOT NULL DEFAULT 'success',
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_bi_data_agent_report_period (report_type, period_start, period_end),
                    INDEX idx_bi_data_agent_report_created (created_at)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_audit_log (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    module_key VARCHAR(64) NOT NULL,
                    module_name VARCHAR(64) NOT NULL,
                    action_key VARCHAR(64) NOT NULL,
                    action_name VARCHAR(128) NOT NULL,
                    target_type VARCHAR(64) NOT NULL DEFAULT '',
                    target_id VARCHAR(128) NOT NULL DEFAULT '',
                    target_name VARCHAR(255) NOT NULL DEFAULT '',
                    result_status VARCHAR(32) NOT NULL DEFAULT 'success',
                    detail_summary VARCHAR(255) NOT NULL DEFAULT '',
                    detail_json LONGTEXT NULL,
                    source_path VARCHAR(255) NOT NULL DEFAULT '',
                    source_method VARCHAR(16) NOT NULL DEFAULT '',
                    triggered_by VARCHAR(64) NULL,
                    affected_count INT NOT NULL DEFAULT 0,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_bi_audit_log_created (created_at, id),
                    INDEX idx_bi_audit_log_module (module_key, created_at),
                    INDEX idx_bi_audit_log_status (result_status, created_at),
                    INDEX idx_bi_audit_log_actor (triggered_by, created_at),
                    INDEX idx_bi_audit_log_target (target_type, target_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_procurement_arrival (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    arrival_no VARCHAR(64) NOT NULL,
                    purchase_order_no VARCHAR(64) NOT NULL,
                    supplier_name VARCHAR(128) NOT NULL,
                    warehouse_code VARCHAR(64) NOT NULL,
                    warehouse_name VARCHAR(128) NOT NULL DEFAULT '',
                    channel_code VARCHAR(64) NOT NULL DEFAULT '',
                    channel_name VARCHAR(128) NOT NULL DEFAULT '',
                    sku_code VARCHAR(64) NOT NULL,
                    sku_name VARCHAR(255) NOT NULL,
                    expected_qty DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    arrived_qty DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    qualified_qty DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    exception_qty DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    unit VARCHAR(16) NOT NULL DEFAULT '?',
                    arrival_date DATE NOT NULL,
                    status VARCHAR(32) NOT NULL DEFAULT 'draft',
                    document_status VARCHAR(32) NOT NULL DEFAULT 'pending',
                    exception_reason VARCHAR(255) NOT NULL DEFAULT '',
                    remark LONGTEXT NULL,
                    source_system VARCHAR(32) NOT NULL DEFAULT 'manual',
                    created_by VARCHAR(64) NULL,
                    updated_by VARCHAR(64) NULL,
                    sort_order INT NOT NULL DEFAULT 100,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_bi_procurement_arrival_no (arrival_no),
                    INDEX idx_bi_procurement_po (purchase_order_no, id),
                    INDEX idx_bi_procurement_status (status, document_status, arrival_date),
                    INDEX idx_bi_procurement_supplier (supplier_name, arrival_date),
                    INDEX idx_bi_procurement_warehouse (warehouse_code, arrival_date)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_inventory_flow_rule (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    rule_name VARCHAR(128) NOT NULL,
                    trigger_source VARCHAR(32) NOT NULL DEFAULT 'manual',
                    trigger_condition VARCHAR(32) NOT NULL DEFAULT 'manual',
                    action_type VARCHAR(32) NOT NULL DEFAULT 'status_transition',
                    source_status_id VARCHAR(64) NOT NULL DEFAULT '',
                    source_status_name VARCHAR(128) NOT NULL DEFAULT '',
                    target_status_id VARCHAR(64) NOT NULL DEFAULT '',
                    target_status_name VARCHAR(128) NOT NULL DEFAULT '',
                    source_warehouse_code VARCHAR(64) NOT NULL DEFAULT '',
                    source_warehouse_name VARCHAR(128) NOT NULL DEFAULT '',
                    target_warehouse_code VARCHAR(64) NOT NULL DEFAULT '',
                    target_warehouse_name VARCHAR(128) NOT NULL DEFAULT '',
                    priority VARCHAR(16) NOT NULL DEFAULT 'normal',
                    auto_create_task TINYINT(1) NOT NULL DEFAULT 0,
                    is_enabled TINYINT(1) NOT NULL DEFAULT 1,
                    sort_order INT NOT NULL DEFAULT 100,
                    note VARCHAR(255) NOT NULL DEFAULT '',
                    created_by VARCHAR(64) NULL,
                    updated_by VARCHAR(64) NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_bi_inventory_flow_rule_name (rule_name),
                    INDEX idx_bi_inventory_flow_rule_sort (sort_order, id),
                    INDEX idx_bi_inventory_flow_rule_source (trigger_source, is_enabled)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_inventory_flow_task (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    task_no VARCHAR(64) NOT NULL,
                    source_record_type VARCHAR(32) NOT NULL DEFAULT 'manual',
                    source_record_id VARCHAR(64) NOT NULL DEFAULT '',
                    source_record_no VARCHAR(64) NOT NULL DEFAULT '',
                    trigger_source VARCHAR(32) NOT NULL DEFAULT 'manual',
                    action_type VARCHAR(32) NOT NULL DEFAULT 'status_transition',
                    task_status VARCHAR(32) NOT NULL DEFAULT 'draft',
                    priority VARCHAR(16) NOT NULL DEFAULT 'normal',
                    sku_code VARCHAR(64) NOT NULL,
                    sku_name VARCHAR(255) NOT NULL,
                    request_qty DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    confirmed_qty DECIMAL(18, 2) NOT NULL DEFAULT 0,
                    source_status_id VARCHAR(64) NOT NULL DEFAULT '',
                    source_status_name VARCHAR(128) NOT NULL DEFAULT '',
                    target_status_id VARCHAR(64) NOT NULL DEFAULT '',
                    target_status_name VARCHAR(128) NOT NULL DEFAULT '',
                    source_warehouse_code VARCHAR(64) NOT NULL DEFAULT '',
                    source_warehouse_name VARCHAR(128) NOT NULL DEFAULT '',
                    target_warehouse_code VARCHAR(64) NOT NULL DEFAULT '',
                    target_warehouse_name VARCHAR(128) NOT NULL DEFAULT '',
                    planned_execute_date DATE NULL,
                    reason_text VARCHAR(255) NOT NULL DEFAULT '',
                    note VARCHAR(255) NOT NULL DEFAULT '',
                    created_by VARCHAR(64) NULL,
                    updated_by VARCHAR(64) NULL,
                    sort_order INT NOT NULL DEFAULT 100,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_bi_inventory_flow_task_no (task_no),
                    INDEX idx_bi_inventory_flow_task_status (task_status, planned_execute_date, id),
                    INDEX idx_bi_inventory_flow_task_source (source_record_type, source_record_id),
                    INDEX idx_bi_inventory_flow_task_action (action_type, priority)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_task_center_item (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    source_module VARCHAR(32) NOT NULL,
                    source_type VARCHAR(32) NOT NULL,
                    source_id VARCHAR(64) NOT NULL,
                    source_no VARCHAR(64) NOT NULL DEFAULT '',
                    task_title VARCHAR(255) NOT NULL,
                    task_category VARCHAR(32) NOT NULL DEFAULT 'procurement_followup',
                    task_status VARCHAR(32) NOT NULL DEFAULT 'open',
                    priority VARCHAR(16) NOT NULL DEFAULT 'normal',
                    owner_name VARCHAR(64) NOT NULL DEFAULT '',
                    owner_role VARCHAR(64) NOT NULL DEFAULT '',
                    due_date DATE NULL,
                    source_status VARCHAR(32) NOT NULL DEFAULT '',
                    source_detail_status VARCHAR(32) NOT NULL DEFAULT '',
                    summary_text VARCHAR(255) NOT NULL DEFAULT '',
                    note LONGTEXT NULL,
                    source_snapshot_json LONGTEXT NULL,
                    sort_order INT NOT NULL DEFAULT 100,
                    created_by VARCHAR(64) NULL,
                    updated_by VARCHAR(64) NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_bi_task_center_source (source_module, source_type, source_id),
                    INDEX idx_bi_task_center_status (task_status, due_date, priority),
                    INDEX idx_bi_task_center_source (source_module, task_category),
                    INDEX idx_bi_task_center_sort (sort_order, updated_at, id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_reconciliation_case (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    source_module VARCHAR(32) NOT NULL,
                    source_type VARCHAR(32) NOT NULL,
                    source_id VARCHAR(64) NOT NULL,
                    source_no VARCHAR(64) NOT NULL DEFAULT '',
                    case_type VARCHAR(32) NOT NULL,
                    case_title VARCHAR(255) NOT NULL,
                    case_status VARCHAR(32) NOT NULL DEFAULT 'open',
                    severity VARCHAR(16) NOT NULL DEFAULT 'normal',
                    diff_summary VARCHAR(255) NOT NULL DEFAULT '',
                    owner_name VARCHAR(64) NOT NULL DEFAULT '',
                    owner_role VARCHAR(64) NOT NULL DEFAULT '',
                    due_date DATE NULL,
                    expected_snapshot_json LONGTEXT NULL,
                    actual_snapshot_json LONGTEXT NULL,
                    last_compensation_action VARCHAR(32) NOT NULL DEFAULT '',
                    compensation_note LONGTEXT NULL,
                    compensated_at DATETIME NULL,
                    compensated_by VARCHAR(64) NULL,
                    sort_order INT NOT NULL DEFAULT 100,
                    created_by VARCHAR(64) NULL,
                    updated_by VARCHAR(64) NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_bi_reconciliation_source (source_module, source_type, source_id, case_type),
                    INDEX idx_bi_reconciliation_status (case_status, severity, due_date),
                    INDEX idx_bi_reconciliation_source (source_module, case_type),
                    INDEX idx_bi_reconciliation_sort (sort_order, updated_at, id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_metric_dictionary (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    metric_key VARCHAR(64) NOT NULL,
                    metric_name VARCHAR(128) NOT NULL,
                    business_domain VARCHAR(64) NOT NULL DEFAULT '????',
                    owner_role VARCHAR(64) NOT NULL DEFAULT '',
                    definition_text LONGTEXT NOT NULL,
                    formula_text LONGTEXT NOT NULL,
                    source_table VARCHAR(255) NOT NULL DEFAULT '',
                    source_fields VARCHAR(512) NOT NULL DEFAULT '',
                    dimension_notes VARCHAR(512) NOT NULL DEFAULT '',
                    version_tag VARCHAR(32) NOT NULL DEFAULT 'v1',
                    effective_date DATE NULL,
                    sort_order INT NOT NULL DEFAULT 100,
                    is_enabled TINYINT(1) NOT NULL DEFAULT 1,
                    created_by VARCHAR(64) NULL,
                    updated_by VARCHAR(64) NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_bi_metric_dictionary_key (metric_key),
                    INDEX idx_bi_metric_dictionary_sort (sort_order, id),
                    INDEX idx_bi_metric_dictionary_domain (business_domain),
                    INDEX idx_bi_metric_dictionary_enabled (is_enabled)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_sku_master (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    sku_code VARCHAR(64) NOT NULL,
                    sku_name VARCHAR(255) NOT NULL,
                    sku_type VARCHAR(32) NOT NULL DEFAULT '',
                    product_line VARCHAR(64) NOT NULL DEFAULT '',
                    model VARCHAR(64) NOT NULL DEFAULT '',
                    spec_version VARCHAR(64) NOT NULL DEFAULT '',
                    lifecycle_status VARCHAR(32) NOT NULL DEFAULT 'active',
                    owner_dept VARCHAR(64) NOT NULL DEFAULT '',
                    sort_order INT NOT NULL DEFAULT 100,
                    is_active TINYINT(1) NOT NULL DEFAULT 1,
                    created_by VARCHAR(64) NULL,
                    updated_by VARCHAR(64) NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_bi_sku_master_code (sku_code),
                    INDEX idx_bi_sku_master_sort (sort_order, id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS bi_channel_shop_master (
                    id BIGINT PRIMARY KEY AUTO_INCREMENT,
                    channel_code VARCHAR(64) NOT NULL,
                    channel_name VARCHAR(128) NOT NULL,
                    shop_name VARCHAR(128) NOT NULL DEFAULT '',
                    platform_name VARCHAR(64) NOT NULL DEFAULT '',
                    owner_dept VARCHAR(64) NOT NULL DEFAULT '',
                    sort_order INT NOT NULL DEFAULT 100,
                    is_active TINYINT(1) NOT NULL DEFAULT 1,
                    created_by VARCHAR(64) NULL,
                    updated_by VARCHAR(64) NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_bi_channel_shop_pair (channel_code, shop_name),
                    INDEX idx_bi_channel_shop_sort (sort_order, id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """
            )
        )
        columns = {row[0] for row in conn.execute(text("SHOW COLUMNS FROM bi_dashboard_view")).fetchall()}
        if "global_filters_json" not in columns:
            conn.execute(text("ALTER TABLE bi_dashboard_view ADD COLUMN global_filters_json LONGTEXT NULL"))
        ensure_sync_schedule_seed(conn)
        ensure_metric_dictionary_seed(conn)
        ensure_master_data_seed(conn)
        ensure_procurement_arrival_seed(conn)
        ensure_inventory_flow_seed(conn)
        count = int(conn.execute(text("SELECT COUNT(*) FROM bi_dashboard_view")).scalar() or 0)
        if count == 0:
            conn.execute(
                text(
                    """
                    INSERT INTO bi_dashboard_view(name, description, global_filters_json)
                    VALUES (:name, :description, :global_filters_json)
                    """
                ),
                {
                    "name": "榛樿鐪嬫澘",
                    "description": "搴撳瓨涓庨攢鍞殑鏃ュ父缁忚惀姒傝",
                    "global_filters_json": "[]",
                },
            )
        ensure_preset_sales_view(conn)
        ensure_preset_inventory_view(conn)
    schema_ready = True


def load_view(conn, view_id: int) -> Dict[str, Any]:
    row = conn.execute(
        text(
            """
            SELECT id, name, description, global_filters_json, created_at, updated_at
            FROM bi_dashboard_view
            WHERE id = :id
            """
        ),
        {"id": view_id},
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail=f"鐪嬫澘涓嶅瓨鍦細{view_id}")
    global_filters = normalize_global_filters(json_loads(row.get("global_filters_json"), []))
    return {
        "id": int(row["id"]),
        "name": row["name"],
        "description": row["description"],
        "global_filters": global_filters,
        "created_at": to_plain(row["created_at"]),
        "updated_at": to_plain(row["updated_at"]),
    }


def widget_from_row(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": int(row["id"]),
        "view_id": int(row["view_id"]),
        "title": row["title"],
        "widget_type": row["widget_type"],
        "dataset": row["dataset"],
        "config": normalize_widget_config(row["widget_type"], json_loads(row.get("config_json"), {})),
        "layout": normalize_layout(json_loads(row.get("layout_json"), default_layout())),
        "sort_order": int(row["sort_order"]),
        "analysis_text": row.get("analysis_text") or "",
        "created_at": to_plain(row.get("created_at")),
        "updated_at": to_plain(row.get("updated_at")),
    }


def load_widget(conn, widget_id: int) -> Dict[str, Any]:
    row = conn.execute(
        text(
            """
            SELECT
                id, view_id, title, widget_type, dataset, config_json, layout_json,
                sort_order, analysis_text, created_at, updated_at
            FROM bi_dashboard_widget
            WHERE id = :id
            """
        ),
        {"id": widget_id},
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail=f"鍥捐〃涓嶅瓨鍦細{widget_id}")
    return widget_from_row(dict(row))


def load_layout_template(conn, template_id: int) -> Dict[str, Any]:
    row = conn.execute(
        text(
            """
            SELECT id, name, description, layout_json, created_at, updated_at
            FROM bi_dashboard_layout_template
            WHERE id = :id
            """
        ),
        {"id": template_id},
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail=f"甯冨眬妯℃澘涓嶅瓨鍦細{template_id}")
    return layout_template_from_row(dict(row), include_layout=True)


def view_detail(conn, view_id: int) -> Dict[str, Any]:
    view = load_view(conn, view_id)
    widgets = conn.execute(
        text(
            """
            SELECT
                id, view_id, title, widget_type, dataset, config_json, layout_json,
                sort_order, analysis_text, created_at, updated_at
            FROM bi_dashboard_widget
            WHERE view_id = :view_id
            ORDER BY sort_order, id
            """
        ),
        {"view_id": view_id},
    ).mappings().all()
    return {**view, "widgets": [widget_from_row(dict(row)) for row in widgets]}


def latest_date(conn, dataset: str) -> date | None:
    if dataset == "inventory_turnover":
        latest_sales = conn.execute(text("SELECT MAX(biz_date) FROM bi_material_sales_daily_cleaning")).scalar()
        latest_inventory = conn.execute(text("SELECT MAX(snapshot_date) FROM bi_inventory_snapshot_daily_cleaning")).scalar()
        latest_candidates = [item for item in (latest_sales, latest_inventory) if item is not None]
        if not latest_candidates:
            return None
        latest_value = max(latest_candidates)
        return latest_value.replace(day=1)
    ds = DATASETS[dataset]
    return conn.execute(text(f"SELECT MAX(`{ds['date_col']}`) FROM `{ds['table']}`")).scalar()


def month_floor(value: date | None) -> date | None:
    if value is None:
        return None
    return value.replace(day=1)


def inventory_turnover_source_rows(conn) -> List[Dict[str, Any]]:
    rows = conn.execute(
        text(
            """
            WITH sales_month AS (
                SELECT
                    DATE_FORMAT(biz_date, '%Y-%m-01') AS month_date,
                    DATE_FORMAT(biz_date, '%Y-%m') AS month,
                    COALESCE(SUM(total_sales_qty), 0) AS monthly_sales_qty
                FROM bi_material_sales_daily_cleaning
                GROUP BY DATE_FORMAT(biz_date, '%Y-%m-01'), DATE_FORMAT(biz_date, '%Y-%m')
            ),
            inventory_day AS (
                SELECT
                    snapshot_date,
                    COALESCE(SUM(qty), 0) AS day_inventory_qty
                FROM bi_inventory_snapshot_daily_cleaning
                GROUP BY snapshot_date
            ),
            inventory_month AS (
                SELECT
                    DATE_FORMAT(snapshot_date, '%Y-%m-01') AS month_date,
                    DATE_FORMAT(snapshot_date, '%Y-%m') AS month,
                    COALESCE(AVG(day_inventory_qty), 0) AS avg_inventory_qty
                FROM inventory_day
                GROUP BY DATE_FORMAT(snapshot_date, '%Y-%m-01'), DATE_FORMAT(snapshot_date, '%Y-%m')
            ),
            month_union AS (
                SELECT month_date, month FROM sales_month
                UNION
                SELECT month_date, month FROM inventory_month
            )
            SELECT
                month_union.month_date,
                month_union.month,
                COALESCE(sales_month.monthly_sales_qty, 0) AS monthly_sales_qty,
                COALESCE(inventory_month.avg_inventory_qty, 0) AS avg_inventory_qty
            FROM month_union
            LEFT JOIN sales_month
                ON sales_month.month_date = month_union.month_date
                AND sales_month.month = month_union.month
            LEFT JOIN inventory_month
                ON inventory_month.month_date = month_union.month_date
                AND inventory_month.month = month_union.month
            ORDER BY month_union.month_date
            """
        )
    ).mappings().all()

    result: List[Dict[str, Any]] = []
    for row in rows:
        month_date = parse_date_or_none(str(row.get("month_date") or ""))
        monthly_sales_qty = float(to_number(row.get("monthly_sales_qty")) or 0)
        avg_inventory_qty = float(to_number(row.get("avg_inventory_qty")) or 0)
        annualized_ratio = 0.0
        inventory_turnover_days = 0.0
        if monthly_sales_qty > 0 and avg_inventory_qty > 0:
            annualized_ratio = (monthly_sales_qty * 12.0) / avg_inventory_qty
            if annualized_ratio > 0:
                inventory_turnover_days = 365.0 / annualized_ratio
        result.append(
            {
                "month_date": month_date,
                "month": str(row.get("month") or ""),
                "monthly_sales_qty": monthly_sales_qty,
                "avg_inventory_qty": avg_inventory_qty,
                "inventory_turnover_days": inventory_turnover_days,
            }
        )
    return result


def virtual_dataset_rows(conn, dataset: str) -> List[Dict[str, Any]]:
    if dataset == "inventory_turnover":
        return inventory_turnover_source_rows(conn)
    raise ValueError(f"unsupported virtual dataset: {dataset}")


def normalize_virtual_cell(field_meta: Dict[str, Any], value: Any) -> Any:
    if field_meta.get("type") == "date":
        if isinstance(value, date):
            return value
        return parse_date_or_none(str(value or ""))
    if field_meta.get("numeric"):
        number = to_number(value)
        return float(number) if number is not None else None
    return "" if value is None else str(value)


def matches_virtual_filter(field_meta: Dict[str, Any], row_value: Any, operator: str, expected: Any) -> bool:
    normalized_row = normalize_virtual_cell(field_meta, row_value)
    if operator == "like":
        source_text = str(normalized_row or "").lower()
        return str(expected or "").lower() in source_text
    if operator == "in":
        values = expected if isinstance(expected, list) else []
        normalized_values = [normalize_virtual_cell(field_meta, item) for item in values]
        return normalized_row in normalized_values
    if operator == "between":
        values = expected if isinstance(expected, list) else []
        if len(values) != 2:
            return True
        low = normalize_virtual_cell(field_meta, values[0])
        high = normalize_virtual_cell(field_meta, values[1])
        if normalized_row is None or low is None or high is None:
            return False
        return low <= normalized_row <= high
    expected_value = normalize_virtual_cell(field_meta, expected)
    if operator == "eq":
        return normalized_row == expected_value
    if operator == "ne":
        return normalized_row != expected_value
    if normalized_row is None or expected_value is None:
        return False
    if operator == "gt":
        return normalized_row > expected_value
    if operator == "gte":
        return normalized_row >= expected_value
    if operator == "lt":
        return normalized_row < expected_value
    if operator == "lte":
        return normalized_row <= expected_value
    return True


def filter_virtual_rows(
    rows: List[Dict[str, Any]],
    *,
    dataset: str,
    filters: List[Dict[str, Any]],
    target_date: date | None,
) -> List[Dict[str, Any]]:
    ds = DATASETS[dataset]
    effective_filters = list(filters)
    if target_date is not None:
        effective_filters.append({"field": ds["date_col"], "op": "eq", "value": target_date})

    result: List[Dict[str, Any]] = []
    for row in rows:
        matched = True
        for item in effective_filters:
            field = item["field"]
            field_meta = ds["fields"].get(field)
            if not field_meta:
                continue
            if not matches_virtual_filter(field_meta, row.get(field), item["op"], item.get("value")):
                matched = False
                break
        if matched:
            result.append(row)
    return result


def query_filter_options(
    conn,
    *,
    dataset: str,
    field: str,
    selected_date: date | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    keyword: str | None = None,
    limit: int = 200,
) -> Tuple[List[Dict[str, Any]], str]:
    if DATASETS[dataset].get("virtual"):
        rows = virtual_dataset_rows(conn, dataset)
        normalized_selected_date = month_floor(selected_date) if dataset == "inventory_turnover" else selected_date
        normalized_start_date = month_floor(start_date) if dataset == "inventory_turnover" else start_date
        normalized_end_date = month_floor(end_date) if dataset == "inventory_turnover" else end_date
        scoped_rows = filter_virtual_rows(
            rows,
            dataset=dataset,
            filters=[],
            target_date=normalized_selected_date if field != DATASETS[dataset]["date_col"] else None,
        )
        if normalized_start_date is not None:
            scoped_rows = filter_virtual_rows(
                scoped_rows,
                dataset=dataset,
                filters=[{"field": DATASETS[dataset]["date_col"], "op": "gte", "value": normalized_start_date}],
                target_date=None,
            )
        if normalized_end_date is not None:
            scoped_rows = filter_virtual_rows(
                scoped_rows,
                dataset=dataset,
                filters=[{"field": DATASETS[dataset]["date_col"], "op": "lte", "value": normalized_end_date}],
                target_date=None,
            )
        keyword_text = str(keyword or "").strip().lower()
        field_meta = DATASETS[dataset]["fields"][field]
        seen: set[Any] = set()
        options: List[Dict[str, Any]] = []
        for row in scoped_rows:
            raw_value = row.get(field)
            plain_value = to_plain(raw_value)
            if plain_value in (None, "") or plain_value in seen:
                continue
            label_text = str(plain_value)
            if keyword_text and keyword_text not in label_text.lower():
                continue
            seen.add(plain_value)
            options.append({"value": plain_value, "label": label_text})
        options.sort(key=lambda item: normalize_virtual_cell(field_meta, item["value"]) or item["label"])
        return options[: max(1, min(limit, 500))], "selected_date"

    ds = DATASETS[dataset]
    field_meta = ds["fields"][field]
    keyword_text = str(keyword or "").strip()
    base_params: Dict[str, Any] = {"_limit": max(1, min(limit, 500))}
    base_clauses: List[str] = [f"`{field}` IS NOT NULL"]

    if field_meta.get("type") == "string":
        base_clauses.append(f"TRIM(CAST(`{field}` AS CHAR)) <> ''")
    if keyword_text:
        base_clauses.append(f"CAST(`{field}` AS CHAR) LIKE :keyword")
        base_params["keyword"] = f"%{keyword_text}%"
    if start_date is not None:
        base_clauses.append(f"`{ds['date_col']}` >= :start_date")
        base_params["start_date"] = start_date
    if end_date is not None:
        base_clauses.append(f"`{ds['date_col']}` <= :end_date")
        base_params["end_date"] = end_date

    def run_query(use_selected_date: bool) -> List[Dict[str, Any]]:
        clauses = list(base_clauses)
        params = dict(base_params)
        if use_selected_date and selected_date is not None and field != ds["date_col"]:
            clauses.append(f"`{ds['date_col']}` = :selected_date")
            params["selected_date"] = selected_date
        sql = (
            f"SELECT DISTINCT `{field}` AS option_value "
            f"FROM `{ds['table']}` "
            f"WHERE {' AND '.join(clauses)} "
            f"ORDER BY `{field}` "
            f"LIMIT :_limit"
        )
        rows = conn.execute(text(sql), params).fetchall()
        options: List[Dict[str, Any]] = []
        for row in rows:
            raw_value = row[0]
            plain_value = to_plain(raw_value)
            if plain_value in (None, ""):
                continue
            options.append({"value": plain_value, "label": str(plain_value)})
        return options

    options = run_query(use_selected_date=True)
    if options or selected_date is None or field == ds["date_col"] or start_date is not None or end_date is not None:
        return options, "selected_date"

    return run_query(use_selected_date=False), "all_dates"


def return_unpack_attendance_summaries(
    conn,
    *,
    start_date: date | None = None,
    end_date: date | None = None,
    limit: int = 60,
) -> List[Dict[str, Any]]:
    clauses = []
    params: Dict[str, Any] = {"limit": max(1, min(limit, 365))}
    if start_date is not None:
        clauses.append("d.biz_date >= :start_date")
        params["start_date"] = start_date
    if end_date is not None:
        clauses.append("d.biz_date <= :end_date")
        params["end_date"] = end_date
    where_clause = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = conn.execute(
        text(
            f"""
            SELECT
                d.biz_date,
                COALESCE(a.attendance_count, 0) AS attendance_count,
                COALESCE(SUM(c.sales_return_warehouse), 0) AS sales_return_warehouse,
                COALESCE(SUM(c.total_return_qty), 0) AS total_return_qty,
                COALESCE(SUM(c.total_sales_qty), 0) AS total_sales_qty
            FROM (
                SELECT DISTINCT biz_date FROM bi_material_sales_daily_cleaning
                UNION
                SELECT biz_date FROM bi_return_unpack_attendance_daily
            ) d
            LEFT JOIN bi_return_unpack_attendance_daily a ON a.biz_date = d.biz_date
            LEFT JOIN bi_material_sales_daily_cleaning c ON c.biz_date = d.biz_date
            {where_clause}
            GROUP BY d.biz_date, a.attendance_count
            ORDER BY d.biz_date DESC
            LIMIT :limit
            """
        ),
        params,
    ).mappings().all()

    summaries: List[Dict[str, Any]] = []
    for row in rows:
        attendance_count = Decimal(str(row["attendance_count"] or 0))
        sales_return_warehouse = Decimal(str(row["sales_return_warehouse"] or 0))
        total_return_qty = Decimal(str(row["total_return_qty"] or 0))
        total_sales_qty = Decimal(str(row["total_sales_qty"] or 0))
        efficiency = Decimal("0")
        if attendance_count > 0:
            efficiency = sales_return_warehouse / attendance_count
        summaries.append(
            {
                "biz_date": to_plain(row["biz_date"]),
                "attendance_count": float(attendance_count),
                "sales_return_warehouse": float(sales_return_warehouse),
                "total_return_qty": float(total_return_qty),
                "total_sales_qty": float(total_sales_qty),
                "return_unpack_efficiency": float(efficiency),
            }
        )
    return summaries


def merge_filters(dataset: str, widget_filters: List[Dict[str, Any]], global_filters: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    merged = list(widget_filters)
    merged.extend(
        {
            "field": item["field"],
            "op": item["op"],
            "value": item.get("value"),
        }
        for item in global_filters
        if item.get("dataset") == dataset
    )
    return merged


def filter_has_value(operator: str, value: Any) -> bool:
    if operator in {"in", "between"}:
        return isinstance(value, list) and any(str(part).strip() for part in value)
    return value not in (None, "")


def should_apply_target_date(dataset: str, dimensions: List[str], filters: List[Dict[str, Any]]) -> bool:
    date_col = DATASETS[dataset]["date_col"]
    if dataset == "inventory_turnover" and any(field in dimensions for field in {"month", "month_date"}):
        return False
    if date_col in dimensions:
        return False
    return not any(
        item.get("field") in {date_col, "month"} and filter_has_value(str(item.get("op") or "eq").lower(), item.get("value"))
        for item in filters
    )


def resolve_widget_date_context(
    conn,
    *,
    dataset: str,
    config: Dict[str, Any],
    selected_date: date | None,
) -> Tuple[List[Dict[str, Any]], date | None, Any, str]:
    date_filter = normalize_widget_date_filter(dataset, config.get("date_filter"))
    config["date_filter"] = date_filter
    mode = date_filter["mode"]
    date_col = DATASETS[dataset]["date_col"]
    normalize_date = month_floor if dataset == "inventory_turnover" else (lambda value: value)

    if mode == "single" and date_filter["date"]:
        target = normalize_date(parse_date_or_none(date_filter["date"]))
        return ([{"field": date_col, "op": "eq", "value": to_plain(target)}], None, to_plain(target), mode)

    if mode == "range":
        start_value = to_plain(normalize_date(parse_date_or_none(date_filter["start_date"])))
        end_value = to_plain(normalize_date(parse_date_or_none(date_filter["end_date"])))
        if start_value and end_value:
            label = f"{start_value} 至 {end_value}"
            return ([{"field": date_col, "op": "between", "value": [start_value, end_value]}], None, label, mode)
        if start_value:
            label = f"{start_value} 起"
            return ([{"field": date_col, "op": "gte", "value": start_value}], None, label, mode)
        if end_value:
            label = f"截至 {end_value}"
            return ([{"field": date_col, "op": "lte", "value": end_value}], None, label, mode)

    if mode == "all":
        return ([], None, "鍏ㄩ儴鏃ユ湡", mode)

    target_date = normalize_date(selected_date or latest_date(conn, dataset))
    return ([], target_date, to_plain(target_date), "follow_page")


def build_where_sql(dataset: str, filters: List[Dict[str, Any]], target_date: date | None) -> Tuple[str, Dict[str, Any]]:
    ds = DATASETS[dataset]
    clauses: List[str] = []
    params: Dict[str, Any] = {}
    if target_date is not None:
        clauses.append(f"`{ds['date_col']}` = :target_date")
        params["target_date"] = target_date
    for idx, item in enumerate(filters):
        field = item["field"]
        operator = item["op"]
        value = item.get("value")
        key = f"f_{idx}"
        if operator == "eq":
            clauses.append(f"`{field}` = :{key}")
            params[key] = value
        elif operator == "ne":
            clauses.append(f"`{field}` <> :{key}")
            params[key] = value
        elif operator == "gt":
            clauses.append(f"`{field}` > :{key}")
            params[key] = value
        elif operator == "gte":
            clauses.append(f"`{field}` >= :{key}")
            params[key] = value
        elif operator == "lt":
            clauses.append(f"`{field}` < :{key}")
            params[key] = value
        elif operator == "lte":
            clauses.append(f"`{field}` <= :{key}")
            params[key] = value
        elif operator == "like":
            clauses.append(f"`{field}` LIKE :{key}")
            params[key] = f"%{value}%"
        elif operator == "in":
            values = value if isinstance(value, list) else []
            if not values:
                continue
            placeholders: List[str] = []
            for sub_idx, sub_value in enumerate(values):
                sub_key = f"{key}_{sub_idx}"
                params[sub_key] = sub_value
                placeholders.append(f":{sub_key}")
            clauses.append(f"`{field}` IN ({', '.join(placeholders)})")
        elif operator == "between":
            values = value if isinstance(value, list) else []
            if len(values) != 2:
                continue
            low_key = f"{key}_low"
            high_key = f"{key}_high"
            params[low_key] = values[0]
            params[high_key] = values[1]
            clauses.append(f"`{field}` BETWEEN :{low_key} AND :{high_key}")
    return (" AND ".join(clauses) if clauses else "1=1"), params


def metric_sql(metric: Dict[str, Any], alias: str) -> str:
    if metric["agg"] == "count":
        return f"COUNT(*) AS `{alias}`" if metric["field"] == "*" else f"COUNT(`{metric['field']}`) AS `{alias}`"
    if metric["agg"] == "sum":
        return f"COALESCE(SUM(`{metric['field']}`), 0) AS `{alias}`"
    if metric["agg"] == "avg":
        return f"COALESCE(AVG(`{metric['field']}`), 0) AS `{alias}`"
    if metric["agg"] == "max":
        return f"COALESCE(MAX(`{metric['field']}`), 0) AS `{alias}`"
    if metric["agg"] == "min":
        return f"COALESCE(MIN(`{metric['field']}`), 0) AS `{alias}`"
    raise ValueError(metric["agg"])


def python_aggregate(raw_rows: List[Dict[str, Any]], dimensions: List[str], metrics: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    for raw_row in raw_rows:
        key = tuple(raw_row.get(dim) for dim in dimensions)
        holder = grouped.setdefault(
            key,
            {
                "dimensions": {dim: raw_row.get(dim) for dim in dimensions},
                "values": {f"metric_{idx}": [] for idx in range(len(metrics))},
            },
        )
        for idx, metric in enumerate(metrics):
            alias = f"metric_{idx}"
            if metric["agg"] == "count":
                if metric["field"] == "*" or raw_row.get(metric["field"]) not in (None, ""):
                    holder["values"][alias].append(1.0)
                continue
            numeric_value = to_number(raw_row.get(metric["field"]))
            if numeric_value is not None:
                holder["values"][alias].append(numeric_value)

    result: List[Dict[str, Any]] = []
    for holder in grouped.values():
        row = dict(holder["dimensions"])
        for idx, metric in enumerate(metrics):
            alias = f"metric_{idx}"
            values = holder["values"][alias]
            if metric["agg"] == "count":
                row[alias] = int(sum(values))
            elif not values:
                row[alias] = 0
            elif metric["agg"] == "sum":
                row[alias] = float(sum(values))
            elif metric["agg"] == "avg":
                row[alias] = float(sum(values) / len(values))
            elif metric["agg"] == "max":
                row[alias] = float(max(values))
            elif metric["agg"] == "min":
                row[alias] = float(min(values))
            elif metric["agg"] == "median":
                row[alias] = float(median(values))
        result.append(row)
    return result


def apply_sort(rows: List[Dict[str, Any]], sort_config: List[Dict[str, Any]], allowed_fields: List[str]) -> None:
    valid_sort = [item for item in sort_config if item.get("field") in allowed_fields]
    for item in reversed(valid_sort):
        field = item["field"]
        reverse = item["direction"] == "desc"
        rows.sort(key=lambda row: (row.get(field) is None, row.get(field)), reverse=reverse)


def pivot_rows_by_series(
    rows: List[Dict[str, Any]],
    *,
    dimensions: List[str],
    metrics: List[Dict[str, Any]],
    series_field: str,
    sort_config: List[Dict[str, Any]],
    limit: int,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, str]]]:
    metric_aliases = [f"metric_{idx}" for idx in range(len(metrics))]
    grouped: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    series_labels: Dict[str, str] = {}

    for raw_row in rows:
        key = tuple(raw_row.get(dim) for dim in dimensions)
        holder = grouped.setdefault(
            key,
            {
                **{dim: raw_row.get(dim) for dim in dimensions},
                **{alias: 0 for alias in metric_aliases},
                "series_values": {},
            },
        )
        series_value = to_plain(raw_row.get(series_field))
        series_key = "" if series_value in (None, "") else str(series_value)
        series_labels.setdefault(series_key, series_key or "未分类")
        slot = holder["series_values"].setdefault(
            series_key,
            {"label": series_labels[series_key], **{alias: 0 for alias in metric_aliases}},
        )
        for alias in metric_aliases:
            numeric = to_number(raw_row.get(alias))
            if numeric is None:
                continue
            slot[alias] = float(slot.get(alias, 0) or 0) + float(numeric)
            holder[alias] = float(holder.get(alias, 0) or 0) + float(numeric)

    pivoted_rows = list(grouped.values())
    apply_sort(pivoted_rows, sort_config, dimensions + metric_aliases)
    limited_rows = pivoted_rows[:limit]

    visible_series_keys: List[str] = []
    for row in limited_rows:
        for series_key in row.get("series_values", {}).keys():
            if series_key not in visible_series_keys:
                visible_series_keys.append(series_key)
    visible_series_keys.sort(key=lambda item: series_labels.get(item, item or "未分类"))

    series_groups = [
        {"key": series_key, "label": series_labels.get(series_key, series_key or "未分类")}
        for series_key in visible_series_keys
    ]
    return limited_rows, series_groups


def query_widget_data(conn, widget: Dict[str, Any], selected_date: date | None, global_filters: List[Dict[str, Any]]) -> Dict[str, Any]:
    config = copy.deepcopy(widget["config"])
    dataset = config["dataset"]
    config["filters"] = merge_filters(dataset, config["filters"], global_filters)
    date_filters, implicit_target_date, target_date, date_filter_scope = resolve_widget_date_context(
        conn,
        dataset=dataset,
        config=config,
        selected_date=selected_date,
    )
    config["filters"].extend(date_filters)

    if widget["widget_type"] == "text":
        return {
            "target_date": target_date,
            "dimensions": [],
            "metrics": [],
            "rows": [],
            "config": config,
            "date_filter_scope": date_filter_scope,
        }

    dimensions = config["dimensions"]
    metrics = config["metrics"]
    series_field = config.get("series_field") or ""
    use_series_breakdown = bool(series_field and supports_series_field(widget["widget_type"]) and dimensions and widget["widget_type"] != "text")
    group_dimensions = dimensions + ([series_field] if use_series_breakdown else [])
    applied_target_date = implicit_target_date if should_apply_target_date(dataset, dimensions, config["filters"]) else None
    if DATASETS[dataset].get("virtual"):
        raw_rows = filter_virtual_rows(
            virtual_dataset_rows(conn, dataset),
            dataset=dataset,
            filters=config["filters"],
            target_date=applied_target_date,
        )
        rows = python_aggregate(raw_rows, group_dimensions, metrics)
        metric_aliases = [f"metric_{idx}" for idx in range(len(metrics))]
        series_groups: List[Dict[str, str]] = []
        if use_series_breakdown:
            rows, series_groups = pivot_rows_by_series(
                rows,
                dimensions=dimensions,
                metrics=metrics,
                series_field=series_field,
                sort_config=config["sort"],
                limit=config["limit"],
            )
        else:
            apply_sort(rows, config["sort"], dimensions + metric_aliases)
            rows = rows[: config["limit"]]
        normalized_rows = [{key: to_plain(value) for key, value in row.items()} for row in rows]
        normalized_metrics = [
            {
                "alias": f"metric_{idx}",
                "field": metric["field"],
                "agg": metric["agg"],
                "label": metric.get("label") or metric_label(dataset, metric["field"], metric["agg"]),
            }
            for idx, metric in enumerate(metrics)
        ]
        return {
            "target_date": target_date,
            "applied_target_date": to_plain(applied_target_date),
            "dimensions": dimensions,
            "series_field": series_field if use_series_breakdown else "",
            "series_groups": series_groups,
            "metrics": normalized_metrics,
            "rows": normalized_rows,
            "config": config,
            "date_filter_scope": date_filter_scope,
        }

    where_clause, params = build_where_sql(dataset, config["filters"], applied_target_date)
    table_name = DATASETS[dataset]["table"]
    has_median = any(metric["agg"] == "median" for metric in metrics)

    if has_median:
        select_columns = [f"`{dim}` AS `{dim}`" for dim in group_dimensions]
        raw_fields = sorted({metric["field"] for metric in metrics if metric["field"] != "*"})
        select_columns.extend(f"`{field}` AS `{field}`" for field in raw_fields)
        sql = f"SELECT {', '.join(select_columns) if select_columns else '*'} FROM `{table_name}` WHERE {where_clause}"
        raw_rows = [dict(row) for row in conn.execute(text(sql), params).mappings().all()]
        rows = python_aggregate(raw_rows, group_dimensions, metrics)
    else:
        metric_aliases = [f"metric_{idx}" for idx in range(len(metrics))]
        select_parts = [f"`{dim}` AS `{dim}`" for dim in group_dimensions]
        select_parts.extend(metric_sql(metric, metric_aliases[idx]) for idx, metric in enumerate(metrics))
        sql = f"SELECT {', '.join(select_parts)} FROM `{table_name}` WHERE {where_clause}"
        if group_dimensions:
            sql += " GROUP BY " + ", ".join(f"`{dim}`" for dim in group_dimensions)
        allowed_fields = dimensions + metric_aliases
        if not use_series_breakdown:
            order_parts = [
                f"`{item['field']}` {item['direction'].upper()}"
                for item in config["sort"]
                if item.get("field") in allowed_fields
            ]
            if not order_parts and metric_aliases:
                order_parts = [f"`{metric_aliases[0]}` DESC"]
            if order_parts:
                sql += " ORDER BY " + ", ".join(order_parts)
            sql += " LIMIT :_limit"
            params["_limit"] = config["limit"]
        rows = [dict(row) for row in conn.execute(text(sql), params).mappings().all()]

    metric_aliases = [f"metric_{idx}" for idx in range(len(metrics))]
    series_groups: List[Dict[str, str]] = []
    if use_series_breakdown:
        rows, series_groups = pivot_rows_by_series(
            rows,
            dimensions=dimensions,
            metrics=metrics,
            series_field=series_field,
            sort_config=config["sort"],
            limit=config["limit"],
        )
    else:
        apply_sort(rows, config["sort"], dimensions + metric_aliases)
        rows = rows[: config["limit"]]
    normalized_rows = [{key: to_plain(value) for key, value in row.items()} for row in rows]
    normalized_metrics = [
        {
            "alias": f"metric_{idx}",
            "field": metric["field"],
            "agg": metric["agg"],
            "label": metric.get("label") or metric_label(dataset, metric["field"], metric["agg"]),
        }
        for idx, metric in enumerate(metrics)
    ]
    return {
        "target_date": target_date,
        "applied_target_date": to_plain(applied_target_date),
        "dimensions": dimensions,
        "series_field": series_field if use_series_breakdown else "",
        "series_groups": series_groups,
        "metrics": normalized_metrics,
        "rows": normalized_rows,
        "config": config,
        "date_filter_scope": date_filter_scope,
    }


def build_ai_text(widget: Dict[str, Any], payload: Dict[str, Any]) -> str:
    if widget["widget_type"] == "text":
        return "文本组件不需要自动分析。"
    rows = payload.get("rows") or []
    metrics = payload.get("metrics") or []
    if not rows:
        return f"{payload.get('target_date') or '未知日期'} 当前组件没有可分析数据，建议检查筛选条件或切换日期。"
    if not metrics:
        return f"{payload.get('target_date') or '未知日期'} 共返回 {len(rows)} 条记录。"
    metric = metrics[0]
    alias = metric["alias"]
    dimension = payload["dimensions"][0] if payload["dimensions"] else None
    total = 0.0
    peak_name = "鏁翠綋"
    peak_value = None
    for row in rows:
        value = to_number(row.get(alias))
        if value is None:
            continue
        total += value
        if peak_value is None or value > peak_value:
            peak_value = value
            peak_name = str(row.get(dimension) or "鏁翠綋") if dimension else "鏁翠綋"
    if peak_value is None:
        return f"{payload.get('target_date') or '未知日期'} {metric['label']} 暂无可量化值。"
    return (
        f"{payload.get('target_date') or '未知日期'}，{metric['label']} 合计 {total:.2f}。"
        f"最高项为 {peak_name}，数值 {peak_value:.2f}。"
        "建议结合组织、仓库或物料继续下钻确认结构变化。"
    )


def data_agent_session_id(prefix: str = "polaris-agent") -> str:
    return f"{prefix}-{int(time.time())}-{secrets.token_hex(4)}"


def call_data_agent_chat(message: str, session_id: str | None = None, timeout: int = 180) -> Dict[str, Any]:
    if not is_data_agent_api_online():
        raise RuntimeError(f"{DATA_AGENT_NAME} 服务未启动")
    session_id = session_id or data_agent_session_id()
    response = requests.post(
        f"{DATA_AGENT_API_URL}/api/chat/stream",
        json={"messages": [{"role": "user", "content": message}], "session_id": session_id},
        headers={"Content-Type": "application/json"},
        stream=True,
        timeout=timeout,
    )
    if response.status_code != 200:
        detail = response.text[:400] if response.text else f"HTTP {response.status_code}"
        raise RuntimeError(f"{DATA_AGENT_NAME} 调用失败：{detail}")

    parts: List[str] = []
    last_message_id = ""
    current_event_type = ""
    for raw_line in response.iter_lines(decode_unicode=True):
        if not raw_line or raw_line.strip() == "":
            continue
        if raw_line.startswith("event:"):
            current_event_type = raw_line.split(":", 1)[1].strip()
            continue
        if not raw_line.startswith("data:"):
            continue
        data_str = raw_line.split(":", 1)[1].strip()
        try:
            payload = json.loads(data_str)
        except Exception:
            continue
        if current_event_type not in {"message", "message_chunk"} and payload.get("type") == "separator":
            continue
        message_id = str(payload.get("id") or "")
        content = str(payload.get("content") or "")
        if not content:
            continue
        if parts and message_id and last_message_id and message_id != last_message_id:
            parts.append("\n")
        parts.append(content)
        if message_id:
            last_message_id = message_id

    result = "".join(parts).strip()
    if not result:
        raise RuntimeError(f"{DATA_AGENT_NAME} 未返回有效内容")
    return {"session_id": session_id, "content": result}


def build_widget_agent_prompt(widget: Dict[str, Any], payload: Dict[str, Any]) -> str:
    dataset_label = DATASETS.get(widget["dataset"], {}).get("label", widget["dataset"])
    preview_rows = list(payload.get("rows") or [])[:12]
    normalized_rows = [{key: to_plain(value) for key, value in row.items()} for row in preview_rows]
    return (
        f"你是北极星系统中的 {DATA_AGENT_NAME}。\n"
        "请基于以下图表数据，输出一段简洁、可直接展示在经营看板右侧的中文分析结论。\n"
        "要求：\n"
        "1. 只输出中文。\n"
        "2. 控制在 4 到 6 句话。\n"
        "3. 先写核心结论，再写可能原因，最后给出一条建议。\n"
        "4. 不要编造未提供的数据。\n\n"
        f"图表标题：{widget['title']}\n"
        f"图表类型：{WIDGET_TYPES.get(widget['widget_type'], widget['widget_type'])}\n"
        f"数据集：{dataset_label}\n"
        f"目标日期：{payload.get('target_date') or payload.get('applied_target_date') or '--'}\n"
        f"维度字段：{json.dumps(payload.get('dimensions') or [], ensure_ascii=False)}\n"
        f"指标字段：{json.dumps(payload.get('metrics') or [], ensure_ascii=False)}\n"
        f"数据预览：{json.dumps(normalized_rows, ensure_ascii=False)}"
    )


def build_data_agent_report_period(conn, report_type: str) -> Dict[str, Any]:
    candidate_dates: List[date] = []
    sales_date = conn.execute(text("SELECT MAX(biz_date) FROM bi_material_sales_daily_cleaning")).scalar()
    inventory_date = conn.execute(text("SELECT MAX(snapshot_date) FROM bi_inventory_snapshot_daily_cleaning")).scalar()
    refurb_date = conn.execute(text("SELECT MAX(biz_date) FROM bi_refurb_production_daily")).scalar()
    for item in (sales_date, inventory_date, refurb_date):
        if isinstance(item, date):
            candidate_dates.append(item)
    anchor = max(candidate_dates) if candidate_dates else date.today()
    report_type = str(report_type or "weekly").strip().lower()
    if report_type == "monthly":
        period_start = anchor.replace(day=1)
        period_end = anchor
        period_label = f"{anchor:%Y-%m} 鏈堟姤"
    else:
        report_type = "weekly"
        period_end = anchor
        period_start = anchor - timedelta(days=6)
        period_label = f"{period_start.isoformat()} ~ {period_end.isoformat()} 鍛ㄦ姤"
    return {
        "report_type": report_type,
        "period_start": period_start,
        "period_end": period_end,
        "period_label": period_label,
        "anchor_date": anchor,
    }


def build_data_agent_report_summary(conn, report_type: str) -> Dict[str, Any]:
    period = build_data_agent_report_period(conn, report_type)
    start_date = period["period_start"]
    end_date = period["period_end"]

    sales_row = conn.execute(
        text(
            """
            SELECT
                COALESCE(SUM(total_sales_qty), 0) AS total_sales_qty,
                COALESCE(SUM(total_return_qty), 0) AS total_return_qty,
                COUNT(DISTINCT biz_date) AS active_days,
                COUNT(DISTINCT material_name) AS material_count
            FROM bi_material_sales_daily_cleaning
            WHERE biz_date BETWEEN :start_date AND :end_date
            """
        ),
        {"start_date": start_date, "end_date": end_date},
    ).mappings().first() or {}

    top_sales_rows = conn.execute(
        text(
            """
            SELECT material_name,
                   ROUND(COALESCE(SUM(total_sales_qty), 0), 2) AS total_sales_qty,
                   ROUND(COALESCE(SUM(total_return_qty), 0), 2) AS total_return_qty
            FROM bi_material_sales_daily_cleaning
            WHERE biz_date BETWEEN :start_date AND :end_date
            GROUP BY material_name
            HAVING COALESCE(SUM(total_sales_qty), 0) > 0
            ORDER BY total_sales_qty DESC, material_name ASC
            LIMIT 5
            """
        ),
        {"start_date": start_date, "end_date": end_date},
    ).mappings().all()

    latest_inventory_date = conn.execute(
        text("SELECT MAX(snapshot_date) FROM bi_inventory_snapshot_daily_cleaning WHERE snapshot_date <= :end_date"),
        {"end_date": end_date},
    ).scalar()
    inventory_total_row = {}
    inventory_by_warehouse: List[Dict[str, Any]] = []
    if isinstance(latest_inventory_date, date):
        inventory_total_row = conn.execute(
            text(
                """
                SELECT ROUND(COALESCE(SUM(qty), 0), 2) AS total_qty,
                       COUNT(DISTINCT material_code) AS material_count
                FROM bi_inventory_snapshot_daily_cleaning
                WHERE snapshot_date = :snapshot_date
                """
            ),
            {"snapshot_date": latest_inventory_date},
        ).mappings().first() or {}
        inventory_by_warehouse = [
            {key: to_plain(value) for key, value in row.items()}
            for row in conn.execute(
                text(
                    """
                    SELECT warehouse_name_clean,
                           ROUND(COALESCE(SUM(qty), 0), 2) AS total_qty
                    FROM bi_inventory_snapshot_daily_cleaning
                    WHERE snapshot_date = :snapshot_date
                    GROUP BY warehouse_name_clean
                    ORDER BY total_qty DESC, warehouse_name_clean ASC
                    LIMIT 8
                    """
                ),
                {"snapshot_date": latest_inventory_date},
            ).mappings().all()
        ]

    refurb_row = conn.execute(
        text(
            """
            SELECT
                ROUND(COALESCE(SUM(plan_qty), 0), 2) AS plan_qty,
                ROUND(COALESCE(SUM(final_good_qty), 0), 2) AS final_good_qty,
                ROUND(COALESCE(AVG(plan_achievement_rate), 0), 4) AS avg_plan_achievement_rate,
                ROUND(COALESCE(AVG(refurb_efficiency), 0), 4) AS avg_refurb_efficiency
            FROM bi_refurb_production_daily
            WHERE biz_date BETWEEN :start_date AND :end_date
            """
        ),
        {"start_date": start_date, "end_date": end_date},
    ).mappings().first() or {}

    alert_snapshot_date = conn.execute(
        text("SELECT MAX(snapshot_date) FROM bi_inventory_alert_log WHERE snapshot_date <= :end_date"),
        {"end_date": end_date},
    ).scalar()
    alert_row = {}
    if isinstance(alert_snapshot_date, date):
        alert_row = conn.execute(
            text(
                """
                SELECT
                    COUNT(*) AS alert_count,
                    ROUND(COALESCE(SUM(shortage_qty_14d), 0), 2) AS shortage_qty_14d
                FROM bi_inventory_alert_log
                WHERE snapshot_date = :snapshot_date AND alert_level <> 'safe'
                """
            ),
            {"snapshot_date": alert_snapshot_date},
        ).mappings().first() or {}

    return {
        **period,
        "sales": {key: to_plain(value) for key, value in dict(sales_row).items()},
        "top_sales_materials": [{key: to_plain(value) for key, value in row.items()} for row in top_sales_rows],
        "inventory": {
            "snapshot_date": to_plain(latest_inventory_date),
            **{key: to_plain(value) for key, value in dict(inventory_total_row).items()},
        },
        "inventory_by_warehouse": inventory_by_warehouse,
        "refurb": {key: to_plain(value) for key, value in dict(refurb_row).items()},
        "alerts": {
            "snapshot_date": to_plain(alert_snapshot_date),
            **{key: to_plain(value) for key, value in dict(alert_row).items()},
        },
    }


def build_data_agent_report_prompt(summary: Dict[str, Any]) -> str:
    return (
        f"你是北极星系统中的 {DATA_AGENT_NAME}。\n"
        "请基于以下经营摘要生成一份中文分析报告。\n"
        "要求：\n"
        "1. 使用中文输出。\n"
        "2. 结构包含：经营概览、销售与退货、库存结构、翻新生产、风险提示、建议动作。\n"
        "3. 只基于提供的数据得出结论，不要编造额外数字。\n"
        "4. 语气专业、简洁，适合作为周报或月报正文。\n\n"
        f"摘要数据：{json.dumps(summary, ensure_ascii=False, default=to_plain)}"
    )

def build_local_agent_report(summary: Dict[str, Any]) -> str:
    sales = summary.get("sales") or {}
    inventory = summary.get("inventory") or {}
    refurb = summary.get("refurb") or {}
    alerts = summary.get("alerts") or {}
    top_items = summary.get("top_sales_materials") or []
    top_text = "、".join(
        f"{item.get('material_name', '--')}（销量 {item.get('total_sales_qty', 0)}）" for item in top_items[:3]
    ) or "暂无头部物料"
    return (
        f"{summary.get('period_label', '--')} 经营分析报告\n\n"
        "一、经营概览\n"
        f"本期累计销量 {sales.get('total_sales_qty', 0)}，累计退货 {sales.get('total_return_qty', 0)}，覆盖 {sales.get('active_days', 0)} 个业务日，涉及 {sales.get('material_count', 0)} 个物料。\n\n"
        "二、销售与退货\n"
        f"销量头部物料主要集中在 {top_text}。建议重点复盘头部 SKU 的退货结构与渠道节奏，确认销量增长是否伴随异常退货。\n\n"
        "三、库存结构\n"
        f"最新库存快照日期为 {inventory.get('snapshot_date') or '--'}，库存总量 {inventory.get('total_qty', 0)}，覆盖物料 {inventory.get('material_count', 0)} 个。建议结合仓库分布继续判断良品、委外与销退仓的结构变化。\n\n"
        "四、翻新生产\n"
        f"本期翻新计划量 {refurb.get('plan_qty', 0)}，最终合格数量 {refurb.get('final_good_qty', 0)}，平均计划达成率 {round(float(refurb.get('avg_plan_achievement_rate') or 0) * 100, 2)}%，平均翻新人效 {refurb.get('avg_refurb_efficiency', 0)}。\n\n"
        "五、风险提示\n"
        f"最近一次预警快照日期 {alerts.get('snapshot_date') or '--'}，风险预警数量 {alerts.get('alert_count', 0)}，14 天缺口合计 {alerts.get('shortage_qty_14d', 0)}。建议优先核对缺口较大的整机和翻新物料。\n\n"
        "六、建议动作\n"
        "建议结合组织、仓库或物料继续下钻确认结构变化。"
    )

def upsert_data_agent_report(conn, report_type: str, trigger_mode: str, summary: Dict[str, Any], report_content: str, generated_by: str) -> Dict[str, Any]:
    title = f"{summary.get('period_label', '--')} {DATA_AGENT_NAME}鎶ュ憡"
    conn.execute(
        text(
            """
            INSERT INTO bi_data_agent_report(
                report_type, period_start, period_end, period_label, trigger_mode, title,
                summary_json, report_content, generated_by, status
            )
            VALUES(
                :report_type, :period_start, :period_end, :period_label, :trigger_mode, :title,
                :summary_json, :report_content, :generated_by, 'success'
            )
            ON DUPLICATE KEY UPDATE
                trigger_mode = VALUES(trigger_mode),
                title = VALUES(title),
                summary_json = VALUES(summary_json),
                report_content = VALUES(report_content),
                generated_by = VALUES(generated_by),
                status = 'success'
            """
        ),
        {
            "report_type": report_type,
            "period_start": summary["period_start"],
            "period_end": summary["period_end"],
            "period_label": summary["period_label"],
            "trigger_mode": trigger_mode,
            "title": title,
            "summary_json": json.dumps(summary, ensure_ascii=False, default=to_plain),
            "report_content": report_content,
            "generated_by": generated_by,
        },
    )
    row = conn.execute(
        text(
            """
            SELECT id, report_type, period_start, period_end, period_label, trigger_mode, title,
                   summary_json, report_content, generated_by, status, created_at, updated_at
            FROM bi_data_agent_report
            WHERE report_type = :report_type AND period_start = :period_start AND period_end = :period_end
            """
        ),
        {
            "report_type": report_type,
            "period_start": summary["period_start"],
            "period_end": summary["period_end"],
        },
    ).mappings().first()
    return dict(row) if row else {}


def report_row_to_payload(row: Dict[str, Any]) -> Dict[str, Any]:
    result = {key: to_plain(value) for key, value in row.items()}
    result["summary"] = json_loads(result.pop("summary_json", "{}"), {})
    return result


def list_data_agent_reports(conn, limit: int = 20) -> List[Dict[str, Any]]:
    rows = conn.execute(
        text(
            """
            SELECT id, report_type, period_start, period_end, period_label, trigger_mode, title,
                   summary_json, report_content, generated_by, status, created_at, updated_at
            FROM bi_data_agent_report
            ORDER BY period_end DESC, created_at DESC
            LIMIT :limit
            """
        ),
        {"limit": max(1, min(int(limit), 100))},
    ).mappings().all()
    return [report_row_to_payload(dict(row)) for row in rows]


def generate_data_agent_report(report_type: str, trigger_mode: str = "manual") -> Dict[str, Any]:
    current_engine = get_engine()
    with current_engine.begin() as conn:
        summary = build_data_agent_report_summary(conn, report_type)
        generated_by = "fallback"
        report_content = build_local_agent_report(summary)
        if is_data_agent_api_online() and data_agent_env_snapshot().get("openai_api_key"):
            try:
                result = call_data_agent_chat(
                    build_data_agent_report_prompt(summary),
                    session_id=data_agent_session_id(f"polaris-{report_type}-report"),
                    timeout=240,
                )
                if result.get("content"):
                    report_content = str(result["content"]).strip()
                    generated_by = "data-agent"
            except Exception as exc:
                app_logger.warning("Data agent report fallback used: %s", exc)
        row = upsert_data_agent_report(conn, report_type, trigger_mode, summary, report_content, generated_by)
    return report_row_to_payload(row)


def run_weekly_data_agent_report() -> None:
    try:
        generate_data_agent_report("weekly", "scheduled")
    except Exception as exc:  # pragma: no cover - scheduler guard
        app_logger.warning("Weekly data agent report generation failed: %s", exc)


def run_monthly_data_agent_report() -> None:
    try:
        generate_data_agent_report("monthly", "scheduled")
    except Exception as exc:  # pragma: no cover - scheduler guard
        app_logger.warning("Monthly data agent report generation failed: %s", exc)


def generate_widget_ai_text(widget: Dict[str, Any], payload: Dict[str, Any]) -> str:
    fallback_text = build_ai_text(widget, payload)
    if widget["widget_type"] == "text":
        return fallback_text
    if not data_agent_env_snapshot().get("openai_api_key") or not is_data_agent_api_online():
        return fallback_text
    try:
        result = call_data_agent_chat(
            build_widget_agent_prompt(widget, payload),
            session_id=data_agent_session_id(f"polaris-widget-{widget['id']}"),
            timeout=120,
        )
        return str(result.get("content") or "").strip() or fallback_text
    except Exception as exc:
        app_logger.warning("Widget AI analysis fallback used: %s", exc)
        return fallback_text


@router.get("/bi-dashboard/login", response_class=HTMLResponse)
async def bi_dashboard_login(request: Request, next: str | None = Query(None)) -> Response:
    target = sanitize_next_path(next)
    if current_dashboard_user(request):
        return RedirectResponse(url=target, status_code=303)
    return HTMLResponse(
        render_template(
            "bi_dashboard_login.html",
            {
                "__NEXT_PATH_JSON__": json.dumps(target, ensure_ascii=False),
                "__BI_LOGO_WORDMARK__": dashboard_logo_wordmark_svg(),
                "__BI_LOGO_BADGE__": dashboard_logo_badge_svg(),
            },
        )
    )


@router.post("/bi-dashboard/login")
async def bi_dashboard_login_submit(payload: Dict[str, Any] = Body(default={})) -> JSONResponse:
    username = str(payload.get("username") or "").strip()
    password = str(payload.get("password") or "")
    remember = bool(payload.get("remember"))
    next_path = sanitize_next_path(payload.get("next"))
    if not authenticate_dashboard_user(username, password):
        raise HTTPException(status_code=401, detail="鐢ㄦ埛鍚嶆垨瀵嗙爜閿欒")

    response = JSONResponse({"ok": True, "redirect_to": next_path, "username": username})
    response.set_cookie(
        key=DASHBOARD_SESSION_COOKIE,
        value=create_dashboard_session(username),
        max_age=DASHBOARD_SESSION_MAX_AGE if remember else None,
        httponly=True,
        samesite="lax",
        path="/",
    )
    return response


@router.post("/bi-dashboard/logout")
async def bi_dashboard_logout() -> JSONResponse:
    response = JSONResponse({"ok": True, "redirect_to": "/financial/bi-dashboard/login"})
    response.delete_cookie(DASHBOARD_SESSION_COOKIE, path="/")
    return response


@router.get("/bi-dashboard/logout")
async def bi_dashboard_logout_redirect() -> RedirectResponse:
    response = RedirectResponse(url="/financial/bi-dashboard/login", status_code=303)
    response.delete_cookie(DASHBOARD_SESSION_COOKIE, path="/")
    return response


@router.get("/bi-dashboard", response_class=HTMLResponse)
async def bi_dashboard(request: Request) -> Response:
    ensure_schema()
    username = current_dashboard_user(request)
    if not username:
        current_path = request.url.path
        if request.url.query:
            current_path = f"{current_path}?{request.url.query}"
        login_url = f"/financial/bi-dashboard/login?next={quote(current_path, safe='')}"
        return RedirectResponse(url=login_url, status_code=303)
    return HTMLResponse(
        render_template(
            "bi_dashboard_runtime.html",
            dashboard_page_context(
                "dashboard",
                {
                "__BI_CURRENT_USER_JSON__": json.dumps(username, ensure_ascii=False),
                "__BI_LOGIN_PATH_JSON__": json.dumps("/financial/bi-dashboard/login", ensure_ascii=False),
                "__BI_LOGOUT_PATH_JSON__": json.dumps("/financial/bi-dashboard/logout", ensure_ascii=False),
                "__BI_EDITOR_PATH_JSON__": json.dumps(DASHBOARD_EDITOR_PATH, ensure_ascii=False),
                "__BI_LOGO_WORDMARK__": dashboard_logo_wordmark_svg(),
                "__BI_LOGO_BADGE_SMALL__": dashboard_logo_badge_small_svg(),
                },
            ),
        )
    )


@router.get("/bi-dashboard/editor", response_class=HTMLResponse)
async def bi_dashboard_editor(request: Request) -> Response:
    ensure_schema()
    username = current_dashboard_user(request)
    if not username:
        current_path = request.url.path
        if request.url.query:
            current_path = f"{current_path}?{request.url.query}"
        login_url = f"/financial/bi-dashboard/login?next={quote(current_path, safe='')}"
        return RedirectResponse(url=login_url, status_code=303)
    return HTMLResponse(
        render_template(
            "bi_dashboard_builder.html",
            dashboard_page_context(
                "editor",
                {
                "__BI_CURRENT_USER_JSON__": json.dumps(username, ensure_ascii=False),
                "__BI_LOGIN_PATH_JSON__": json.dumps("/financial/bi-dashboard/login", ensure_ascii=False),
                "__BI_LOGOUT_PATH_JSON__": json.dumps("/financial/bi-dashboard/logout", ensure_ascii=False),
                "__BI_LOGO_WORDMARK__": dashboard_logo_wordmark_svg(),
                "__BI_LOGO_BADGE_SMALL__": dashboard_logo_badge_small_svg(),
                },
            ),
        )
    )


@router.get("/bi-dashboard/attendance", response_class=HTMLResponse)
async def bi_dashboard_attendance_entry(request: Request) -> Response:
    ensure_schema()
    username = current_dashboard_user(request)
    if not username:
        current_path = request.url.path
        if request.url.query:
            current_path = f"{current_path}?{request.url.query}"
        login_url = f"/financial/bi-dashboard/login?next={quote(current_path, safe='')}"
        return RedirectResponse(url=login_url, status_code=303)
    return HTMLResponse(
        render_template(
            "bi_attendance_entry.html",
            dashboard_page_context(
                "attendance",
                {
                "__BI_CURRENT_USER_JSON__": json.dumps(username, ensure_ascii=False),
                "__BI_DASHBOARD_PATH_JSON__": json.dumps("/financial/bi-dashboard", ensure_ascii=False),
                "__BI_LOGOUT_PATH_JSON__": json.dumps("/financial/bi-dashboard/logout", ensure_ascii=False),
                },
            ),
        )
    )


@router.get("/bi-dashboard/inventory-mappings", response_class=HTMLResponse)
async def bi_dashboard_inventory_mappings(request: Request) -> Response:
    ensure_schema()
    username = current_dashboard_user(request)
    if not username:
        current_path = request.url.path
        if request.url.query:
            current_path = f"{current_path}?{request.url.query}"
        login_url = f"/financial/bi-dashboard/login?next={quote(current_path, safe='')}"
        return RedirectResponse(url=login_url, status_code=303)
    return HTMLResponse(
        render_template(
            "bi_inventory_mapping_entry.html",
            dashboard_page_context(
                "inventory",
                {
                "__BI_CURRENT_USER_JSON__": json.dumps(username, ensure_ascii=False),
                "__BI_DASHBOARD_PATH_JSON__": json.dumps("/financial/bi-dashboard", ensure_ascii=False),
                "__BI_LOGOUT_PATH_JSON__": json.dumps("/financial/bi-dashboard/logout", ensure_ascii=False),
                },
            ),
        )
    )


@router.get("/bi-dashboard/metric-dictionary", response_class=HTMLResponse)
async def bi_dashboard_metric_dictionary(request: Request) -> Response:
    ensure_schema()
    username = current_dashboard_user(request)
    if not username:
        current_path = request.url.path
        if request.url.query:
            current_path = f"{current_path}?{request.url.query}"
        login_url = f"/financial/bi-dashboard/login?next={quote(current_path, safe='')}"
        return RedirectResponse(url=login_url, status_code=303)
    return HTMLResponse(
        render_template(
            "bi_metric_dictionary_entry.html",
            dashboard_page_context(
                "metric-dictionary",
                {
                    "__BI_CURRENT_USER_JSON__": json.dumps(username, ensure_ascii=False),
                    "__BI_DASHBOARD_PATH_JSON__": json.dumps(DASHBOARD_DEFAULT_PATH, ensure_ascii=False),
                    "__BI_LOGOUT_PATH_JSON__": json.dumps("/financial/bi-dashboard/logout", ensure_ascii=False),
                    "__BI_METRIC_DICTIONARY_API_PATH_JSON__": json.dumps(METRIC_DICTIONARY_API_PATH, ensure_ascii=False),
                },
            ),
        )
    )


@router.get("/bi-dashboard/master-data", response_class=HTMLResponse)
async def bi_dashboard_master_data(request: Request) -> Response:
    ensure_schema()
    username = current_dashboard_user(request)
    if not username:
        current_path = request.url.path
        if request.url.query:
            current_path = f"{current_path}?{request.url.query}"
        login_url = f"/financial/bi-dashboard/login?next={quote(current_path, safe='')}"
        return RedirectResponse(url=login_url, status_code=303)
    return HTMLResponse(
        render_template(
            "bi_master_data_entry.html",
            dashboard_page_context(
                "master-data",
                {
                    "__BI_CURRENT_USER_JSON__": json.dumps(username, ensure_ascii=False),
                    "__BI_DASHBOARD_PATH_JSON__": json.dumps(DASHBOARD_DEFAULT_PATH, ensure_ascii=False),
                    "__BI_LOGOUT_PATH_JSON__": json.dumps("/financial/bi-dashboard/logout", ensure_ascii=False),
                    "__BI_METRIC_DICTIONARY_PATH_JSON__": json.dumps(METRIC_DICTIONARY_ENTRY_PATH, ensure_ascii=False),
                    "__BI_MASTER_DATA_API_PATH_JSON__": json.dumps(MASTER_DATA_API_PATH, ensure_ascii=False),
                },
            ),
        )
    )


@router.get("/bi-dashboard/audit-logs", response_class=HTMLResponse)
async def bi_dashboard_audit_logs(request: Request) -> Response:
    ensure_schema()
    username = current_dashboard_user(request)
    if not username:
        current_path = request.url.path
        if request.url.query:
            current_path = f"{current_path}?{request.url.query}"
        login_url = f"/financial/bi-dashboard/login?next={quote(current_path, safe='')}"
        return RedirectResponse(url=login_url, status_code=303)
    return HTMLResponse(
        render_template(
            "bi_audit_log_center.html",
            dashboard_page_context(
                "audit-logs",
                {
                    "__BI_CURRENT_USER_JSON__": json.dumps(username, ensure_ascii=False),
                    "__BI_DASHBOARD_PATH_JSON__": json.dumps(DASHBOARD_DEFAULT_PATH, ensure_ascii=False),
                    "__BI_LOGOUT_PATH_JSON__": json.dumps("/financial/bi-dashboard/logout", ensure_ascii=False),
                    "__BI_MASTER_DATA_PATH_JSON__": json.dumps(MASTER_DATA_ENTRY_PATH, ensure_ascii=False),
                    "__BI_SYNC_SCHEDULE_PATH_JSON__": json.dumps("/financial/bi-dashboard/sync-schedule", ensure_ascii=False),
                    "__BI_AUDIT_LOG_API_PATH_JSON__": json.dumps(AUDIT_LOG_API_PATH, ensure_ascii=False),
                },
            ),
        )
    )


@router.get("/bi-dashboard/refurb-production", response_class=HTMLResponse)
async def bi_dashboard_refurb_production(request: Request) -> Response:
    ensure_schema()
    username = current_dashboard_user(request)
    if not username:
        current_path = request.url.path
        if request.url.query:
            current_path = f"{current_path}?{request.url.query}"
        login_url = f"/financial/bi-dashboard/login?next={quote(current_path, safe='')}"
        return RedirectResponse(url=login_url, status_code=303)
    return HTMLResponse(
        render_template(
            "bi_refurb_production_entry.html",
            dashboard_page_context(
                "refurb",
                {
                "__BI_CURRENT_USER_JSON__": json.dumps(username, ensure_ascii=False),
                "__BI_DASHBOARD_PATH_JSON__": json.dumps("/financial/bi-dashboard", ensure_ascii=False),
                "__BI_LOGOUT_PATH_JSON__": json.dumps("/financial/bi-dashboard/logout", ensure_ascii=False),
                },
            ),
        )
    )


@router.get("/bi-dashboard/forecast-alerts", response_class=HTMLResponse)
async def bi_dashboard_forecast_alerts(request: Request) -> Response:
    ensure_schema()
    username = current_dashboard_user(request)
    if not username:
        current_path = request.url.path
        if request.url.query:
            current_path = f"{current_path}?{request.url.query}"
        login_url = f"/financial/bi-dashboard/login?next={quote(current_path, safe='')}"
        return RedirectResponse(url=login_url, status_code=303)
    return HTMLResponse(
        render_template(
            "bi_forecast_alert_entry.html",
            dashboard_page_context(
                "forecast",
                {
                "__BI_CURRENT_USER_JSON__": json.dumps(username, ensure_ascii=False),
                "__BI_DASHBOARD_PATH_JSON__": json.dumps("/financial/bi-dashboard", ensure_ascii=False),
                "__BI_LOGOUT_PATH_JSON__": json.dumps("/financial/bi-dashboard/logout", ensure_ascii=False),
                },
            ),
        )
    )


@router.get("/bi-dashboard/sync-schedule", response_class=HTMLResponse)
async def bi_dashboard_sync_schedule(request: Request) -> Response:
    ensure_schema()
    username = current_dashboard_user(request)
    if not username:
        current_path = request.url.path
        if request.url.query:
            current_path = f"{current_path}?{request.url.query}"
        login_url = f"/financial/bi-dashboard/login?next={quote(current_path, safe='')}"
        return RedirectResponse(url=login_url, status_code=303)
    return HTMLResponse(
        render_template(
            "bi_sync_schedule_entry.html",
            dashboard_page_context(
                "sync",
                {
                "__BI_CURRENT_USER_JSON__": json.dumps(username, ensure_ascii=False),
                "__BI_DASHBOARD_PATH_JSON__": json.dumps("/financial/bi-dashboard", ensure_ascii=False),
                "__BI_LOGOUT_PATH_JSON__": json.dumps("/financial/bi-dashboard/logout", ensure_ascii=False),
                },
            ),
        )
    )


@router.get("/bi-dashboard/data-agent", response_class=HTMLResponse)
async def bi_dashboard_data_agent_entry(request: Request) -> Response:
    ensure_schema()
    username = current_dashboard_user(request)
    if not username:
        current_path = request.url.path
        if request.url.query:
            current_path = f"{current_path}?{request.url.query}"
        login_url = f"/financial/bi-dashboard/login?next={quote(current_path, safe='')}"
        return RedirectResponse(url=login_url, status_code=303)
    return HTMLResponse(
        render_template(
            "bi_data_agent_entry.html",
            dashboard_page_context(
                "data-agent",
                {
                "__BI_CURRENT_USER_JSON__": json.dumps(username, ensure_ascii=False),
                "__BI_DASHBOARD_PATH_JSON__": json.dumps(DASHBOARD_DEFAULT_PATH, ensure_ascii=False),
                "__BI_EDITOR_PATH_JSON__": json.dumps(DASHBOARD_EDITOR_PATH, ensure_ascii=False),
                "__BI_LOGOUT_PATH_JSON__": json.dumps("/financial/bi-dashboard/logout", ensure_ascii=False),
                "__BI_DATA_AGENT_STATUS_API_PATH_JSON__": json.dumps(DATA_AGENT_STATUS_API_PATH, ensure_ascii=False),
                "__BI_DATA_AGENT_CHAT_API_PATH_JSON__": json.dumps(DATA_AGENT_CHAT_API_PATH, ensure_ascii=False),
                "__BI_DATA_AGENT_REPORTS_API_PATH_JSON__": json.dumps(DATA_AGENT_REPORTS_API_PATH, ensure_ascii=False),
                "__BI_DATA_AGENT_REPORT_GENERATE_API_PATH_JSON__": json.dumps(DATA_AGENT_REPORT_GENERATE_API_PATH, ensure_ascii=False),
                "__BI_LOGO_WORDMARK__": dashboard_logo_wordmark_svg(),
                "__BI_LOGO_BADGE_SMALL__": dashboard_logo_badge_small_svg(),
                },
            ),
        )
    )


@router.get("/bi-dashboard/api/data-agent/status")
async def bi_dashboard_data_agent_status(_auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    return JSONResponse(data_agent_status_payload())


@router.post("/bi-dashboard/api/data-agent/chat")
async def bi_dashboard_data_agent_chat(
    payload: Dict[str, Any] = Body(default={}),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    message = str(payload.get("message") or "").strip()
    session_id = str(payload.get("session_id") or "").strip() or data_agent_session_id()
    if not message:
        raise HTTPException(status_code=400, detail="请输入问题")
    if not data_agent_env_snapshot().get("openai_api_key"):
        raise HTTPException(status_code=400, detail="数据分析 Agent 尚未配置 OPENAI_API_KEY，当前无法进行问答")
    try:
        result = call_data_agent_chat(message, session_id=session_id, timeout=240)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    return JSONResponse(
        {
            "session_id": result["session_id"],
            "answer": result["content"],
            "source": "data-agent",
        }
    )


@router.get("/bi-dashboard/api/data-agent/reports")
async def bi_dashboard_data_agent_reports(
    limit: int = Query(20, ge=1, le=100),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        rows = list_data_agent_reports(conn, limit=limit)
    return JSONResponse({"items": rows})


@router.post("/bi-dashboard/api/data-agent/reports/generate")
async def bi_dashboard_data_agent_generate_report(
    payload: Dict[str, Any] = Body(default={}),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    report_type = str(payload.get("report_type") or "weekly").strip().lower()
    if report_type not in {"weekly", "monthly"}:
        raise HTTPException(status_code=400, detail="浠呮敮鎸?weekly 鎴?monthly")
    item = generate_data_agent_report(report_type, "manual")
    return JSONResponse({"item": item})


@router.get("/bi-dashboard/api/sync-schedule")
async def get_sync_schedule(_auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        schedule = load_sync_schedule(conn)
    payload = serialize_sync_schedule(schedule)
    payload["mode_options"] = [
        {"value": "all", "label": "库存 + 销售"},
        {"value": "inventory", "label": "仅库存原始数据"},
        {"value": "sales", "label": "仅销售原始数据"},
    ]
    payload["status_options"] = [
        {"value": "idle", "label": "未执行"},
        {"value": "running", "label": "执行中"},
        {"value": "success", "label": "成功"},
        {"value": "failed", "label": "失败"},
        {"value": "skipped", "label": "已跳过"},
    ]
    return JSONResponse(payload)


@router.put("/bi-dashboard/api/sync-schedule")
async def save_sync_schedule(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    normalized = validate_sync_schedule_payload(payload or {})
    current_engine = get_engine()
    with current_engine.begin() as conn:
        ensure_sync_schedule_seed(conn)
        conn.execute(
            text(
                """
                UPDATE bi_raw_sync_schedule_config
                SET
                    is_enabled = :is_enabled,
                    mode = :mode,
                    cron_expr = :cron_expr,
                    sales_days_behind = :sales_days_behind,
                    sales_window_days = :sales_window_days,
                    snapshot_days_behind = :snapshot_days_behind,
                    updated_by = :updated_by
                WHERE schedule_key = :schedule_key
                """
            ),
            {
                "schedule_key": SYNC_SCHEDULE_KEY,
                "is_enabled": 1 if normalized["is_enabled"] else 0,
                "mode": normalized["mode"],
                "cron_expr": normalized["cron_expr"],
                "sales_days_behind": normalized["sales_days_behind"],
                "sales_window_days": normalized["sales_window_days"],
                "snapshot_days_behind": normalized["snapshot_days_behind"],
                "updated_by": username,
            },
        )
        schedule = load_sync_schedule(conn)
    refreshed = refresh_sync_scheduler()
    record_dashboard_audit(
        module_key="sync",
        module_name="鍚屾璋冨害",
        action_key="schedule.save",
        action_name="淇濆瓨鍘熷鍚屾璁″垝",
        target_type="schedule",
        target_id=SYNC_SCHEDULE_KEY,
        target_name="鍘熷鏁版嵁鍚屾璁″垝",
        detail_summary=f"{normalized['mode']} / {normalized['cron_expr']}",
        detail={
            "is_enabled": normalized["is_enabled"],
            "mode": normalized["mode"],
            "cron_expr": normalized["cron_expr"],
            "sales_days_behind": normalized["sales_days_behind"],
            "sales_window_days": normalized["sales_window_days"],
            "snapshot_days_behind": normalized["snapshot_days_behind"],
            "runtime": refreshed,
        },
        triggered_by=username,
        source_path="/financial/bi-dashboard/api/sync-schedule",
        source_method="PUT",
        affected_count=1,
    )
    return JSONResponse(
        {
            "message": "瀹氭椂閰嶇疆宸蹭繚瀛樺苟鐢熸晥",
            "schedule": {**schedule, **sync_scheduler_snapshot()},
            "runtime": refreshed,
        }
    )


@router.post("/bi-dashboard/api/sync-schedule/run-now")
async def run_sync_schedule_now(username: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    if sync_run_lock.locked():
        raise HTTPException(status_code=409, detail="已有原始数据同步任务在执行，请稍后再试")

    def run_in_background() -> None:
        try:
            execute_raw_sync("manual")
        except Exception:
            pass

    threading.Thread(target=run_in_background, name="bi-raw-sync-manual", daemon=True).start()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        schedule = load_sync_schedule(conn)
    record_dashboard_audit(
        module_key="sync",
        module_name="同步调度",
        action_key="schedule.run_now",
        action_name="手动触发原始同步",
        target_type="schedule",
        target_id=SYNC_SCHEDULE_KEY,
        target_name="原始数据同步计划",
        result_status="submitted",
        detail_summary=f"已提交 {schedule.get('mode') or 'all'} 模式同步",
        detail={"schedule": serialize_sync_schedule(schedule)},
        triggered_by=username,
        source_path="/financial/bi-dashboard/api/sync-schedule/run-now",
        source_method="POST",
        affected_count=1,
    )
    return JSONResponse(
        {
            "message": "已提交原始数据同步任务，请稍后刷新查看执行结果",
            "schedule": serialize_sync_schedule(schedule),
        },
        status_code=202,
    )


@router.get("/bi-dashboard/api/refurb-production")
async def list_refurb_production(
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    limit: int = Query(200, ge=1, le=1000),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    start = parse_date_or_none(start_date)
    end = parse_date_or_none(end_date)
    current_engine = get_engine()
    with current_engine.connect() as conn:
        rows = refurb_production_summaries(conn, start_date=start, end_date=end, limit=limit)
    total_final_good_qty = sum(float(row.get("final_good_qty") or 0) for row in rows)
    total_plan_qty = sum(float(row.get("plan_qty") or 0) for row in rows)
    return JSONResponse(
        {
            "rows": rows,
            "summary": {
                "row_count": len(rows),
                "total_final_good_qty": total_final_good_qty,
                "total_plan_qty": total_plan_qty,
            },
        }
    )


@router.post("/bi-dashboard/api/refurb-production")
async def upsert_refurb_production(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    normalized = normalize_refurb_production_payload(payload)
    current_engine = get_engine()
    saved_rows = save_refurb_production_rows(current_engine, [normalized], username)
    with current_engine.connect() as conn:
        rows = refurb_production_summaries(
            conn,
            start_date=normalized["biz_date"],
            end_date=normalized["biz_date"],
            limit=500,
        )
    current_row = next(
        (
            row
            for row in rows
            if row["biz_date"] == normalized["biz_date"].isoformat()
            and row["refurb_category"] == normalized["refurb_category"]
            and row["material_name"] == normalized["material_name"]
        ),
        None,
    )
    record_dashboard_audit(
        module_key="refurb",
        module_name="缈绘柊鐢熶骇",
        action_key="daily.upsert",
        action_name="保存翻新生产日报",
        target_type="refurb_daily",
        target_id=f"{normalized['biz_date'].isoformat()}|{normalized['refurb_category']}|{normalized['material_name']}",
        target_name=f"{normalized['biz_date'].isoformat()} / {normalized['refurb_category']} / {normalized['material_name']}",
        detail_summary=f"写入 {saved_rows} 行翻新生产数据",
        detail={
            "biz_date": normalized["biz_date"],
            "refurb_category": normalized["refurb_category"],
            "material_name": normalized["material_name"],
            "saved_rows": saved_rows,
        },
        triggered_by=username,
        source_path="/financial/bi-dashboard/api/refurb-production",
        source_method="POST",
        affected_count=saved_rows,
    )
    return JSONResponse(
        {
            "message": "翻新生产数据已保存",
            "saved_rows": saved_rows,
            "row": current_row or {
                "biz_date": normalized["biz_date"].isoformat(),
                "refurb_category": normalized["refurb_category"],
                "material_name": normalized["material_name"],
            },
        }
    )


@router.post("/bi-dashboard/api/refurb-production/import")
async def import_refurb_production(
    file: UploadFile = File(...),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    if not str(file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="浠呮敮鎸佸鍏?.xlsx 鏂囦欢")
    content = await file.read()
    rows = parse_refurb_excel_rows(file, content)
    current_engine = get_engine()
    saved_rows = save_refurb_production_rows(current_engine, rows, username)
    date_values = [row["biz_date"] for row in rows]
    with current_engine.connect() as conn:
        refreshed_rows = refurb_production_summaries(
            conn,
            start_date=min(date_values),
            end_date=max(date_values),
            limit=1000,
        )
    record_dashboard_audit(
        module_key="refurb",
        module_name="缈绘柊鐢熶骇",
        action_key="daily.import_excel",
        action_name="导入翻新生产 Excel",
        target_type="refurb_import",
        target_id=str(file.filename or ""),
        target_name=str(file.filename or "refurb-production.xlsx"),
        detail_summary=f"Excel 导入 {saved_rows} 行",
        detail={
            "filename": file.filename or "",
            "saved_rows": saved_rows,
            "start_date": min(date_values),
            "end_date": max(date_values),
            "categories": audit_preview_values(rows, "refurb_category"),
            "materials": audit_preview_values(rows, "material_name"),
        },
        triggered_by=username,
        source_path="/financial/bi-dashboard/api/refurb-production/import",
        source_method="POST",
        affected_count=saved_rows,
    )
    return JSONResponse(
        {
            "message": f"Excel 导入完成，共写入 {saved_rows} 行",
            "saved_rows": saved_rows,
            "date_range": {
                "start_date": min(date_values).isoformat(),
                "end_date": max(date_values).isoformat(),
            },
            "rows": refreshed_rows,
        }
    )


@router.get("/bi-dashboard/api/forecast-alerts/overview")
async def forecast_alert_overview(_auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        profiles = list_forecast_profiles(conn)
        events = list_promotion_events(conn)
        alerts = list_inventory_alerts(conn, limit=50)
        today = date.today()
        forecasts = list_ai_forecasts(conn, start_date=today, end_date=today + timedelta(days=14), limit=300)
        latest_forecast_date = conn.execute(text("SELECT MAX(forecast_date) FROM bi_sales_forecast_ai_daily")).scalar()
        latest_alert_date = conn.execute(text("SELECT MAX(snapshot_date) FROM bi_inventory_alert_log")).scalar()
    return JSONResponse(
        {
            "profiles": profiles,
            "events": events,
            "alerts": alerts,
            "forecasts": forecasts,
            "summary": {
                "profile_count": len(profiles),
                "event_count": len(events),
                "alert_count": len(alerts),
                "forecast_count": len(forecasts),
                "latest_forecast_date": _plain_date_value(latest_forecast_date),
                "latest_alert_date": _plain_date_value(latest_alert_date),
            },
        }
    )


def _plain_date_value(value: Any) -> Any:
    return value.isoformat() if isinstance(value, (date, datetime)) else value


@router.get("/bi-dashboard/api/forecast-alerts/manual")
async def list_forecast_manual(
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    limit: int = Query(300, ge=1, le=1000),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        rows = list_forecast_manual_rows(
            conn,
            start_date=parse_date_or_none(start_date),
            end_date=parse_date_or_none(end_date),
            limit=limit,
        )
    return JSONResponse({"rows": rows})


@router.post("/bi-dashboard/api/forecast-alerts/manual")
async def upsert_forecast_manual(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    normalized = normalize_forecast_manual_payload(payload)
    current_engine = get_engine()
    save_manual_forecast(current_engine, normalized, username)
    recalculate_forecasts_and_alerts(current_engine, updated_by=username, send_notifications=False)
    with current_engine.connect() as conn:
        rows = list_forecast_manual_rows(
            conn,
            start_date=normalized["forecast_date"],
            end_date=normalized["forecast_date"],
            limit=100,
        )
    current_row = next(
        (
            row
            for row in rows
            if row["forecast_date"] == normalized["forecast_date"].isoformat()
            and row["material_name"] == normalized["material_name"]
            and row["demand_type"] == normalized["demand_type"]
        ),
        None,
    )
    record_dashboard_audit(
        module_key="forecast",
        module_name="棰勬祴棰勮",
        action_key="manual.upsert",
        action_name="保存手动预测",
        target_type="manual_forecast",
        target_id=f"{normalized['forecast_date'].isoformat()}|{normalized['material_name']}|{normalized['demand_type']}",
        target_name=f"{normalized['forecast_date'].isoformat()} / {normalized['material_name']} / {normalized['demand_type']}",
        detail_summary="保存手动预测并完成预测重算",
        detail={
            "forecast_date": normalized["forecast_date"],
            "material_name": normalized["material_name"],
            "demand_type": normalized["demand_type"],
            "manual_qty": normalized.get("manual_qty"),
        },
        triggered_by=username,
        source_path="/financial/bi-dashboard/api/forecast-alerts/manual",
        source_method="POST",
        affected_count=1,
    )
    return JSONResponse({"message": "手动预测已保存并重算 AI 预测", "row": current_row})


@router.get("/bi-dashboard/api/forecast-alerts/events")
async def get_forecast_events(_auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        rows = list_promotion_events(conn)
    return JSONResponse({"rows": rows})


@router.put("/bi-dashboard/api/forecast-alerts/events")
async def put_forecast_events(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    rows = normalize_promotion_event_payload(payload.get("rows"))
    current_engine = get_engine()
    save_promotion_events(current_engine, rows)
    result = recalculate_forecasts_and_alerts(current_engine, updated_by=username, send_notifications=False)
    record_dashboard_audit(
        module_key="forecast",
        module_name="预测预警",
        action_key="events.bulk_save",
        action_name="保存促销事件",
        target_type="promotion_events",
        target_name="促销事件批量维护",
        detail_summary=f"保存 {len(rows)} 条促销事件并完成重算",
        detail={
            "row_count": len(rows),
            "event_names": audit_preview_values(rows, "event_name"),
            "result": result,
        },
        triggered_by=username,
        source_path="/financial/bi-dashboard/api/forecast-alerts/events",
        source_method="PUT",
        affected_count=len(rows),
    )
    return JSONResponse({"message": "促销事件已保存并完成预测重算", "result": result})


@router.post("/bi-dashboard/api/forecast-alerts/recalculate")
async def post_forecast_recalculate(username: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    result = recalculate_forecasts_and_alerts(current_engine, updated_by=username, send_notifications=True)
    record_dashboard_audit(
        module_key="forecast",
        module_name="预测预警",
        action_key="recalculate.run",
        action_name="重算预测与预警",
        target_type="forecast_engine",
        target_name="预测与安全库存预警",
        detail_summary="执行预测与预警重算",
        detail={"result": result, "send_notifications": True},
        triggered_by=username,
        source_path="/financial/bi-dashboard/api/forecast-alerts/recalculate",
        source_method="POST",
        affected_count=int(result.get("saved_rows") or result.get("forecast_count") or 0) if isinstance(result, dict) else 0,
    )
    return JSONResponse({"message": "预测与安全库存预警已重算", "result": result})


@router.get("/bi-dashboard/api/inventory-mappings")
async def list_inventory_mappings(_auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        warehouse_rows = conn.execute(
            text(
                """
                SELECT id, source_warehouse_name, warehouse_name_clean, sort_order, is_enabled, created_at, updated_at
                FROM bi_inventory_warehouse_map
                ORDER BY sort_order, id
                """
            )
        ).mappings().all()
        status_rows = conn.execute(
            text(
                """
                SELECT id, stock_status_id, stock_status_name, sort_order, is_enabled, created_at, updated_at
                FROM bi_inventory_status_map
                ORDER BY sort_order, id
                """
            )
        ).mappings().all()
        latest_cleaning = to_plain(latest_date(conn, "inventory_cleaning"))
        latest_raw = to_plain(latest_date(conn, "inventory"))
    return JSONResponse(
        {
            "warehouses": [
                {
                    "id": int(row["id"]),
                    "source_warehouse_name": row["source_warehouse_name"],
                    "warehouse_name_clean": row["warehouse_name_clean"],
                    "sort_order": int(row["sort_order"] or 0),
                    "is_enabled": bool(row["is_enabled"]),
                    "created_at": to_plain(row["created_at"]),
                    "updated_at": to_plain(row["updated_at"]),
                }
                for row in warehouse_rows
            ],
            "statuses": [
                {
                    "id": int(row["id"]),
                    "stock_status_id": row["stock_status_id"],
                    "stock_status_name": row["stock_status_name"],
                    "sort_order": int(row["sort_order"] or 0),
                    "is_enabled": bool(row["is_enabled"]),
                    "created_at": to_plain(row["created_at"]),
                    "updated_at": to_plain(row["updated_at"]),
                }
                for row in status_rows
            ],
            "latest_cleaning_date": latest_cleaning,
            "latest_raw_date": latest_raw,
        }
    )


@router.put("/bi-dashboard/api/inventory-mappings")
async def save_inventory_mappings(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    warehouse_rows = normalize_inventory_mapping_payload(payload.get("warehouses"), mapping_type="warehouse")
    status_rows = normalize_inventory_mapping_payload(payload.get("statuses"), mapping_type="status")
    current_engine = get_engine()
    with current_engine.begin() as conn:
        existing_warehouse_ids = {
            int(row[0])
            for row in conn.execute(text("SELECT id FROM bi_inventory_warehouse_map")).fetchall()
        }
        existing_status_ids = {
            int(row[0])
            for row in conn.execute(text("SELECT id FROM bi_inventory_status_map")).fetchall()
        }
        submitted_warehouse_ids = {item["id"] for item in warehouse_rows if item["id"] > 0}
        submitted_status_ids = {item["id"] for item in status_rows if item["id"] > 0}

        for item in warehouse_rows:
            if item["id"] > 0:
                result = conn.execute(
                    text(
                        """
                        UPDATE bi_inventory_warehouse_map
                        SET
                            source_warehouse_name = :source_warehouse_name,
                            warehouse_name_clean = :warehouse_name_clean,
                            sort_order = :sort_order,
                            is_enabled = :is_enabled
                        WHERE id = :id
                        """
                    ),
                    item,
                )
                if result.rowcount:
                    continue
            conn.execute(
                text(
                    """
                    INSERT INTO bi_inventory_warehouse_map(
                        source_warehouse_name, warehouse_name_clean, sort_order, is_enabled
                    ) VALUES (
                        :source_warehouse_name, :warehouse_name_clean, :sort_order, :is_enabled
                    )
                    """
                ),
                {key: item[key] for key in ("source_warehouse_name", "warehouse_name_clean", "sort_order", "is_enabled")},
            )

        for item in status_rows:
            if item["id"] > 0:
                result = conn.execute(
                    text(
                        """
                        UPDATE bi_inventory_status_map
                        SET
                            stock_status_id = :stock_status_id,
                            stock_status_name = :stock_status_name,
                            sort_order = :sort_order,
                            is_enabled = :is_enabled
                        WHERE id = :id
                        """
                    ),
                    item,
                )
                if result.rowcount:
                    continue
            conn.execute(
                text(
                    """
                    INSERT INTO bi_inventory_status_map(
                        stock_status_id, stock_status_name, sort_order, is_enabled
                    ) VALUES (
                        :stock_status_id, :stock_status_name, :sort_order, :is_enabled
                    )
                    """
                ),
                {key: item[key] for key in ("stock_status_id", "stock_status_name", "sort_order", "is_enabled")},
            )

        disabled_warehouse_ids = sorted(existing_warehouse_ids - submitted_warehouse_ids)
        disabled_status_ids = sorted(existing_status_ids - submitted_status_ids)
        if disabled_warehouse_ids:
            placeholders = ", ".join(f":wid_{idx}" for idx, _ in enumerate(disabled_warehouse_ids))
            params = {f"wid_{idx}": row_id for idx, row_id in enumerate(disabled_warehouse_ids)}
            conn.execute(
                text(f"UPDATE bi_inventory_warehouse_map SET is_enabled = 0 WHERE id IN ({placeholders})"),
                params,
            )
        if disabled_status_ids:
            placeholders = ", ".join(f":sid_{idx}" for idx, _ in enumerate(disabled_status_ids))
            params = {f"sid_{idx}": row_id for idx, row_id in enumerate(disabled_status_ids)}
            conn.execute(
                text(f"UPDATE bi_inventory_status_map SET is_enabled = 0 WHERE id IN ({placeholders})"),
                params,
            )

    refreshed_rows = refresh_inventory_cleaning(current_engine)
    with current_engine.connect() as conn:
        latest_cleaning = to_plain(latest_date(conn, "inventory_cleaning"))
    record_dashboard_audit(
        module_key="governance",
        module_name="治理中心",
        action_key="inventory_mappings.save",
        action_name="保存库存映射",
        target_type="inventory_mappings",
        target_name="仓库与状态映射",
        detail_summary=f"保存 {len(warehouse_rows)} 条仓库映射、{len(status_rows)} 条状态映射",
        detail={
            "warehouse_count": len(warehouse_rows),
            "status_count": len(status_rows),
            "warehouse_names": audit_preview_values(warehouse_rows, "warehouse_name_clean"),
            "status_names": audit_preview_values(status_rows, "stock_status_name"),
            "refreshed_rows": int(refreshed_rows),
            "latest_cleaning_date": latest_cleaning,
        },
        triggered_by=username,
        source_path="/financial/bi-dashboard/api/inventory-mappings",
        source_method="PUT",
        affected_count=len(warehouse_rows) + len(status_rows),
    )
    return JSONResponse(
        {
            "saved": True,
            "warehouse_count": len(warehouse_rows),
            "status_count": len(status_rows),
            "refreshed_rows": int(refreshed_rows),
            "latest_cleaning_date": latest_cleaning,
        }
    )


@router.get("/bi-dashboard/api/metric-dictionary")
async def list_metric_dictionary(_auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT
                    id, metric_key, metric_name, business_domain, owner_role,
                    definition_text, formula_text, source_table, source_fields,
                    dimension_notes, version_tag, effective_date, sort_order,
                    is_enabled, created_by, updated_by, created_at, updated_at
                FROM bi_metric_dictionary
                ORDER BY sort_order, id
                """
            )
        ).mappings().all()
    items = [
        {
            "id": int(row["id"]),
            "metric_key": str(row["metric_key"] or ""),
            "metric_name": str(row["metric_name"] or ""),
            "business_domain": str(row["business_domain"] or ""),
            "owner_role": str(row["owner_role"] or ""),
            "definition_text": str(row["definition_text"] or ""),
            "formula_text": str(row["formula_text"] or ""),
            "source_table": str(row["source_table"] or ""),
            "source_fields": str(row["source_fields"] or ""),
            "dimension_notes": str(row["dimension_notes"] or ""),
            "version_tag": str(row["version_tag"] or "v1"),
            "effective_date": to_plain(row["effective_date"]),
            "sort_order": int(row["sort_order"] or 0),
            "is_enabled": bool(row["is_enabled"]),
            "created_by": to_plain(row["created_by"]),
            "updated_by": to_plain(row["updated_by"]),
            "created_at": to_plain(row["created_at"]),
            "updated_at": to_plain(row["updated_at"]),
        }
        for row in rows
    ]
    active_count = sum(1 for item in items if item["is_enabled"])
    domain_count = len({item["business_domain"] for item in items if item["business_domain"]})
    latest_updated_at = max((item["updated_at"] for item in items if item["updated_at"]), default=None)
    return JSONResponse(
        {
            "items": items,
            "summary": {
                "total_count": len(items),
                "active_count": active_count,
                "domain_count": domain_count,
                "latest_updated_at": latest_updated_at,
            },
        }
    )


@router.put("/bi-dashboard/api/metric-dictionary")
async def save_metric_dictionary(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    rows = normalize_metric_dictionary_payload(payload.get("items"))
    current_engine = get_engine()
    with current_engine.begin() as conn:
        existing_ids = {
            int(row[0])
            for row in conn.execute(text("SELECT id FROM bi_metric_dictionary")).fetchall()
        }
        submitted_ids = {item["id"] for item in rows if item["id"] > 0}
        for item in rows:
            record = {**item, "updated_by": username}
            if item["id"] > 0:
                result = conn.execute(
                    text(
                        """
                        UPDATE bi_metric_dictionary
                        SET
                            metric_key = :metric_key,
                            metric_name = :metric_name,
                            business_domain = :business_domain,
                            owner_role = :owner_role,
                            definition_text = :definition_text,
                            formula_text = :formula_text,
                            source_table = :source_table,
                            source_fields = :source_fields,
                            dimension_notes = :dimension_notes,
                            version_tag = :version_tag,
                            effective_date = :effective_date,
                            sort_order = :sort_order,
                            is_enabled = :is_enabled,
                            updated_by = :updated_by
                        WHERE id = :id
                        """
                    ),
                    record,
                )
                if result.rowcount:
                    continue
            record["created_by"] = username
            conn.execute(
                text(
                    """
                    INSERT INTO bi_metric_dictionary(
                        metric_key, metric_name, business_domain, owner_role,
                        definition_text, formula_text, source_table, source_fields,
                        dimension_notes, version_tag, effective_date, sort_order,
                        is_enabled, created_by, updated_by
                    ) VALUES (
                        :metric_key, :metric_name, :business_domain, :owner_role,
                        :definition_text, :formula_text, :source_table, :source_fields,
                        :dimension_notes, :version_tag, :effective_date, :sort_order,
                        :is_enabled, :created_by, :updated_by
                    )
                    """
                ),
                record,
            )
        disabled_ids = sorted(existing_ids - submitted_ids)
        if disabled_ids:
            placeholders = ", ".join(f":mid_{idx}" for idx, _ in enumerate(disabled_ids))
            params = {f"mid_{idx}": row_id for idx, row_id in enumerate(disabled_ids)}
            params["updated_by"] = username
            conn.execute(
                text(
                    f"""
                    UPDATE bi_metric_dictionary
                    SET is_enabled = 0, updated_by = :updated_by
                    WHERE id IN ({placeholders})
                    """
                ),
                params,
            )
    record_dashboard_audit(
        module_key="governance",
        module_name="治理中心",
        action_key="metric_dictionary.save",
        action_name="保存指标口径",
        target_type="metric_dictionary",
        target_name="指标口径中心",
        detail_summary=f"保存 {len(rows)} 条指标口径",
        detail={
            "item_count": len(rows),
            "metric_keys": audit_preview_values(rows, "metric_key"),
            "domains": audit_preview_values(rows, "business_domain"),
        },
        triggered_by=username,
        source_path="/financial/bi-dashboard/api/metric-dictionary",
        source_method="PUT",
        affected_count=len(rows),
    )
    return JSONResponse({"saved": True, "item_count": len(rows)})


@router.get("/bi-dashboard/api/master-data")
async def list_master_data(_auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        sku_rows = conn.execute(
            text(
                """
                SELECT
                    id, sku_code, sku_name, sku_type, product_line, model, spec_version,
                    lifecycle_status, owner_dept, sort_order, is_active, created_at, updated_at
                FROM bi_sku_master
                ORDER BY sort_order, id
                """
            )
        ).mappings().all()
        warehouse_rows = conn.execute(
            text(
                """
                SELECT
                    id, source_warehouse_name, warehouse_name_clean, warehouse_code,
                    warehouse_type, platform_owner, city, is_sellable_warehouse,
                    is_reverse_warehouse, sort_order, is_enabled, created_at, updated_at
                FROM bi_inventory_warehouse_map
                ORDER BY sort_order, id
                """
            )
        ).mappings().all()
        status_rows = conn.execute(
            text(
                """
                SELECT
                    id, stock_status_id, stock_status_name, status_group, can_sell,
                    can_forecast_supply, need_quality_check, next_default_status,
                    sort_order, is_enabled, created_at, updated_at
                FROM bi_inventory_status_map
                ORDER BY sort_order, id
                """
            )
        ).mappings().all()
        channel_rows = conn.execute(
            text(
                """
                SELECT
                    id, channel_code, channel_name, shop_name, platform_name,
                    owner_dept, sort_order, is_active, created_at, updated_at
                FROM bi_channel_shop_master
                ORDER BY sort_order, id
                """
            )
        ).mappings().all()
        latest_cleaning = to_plain(latest_date(conn, "inventory_cleaning"))
        latest_sales = to_plain(latest_date(conn, "sales"))
    return JSONResponse(
        {
            "skus": [{key: to_plain(value) for key, value in row.items()} for row in sku_rows],
            "warehouses": [{key: to_plain(value) for key, value in row.items()} for row in warehouse_rows],
            "statuses": [{key: to_plain(value) for key, value in row.items()} for row in status_rows],
            "channels": [{key: to_plain(value) for key, value in row.items()} for row in channel_rows],
            "summary": {
                "sku_count": len(sku_rows),
                "warehouse_count": len(warehouse_rows),
                "status_count": len(status_rows),
                "channel_count": len(channel_rows),
                "latest_inventory_cleaning_date": latest_cleaning,
                "latest_sales_date": latest_sales,
            },
        }
    )


@router.put("/bi-dashboard/api/master-data")
async def save_master_data(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    sku_rows = normalize_sku_master_payload(payload.get("skus"))
    warehouse_rows = normalize_master_warehouse_payload(payload.get("warehouses"))
    status_rows = normalize_master_status_payload(payload.get("statuses"))
    channel_rows = normalize_channel_shop_payload(payload.get("channels"))
    current_engine = get_engine()
    with current_engine.begin() as conn:
        existing_sku_ids = {int(row[0]) for row in conn.execute(text("SELECT id FROM bi_sku_master")).fetchall()}
        existing_warehouse_ids = {int(row[0]) for row in conn.execute(text("SELECT id FROM bi_inventory_warehouse_map")).fetchall()}
        existing_status_ids = {int(row[0]) for row in conn.execute(text("SELECT id FROM bi_inventory_status_map")).fetchall()}
        existing_channel_ids = {int(row[0]) for row in conn.execute(text("SELECT id FROM bi_channel_shop_master")).fetchall()}

        submitted_sku_ids = {item["id"] for item in sku_rows if item["id"] > 0}
        submitted_warehouse_ids = {item["id"] for item in warehouse_rows if item["id"] > 0}
        submitted_status_ids = {item["id"] for item in status_rows if item["id"] > 0}
        submitted_channel_ids = {item["id"] for item in channel_rows if item["id"] > 0}

        for item in sku_rows:
            record = {**item, "updated_by": username}
            if item["id"] > 0:
                result = conn.execute(
                    text(
                        """
                        UPDATE bi_sku_master
                        SET
                            sku_code = :sku_code,
                            sku_name = :sku_name,
                            sku_type = :sku_type,
                            product_line = :product_line,
                            model = :model,
                            spec_version = :spec_version,
                            lifecycle_status = :lifecycle_status,
                            owner_dept = :owner_dept,
                            sort_order = :sort_order,
                            is_active = :is_active,
                            updated_by = :updated_by
                        WHERE id = :id
                        """
                    ),
                    record,
                )
                if result.rowcount:
                    continue
            record["created_by"] = username
            conn.execute(
                text(
                    """
                    INSERT INTO bi_sku_master(
                        sku_code, sku_name, sku_type, product_line, model, spec_version,
                        lifecycle_status, owner_dept, sort_order, is_active, created_by, updated_by
                    ) VALUES (
                        :sku_code, :sku_name, :sku_type, :product_line, :model, :spec_version,
                        :lifecycle_status, :owner_dept, :sort_order, :is_active, :created_by, :updated_by
                    )
                    """
                ),
                record,
            )

        for item in warehouse_rows:
            if item["id"] > 0:
                result = conn.execute(
                    text(
                        """
                        UPDATE bi_inventory_warehouse_map
                        SET
                            source_warehouse_name = :source_warehouse_name,
                            warehouse_name_clean = :warehouse_name_clean,
                            warehouse_code = :warehouse_code,
                            warehouse_type = :warehouse_type,
                            platform_owner = :platform_owner,
                            city = :city,
                            is_sellable_warehouse = :is_sellable_warehouse,
                            is_reverse_warehouse = :is_reverse_warehouse,
                            sort_order = :sort_order,
                            is_enabled = :is_enabled
                        WHERE id = :id
                        """
                    ),
                    item,
                )
                if result.rowcount:
                    continue
            conn.execute(
                text(
                    """
                    INSERT INTO bi_inventory_warehouse_map(
                        source_warehouse_name, warehouse_name_clean, warehouse_code, warehouse_type,
                        platform_owner, city, is_sellable_warehouse, is_reverse_warehouse, sort_order, is_enabled
                    ) VALUES (
                        :source_warehouse_name, :warehouse_name_clean, :warehouse_code, :warehouse_type,
                        :platform_owner, :city, :is_sellable_warehouse, :is_reverse_warehouse, :sort_order, :is_enabled
                    )
                    """
                ),
                item,
            )

        for item in status_rows:
            if item["id"] > 0:
                result = conn.execute(
                    text(
                        """
                        UPDATE bi_inventory_status_map
                        SET
                            stock_status_id = :stock_status_id,
                            stock_status_name = :stock_status_name,
                            status_group = :status_group,
                            can_sell = :can_sell,
                            can_forecast_supply = :can_forecast_supply,
                            need_quality_check = :need_quality_check,
                            next_default_status = :next_default_status,
                            sort_order = :sort_order,
                            is_enabled = :is_enabled
                        WHERE id = :id
                        """
                    ),
                    item,
                )
                if result.rowcount:
                    continue
            conn.execute(
                text(
                    """
                    INSERT INTO bi_inventory_status_map(
                        stock_status_id, stock_status_name, status_group, can_sell,
                        can_forecast_supply, need_quality_check, next_default_status, sort_order, is_enabled
                    ) VALUES (
                        :stock_status_id, :stock_status_name, :status_group, :can_sell,
                        :can_forecast_supply, :need_quality_check, :next_default_status, :sort_order, :is_enabled
                    )
                    """
                ),
                item,
            )

        for item in channel_rows:
            record = {**item, "updated_by": username}
            if item["id"] > 0:
                result = conn.execute(
                    text(
                        """
                        UPDATE bi_channel_shop_master
                        SET
                            channel_code = :channel_code,
                            channel_name = :channel_name,
                            shop_name = :shop_name,
                            platform_name = :platform_name,
                            owner_dept = :owner_dept,
                            sort_order = :sort_order,
                            is_active = :is_active,
                            updated_by = :updated_by
                        WHERE id = :id
                        """
                    ),
                    record,
                )
                if result.rowcount:
                    continue
            record["created_by"] = username
            conn.execute(
                text(
                    """
                    INSERT INTO bi_channel_shop_master(
                        channel_code, channel_name, shop_name, platform_name,
                        owner_dept, sort_order, is_active, created_by, updated_by
                    ) VALUES (
                        :channel_code, :channel_name, :shop_name, :platform_name,
                        :owner_dept, :sort_order, :is_active, :created_by, :updated_by
                    )
                    """
                ),
                record,
            )

        disabled_sku_ids = sorted(existing_sku_ids - submitted_sku_ids)
        disabled_warehouse_ids = sorted(existing_warehouse_ids - submitted_warehouse_ids)
        disabled_status_ids = sorted(existing_status_ids - submitted_status_ids)
        disabled_channel_ids = sorted(existing_channel_ids - submitted_channel_ids)
        if disabled_sku_ids:
            placeholders = ", ".join(f":sku_{idx}" for idx, _ in enumerate(disabled_sku_ids))
            params = {f"sku_{idx}": row_id for idx, row_id in enumerate(disabled_sku_ids)}
            conn.execute(text(f"UPDATE bi_sku_master SET is_active = 0 WHERE id IN ({placeholders})"), params)
        if disabled_warehouse_ids:
            placeholders = ", ".join(f":wid_{idx}" for idx, _ in enumerate(disabled_warehouse_ids))
            params = {f"wid_{idx}": row_id for idx, row_id in enumerate(disabled_warehouse_ids)}
            conn.execute(text(f"UPDATE bi_inventory_warehouse_map SET is_enabled = 0 WHERE id IN ({placeholders})"), params)
        if disabled_status_ids:
            placeholders = ", ".join(f":sid_{idx}" for idx, _ in enumerate(disabled_status_ids))
            params = {f"sid_{idx}": row_id for idx, row_id in enumerate(disabled_status_ids)}
            conn.execute(text(f"UPDATE bi_inventory_status_map SET is_enabled = 0 WHERE id IN ({placeholders})"), params)
        if disabled_channel_ids:
            placeholders = ", ".join(f":cid_{idx}" for idx, _ in enumerate(disabled_channel_ids))
            params = {f"cid_{idx}": row_id for idx, row_id in enumerate(disabled_channel_ids)}
            conn.execute(text(f"UPDATE bi_channel_shop_master SET is_active = 0 WHERE id IN ({placeholders})"), params)

    refreshed_rows = refresh_inventory_cleaning(current_engine)
    record_dashboard_audit(
        module_key="governance",
        module_name="治理中心",
        action_key="master_data.save",
        action_name="保存主数据",
        target_type="master_data",
        target_name="SKU / 仓库 / 状态 / 渠道主数据",
        detail_summary=(
            f"SKU {len(sku_rows)} / 仓库 {len(warehouse_rows)} / 状态 {len(status_rows)} / 渠道 {len(channel_rows)}"
        ),
        detail={
            "sku_count": len(sku_rows),
            "warehouse_count": len(warehouse_rows),
            "status_count": len(status_rows),
            "channel_count": len(channel_rows),
            "sku_codes": audit_preview_values(sku_rows, "sku_code"),
            "warehouse_names": audit_preview_values(warehouse_rows, "warehouse_name_clean"),
            "status_names": audit_preview_values(status_rows, "stock_status_name"),
            "channel_names": audit_preview_values(channel_rows, "channel_name"),
            "refreshed_rows": int(refreshed_rows),
        },
        triggered_by=username,
        source_path="/financial/bi-dashboard/api/master-data",
        source_method="PUT",
        affected_count=len(sku_rows) + len(warehouse_rows) + len(status_rows) + len(channel_rows),
    )
    return JSONResponse(
        {
            "saved": True,
            "sku_count": len(sku_rows),
            "warehouse_count": len(warehouse_rows),
            "status_count": len(status_rows),
            "channel_count": len(channel_rows),
            "refreshed_rows": int(refreshed_rows),
        }
    )


@router.get("/bi-dashboard/api/audit-logs")
async def list_audit_logs(
    module_key: str | None = Query(None),
    result_status: str | None = Query(None),
    actor: str | None = Query(None),
    keyword: str | None = Query(None),
    limit: int = Query(120, ge=1, le=400),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    conditions = ["1 = 1"]
    params: Dict[str, Any] = {}
    if module_key:
        conditions.append("module_key = :module_key")
        params["module_key"] = str(module_key).strip()
    if result_status:
        conditions.append("result_status = :result_status")
        params["result_status"] = str(result_status).strip()
    if actor:
        conditions.append("triggered_by = :actor")
        params["actor"] = str(actor).strip()
    if keyword:
        conditions.append(
            "(action_name LIKE :keyword OR target_name LIKE :keyword OR detail_summary LIKE :keyword OR target_id LIKE :keyword)"
        )
        params["keyword"] = f"%{str(keyword).strip()}%"
    where_sql = " AND ".join(conditions)
    current_engine = get_engine()
    with current_engine.connect() as conn:
        rows = conn.execute(
            text(
                f"""
                SELECT
                    id, module_key, module_name, action_key, action_name,
                    target_type, target_id, target_name, result_status,
                    detail_summary, detail_json, source_path, source_method,
                    triggered_by, affected_count, created_at
                FROM bi_audit_log
                WHERE {where_sql}
                ORDER BY id DESC
                LIMIT :limit
                """
            ),
            {**params, "limit": int(limit)},
        ).mappings().all()
        summary_row = conn.execute(
            text(
                f"""
                SELECT
                    COUNT(*) AS total_count,
                    SUM(CASE WHEN result_status = 'success' THEN 1 ELSE 0 END) AS success_count,
                    SUM(CASE WHEN result_status = 'submitted' THEN 1 ELSE 0 END) AS submitted_count,
                    SUM(CASE WHEN result_status = 'failed' THEN 1 ELSE 0 END) AS failed_count,
                    COUNT(DISTINCT module_key) AS module_count,
                    MAX(created_at) AS latest_at
                FROM bi_audit_log
                WHERE {where_sql}
                """
            ),
            params,
        ).mappings().first()
        module_rows = conn.execute(
            text(
                """
                SELECT module_key, module_name, COUNT(*) AS item_count
                FROM bi_audit_log
                GROUP BY module_key, module_name
                ORDER BY MAX(created_at) DESC, module_name ASC
                """
            )
        ).mappings().all()
    items = [
        {
            "id": int(row["id"]),
            "module_key": row["module_key"],
            "module_name": row["module_name"],
            "action_key": row["action_key"],
            "action_name": row["action_name"],
            "target_type": row["target_type"],
            "target_id": row["target_id"],
            "target_name": row["target_name"],
            "result_status": row["result_status"],
            "detail_summary": row["detail_summary"],
            "detail": json_loads(row.get("detail_json"), {}),
            "source_path": row["source_path"],
            "source_method": row["source_method"],
            "triggered_by": row["triggered_by"] or "",
            "affected_count": int(row["affected_count"] or 0),
            "created_at": to_plain(row["created_at"]),
        }
        for row in rows
    ]
    summary = summary_row or {}
    return JSONResponse(
        {
            "items": items,
            "summary": {
                "total_count": int(summary.get("total_count") or 0),
                "success_count": int(summary.get("success_count") or 0),
                "submitted_count": int(summary.get("submitted_count") or 0),
                "failed_count": int(summary.get("failed_count") or 0),
                "module_count": int(summary.get("module_count") or 0),
                "latest_at": to_plain(summary.get("latest_at")),
            },
            "module_options": [
                {
                    "value": row["module_key"],
                    "label": row["module_name"],
                    "item_count": int(row["item_count"] or 0),
                }
                for row in module_rows
            ],
            "status_options": [
                {"value": "success", "label": "鎴愬姛"},
                {"value": "submitted", "label": "已提交"},
                {"value": "failed", "label": "澶辫触"},
            ],
        }
    )


@router.get("/bi-dashboard/api/procurement-arrivals")
async def list_procurement_arrivals(
    status: str | None = Query(None),
    document_status: str | None = Query(None),
    warehouse_code: str | None = Query(None),
    keyword: str | None = Query(None),
    limit: int = Query(80, ge=1, le=300),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    conditions = ["1 = 1"]
    params: Dict[str, Any] = {}
    if status:
        conditions.append("status = :status")
        params["status"] = str(status).strip()
    if document_status:
        conditions.append("document_status = :document_status")
        params["document_status"] = str(document_status).strip()
    if warehouse_code:
        conditions.append("warehouse_code = :warehouse_code")
        params["warehouse_code"] = str(warehouse_code).strip()
    if keyword:
        conditions.append(
            "(arrival_no LIKE :keyword OR purchase_order_no LIKE :keyword OR supplier_name LIKE :keyword OR sku_code LIKE :keyword OR sku_name LIKE :keyword)"
        )
        params["keyword"] = f"%{str(keyword).strip()}%"
    where_sql = " AND ".join(conditions)

    current_engine = get_engine()
    with current_engine.connect() as conn:
        rows = conn.execute(
            text(
                f"""
                SELECT
                    id, arrival_no, purchase_order_no, supplier_name, warehouse_code, warehouse_name,
                    channel_code, channel_name, sku_code, sku_name, expected_qty, arrived_qty,
                    qualified_qty, exception_qty, unit, arrival_date, status, document_status,
                    exception_reason, remark, source_system, created_by, updated_by,
                    sort_order, created_at, updated_at
                FROM bi_procurement_arrival
                WHERE {where_sql}
                ORDER BY arrival_date DESC, sort_order, id DESC
                LIMIT :limit
                """
            ),
            {**params, "limit": int(limit)},
        ).mappings().all()
        warehouse_rows = conn.execute(
            text(
                """
                SELECT DISTINCT warehouse_code, warehouse_name_clean
                FROM bi_inventory_warehouse_map
                WHERE is_enabled = 1
                ORDER BY warehouse_name_clean, warehouse_code
                """
            )
        ).mappings().all()
        channel_rows = conn.execute(
            text(
                """
                SELECT DISTINCT channel_code, channel_name
                FROM bi_channel_shop_master
                WHERE is_active = 1
                ORDER BY channel_name, channel_code
                """
            )
        ).mappings().all()
        supplier_rows = conn.execute(
            text(
                """
                SELECT DISTINCT supplier_name
                FROM bi_procurement_arrival
                WHERE supplier_name <> ''
                ORDER BY supplier_name
                """
            )
        ).fetchall()

    items = [serialize_procurement_arrival_row(row) for row in rows]
    summary = {
        "total_count": len(items),
        "draft_count": sum(1 for item in items if item["status"] == "draft"),
        "ready_count": sum(1 for item in items if item["status"] == "ready"),
        "completed_count": sum(1 for item in items if item["status"] == "completed"),
        "exception_count": sum(1 for item in items if item["status"] == "exception"),
        "pending_document_count": sum(1 for item in items if item["document_status"] in {"pending", "failed"}),
        "total_expected_qty": round(sum(float(item["expected_qty"]) for item in items), 2),
        "total_arrived_qty": round(sum(float(item["arrived_qty"]) for item in items), 2),
        "total_qualified_qty": round(sum(float(item["qualified_qty"]) for item in items), 2),
        "latest_arrival_date": max((item["arrival_date"] for item in items if item["arrival_date"]), default=None),
    }

    return JSONResponse(
        {
            "items": items,
            "summary": summary,
            "status_options": procurement_arrival_status_options(),
            "document_status_options": procurement_document_status_options(),
            "warehouse_options": [
                {
                    "value": str(row["warehouse_code"] or ""),
                    "label": str(row["warehouse_name_clean"] or row["warehouse_code"] or ""),
                }
                for row in warehouse_rows
                if str(row["warehouse_code"] or "").strip()
            ],
            "channel_options": [
                {
                    "value": str(row["channel_code"] or ""),
                    "label": str(row["channel_name"] or row["channel_code"] or ""),
                }
                for row in channel_rows
                if str(row["channel_code"] or "").strip()
            ],
            "supplier_options": [
                {"value": str(row[0] or ""), "label": str(row[0] or "")}
                for row in supplier_rows
                if str(row[0] or "").strip()
            ],
        }
    )


@router.post("/bi-dashboard/api/procurement-arrivals")
async def save_procurement_arrival(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    record = normalize_procurement_arrival_payload(payload)
    current_engine = get_engine()
    with current_engine.begin() as conn:
        duplicate = conn.execute(
            text(
                """
                SELECT id
                FROM bi_procurement_arrival
                WHERE arrival_no = :arrival_no AND id <> :id
                LIMIT 1
                """
            ),
            {"arrival_no": record["arrival_no"], "id": record["id"]},
        ).fetchone()
        if duplicate:
            raise HTTPException(status_code=400, detail=f"鍒拌揣鍗曞彿宸插瓨鍦細{record['arrival_no']}")

        payload_with_user = {**record, "updated_by": username}
        created = False
        saved_id = int(record["id"] or 0)
        if record["id"] > 0:
            result = conn.execute(
                text(
                    """
                    UPDATE bi_procurement_arrival
                    SET
                        arrival_no = :arrival_no,
                        purchase_order_no = :purchase_order_no,
                        supplier_name = :supplier_name,
                        warehouse_code = :warehouse_code,
                        warehouse_name = :warehouse_name,
                        channel_code = :channel_code,
                        channel_name = :channel_name,
                        sku_code = :sku_code,
                        sku_name = :sku_name,
                        expected_qty = :expected_qty,
                        arrived_qty = :arrived_qty,
                        qualified_qty = :qualified_qty,
                        exception_qty = :exception_qty,
                        unit = :unit,
                        arrival_date = :arrival_date,
                        status = :status,
                        document_status = :document_status,
                        exception_reason = :exception_reason,
                        remark = :remark,
                        source_system = :source_system,
                        updated_by = :updated_by,
                        sort_order = :sort_order
                    WHERE id = :id
                    """
                ),
                payload_with_user,
            )
            if not result.rowcount:
                saved_id = 0

        if saved_id <= 0:
            created = True
            payload_with_user["created_by"] = username
            insert_result = conn.execute(
                text(
                    """
                    INSERT INTO bi_procurement_arrival(
                        arrival_no, purchase_order_no, supplier_name, warehouse_code, warehouse_name,
                        channel_code, channel_name, sku_code, sku_name, expected_qty, arrived_qty,
                        qualified_qty, exception_qty, unit, arrival_date, status, document_status,
                        exception_reason, remark, source_system, created_by, updated_by, sort_order
                    ) VALUES (
                        :arrival_no, :purchase_order_no, :supplier_name, :warehouse_code, :warehouse_name,
                        :channel_code, :channel_name, :sku_code, :sku_name, :expected_qty, :arrived_qty,
                        :qualified_qty, :exception_qty, :unit, :arrival_date, :status, :document_status,
                        :exception_reason, :remark, :source_system, :created_by, :updated_by, :sort_order
                    )
                    """
                ),
                payload_with_user,
            )
            saved_id = int(insert_result.lastrowid or 0)

        saved_row = conn.execute(
            text(
                """
                SELECT
                    id, arrival_no, purchase_order_no, supplier_name, warehouse_code, warehouse_name,
                    channel_code, channel_name, sku_code, sku_name, expected_qty, arrived_qty,
                    qualified_qty, exception_qty, unit, arrival_date, status, document_status,
                    exception_reason, remark, source_system, created_by, updated_by,
                    sort_order, created_at, updated_at
                FROM bi_procurement_arrival
                WHERE id = :id
                """
            ),
            {"id": saved_id},
        ).mappings().first()
        item = serialize_procurement_arrival_row(saved_row)
        flow_sync = sync_procurement_inventory_flow_tasks(conn, item, username)
        task_center_sync = sync_task_center_snapshot(conn, username)
        reconciliation_sync = sync_reconciliation_snapshot(conn, username)

    record_dashboard_audit(
        module_key="procurement",
        module_name="采购到货",
        action_key="arrival.upsert",
        action_name="保存采购到货单",
        target_type="procurement_arrival",
        target_id=item["arrival_no"],
        target_name=f"{item['purchase_order_no']} / {item['sku_name']}",
        detail_summary=f"{item['status']} / {item['document_status']} / 鍒拌揣 {item['arrived_qty']} {item['unit']}",
        detail={
            "arrival_no": item["arrival_no"],
            "purchase_order_no": item["purchase_order_no"],
            "supplier_name": item["supplier_name"],
            "warehouse_name": item["warehouse_name"],
            "sku_code": item["sku_code"],
            "sku_name": item["sku_name"],
            "status": item["status"],
            "document_status": item["document_status"],
            "expected_qty": item["expected_qty"],
            "arrived_qty": item["arrived_qty"],
            "qualified_qty": item["qualified_qty"],
            "exception_qty": item["exception_qty"],
            "inventory_flow_sync": flow_sync,
        },
        triggered_by=username,
        source_path=PROCUREMENT_ARRIVAL_API_PATH,
        source_method="POST",
        affected_count=1,
    )
    if any(int(flow_sync[key]) for key in ("created_count", "updated_count", "cancelled_count")):
        record_dashboard_audit(
            module_key="inventory_flow",
            module_name="库存流转",
            action_key="task.auto_sync",
            action_name="采购到货触发库存流转任务同步",
            target_type="inventory_flow_task",
            target_id=item["arrival_no"],
            target_name=item["sku_name"],
            detail_summary=(
                f"新增 / 更新 / 取消 {flow_sync['created_count']} / {flow_sync['updated_count']} / "
                f"{flow_sync['cancelled_count']}"
            ),
            detail={
                "arrival_no": item["arrival_no"],
                "task_nos": flow_sync["task_nos"],
                "created_count": flow_sync["created_count"],
                "updated_count": flow_sync["updated_count"],
                "cancelled_count": flow_sync["cancelled_count"],
            },
            triggered_by=username,
            source_path=PROCUREMENT_ARRIVAL_API_PATH,
            source_method="POST",
            affected_count=int(flow_sync["created_count"] + flow_sync["updated_count"]),
        )
    return JSONResponse(
        {
            "saved": True,
            "created": created,
            "item": item,
            "inventory_flow_sync": flow_sync,
            "task_center_sync": task_center_sync,
            "reconciliation_sync": reconciliation_sync,
        }
    )


@router.get("/bi-dashboard/api/inventory-flows")
async def list_inventory_flows(
    task_status: str | None = Query(None),
    action_type: str | None = Query(None),
    keyword: str | None = Query(None),
    limit: int = Query(120, ge=1, le=400),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    conditions = ["1 = 1"]
    params: Dict[str, Any] = {}
    if task_status:
        conditions.append("task_status = :task_status")
        params["task_status"] = str(task_status).strip()
    if action_type:
        conditions.append("action_type = :action_type")
        params["action_type"] = str(action_type).strip()
    if keyword:
        conditions.append(
            "(task_no LIKE :keyword OR source_record_no LIKE :keyword OR sku_code LIKE :keyword OR sku_name LIKE :keyword OR note LIKE :keyword)"
        )
        params["keyword"] = f"%{str(keyword).strip()}%"
    where_sql = " AND ".join(conditions)
    current_engine = get_engine()
    with current_engine.connect() as conn:
        rule_rows = conn.execute(
            text(
                """
                SELECT
                    id, rule_name, trigger_source, trigger_condition, action_type,
                    source_status_id, source_status_name, target_status_id, target_status_name,
                    source_warehouse_code, source_warehouse_name, target_warehouse_code, target_warehouse_name,
                    priority, auto_create_task, is_enabled, sort_order, note, created_by, updated_by,
                    created_at, updated_at
                FROM bi_inventory_flow_rule
                ORDER BY sort_order, id
                """
            )
        ).mappings().all()
        task_rows = conn.execute(
            text(
                f"""
                SELECT
                    id, task_no, source_record_type, source_record_id, source_record_no, trigger_source,
                    action_type, task_status, priority, sku_code, sku_name, request_qty, confirmed_qty,
                    source_status_id, source_status_name, target_status_id, target_status_name,
                    source_warehouse_code, source_warehouse_name, target_warehouse_code, target_warehouse_name,
                    planned_execute_date, reason_text, note, created_by, updated_by, sort_order, created_at, updated_at
                FROM bi_inventory_flow_task
                WHERE {where_sql}
                ORDER BY FIELD(task_status, 'pending', 'blocked', 'draft', 'completed', 'cancelled'),
                         planned_execute_date DESC, updated_at DESC, id DESC
                LIMIT :limit
                """
            ),
            {**params, "limit": int(limit)},
        ).mappings().all()
        status_rows = conn.execute(
            text(
                """
                SELECT stock_status_id, stock_status_name
                FROM bi_inventory_status_map
                WHERE is_enabled = 1
                ORDER BY sort_order, id
                """
            )
        ).mappings().all()
        warehouse_rows = conn.execute(
            text(
                """
                SELECT warehouse_code, warehouse_name_clean
                FROM bi_inventory_warehouse_map
                WHERE is_enabled = 1
                ORDER BY sort_order, id
                """
            )
        ).mappings().all()
    rules = [serialize_inventory_flow_rule_row(row) for row in rule_rows]
    tasks = [serialize_inventory_flow_task_row(row) for row in task_rows]
    summary = {
        "task_count": len(tasks),
        "pending_count": sum(1 for item in tasks if item["task_status"] == "pending"),
        "blocked_count": sum(1 for item in tasks if item["task_status"] == "blocked"),
        "completed_count": sum(1 for item in tasks if item["task_status"] == "completed"),
        "enabled_rule_count": sum(1 for item in rules if item["is_enabled"]),
        "auto_rule_count": sum(1 for item in rules if item["is_enabled"] and item["auto_create_task"]),
        "transfer_count": sum(1 for item in tasks if item["action_type"] == "warehouse_transfer"),
    }
    return JSONResponse(
        {
            "rules": rules,
            "tasks": tasks,
            "summary": summary,
            "action_options": inventory_flow_action_options(),
            "task_status_options": inventory_flow_task_status_options(),
            "priority_options": inventory_flow_priority_options(),
            "trigger_source_options": inventory_flow_trigger_options(),
            "status_options": [
                {"value": str(row["stock_status_id"] or ""), "label": str(row["stock_status_name"] or "")}
                for row in status_rows
                if str(row["stock_status_id"] or "").strip()
            ],
            "warehouse_options": [
                {"value": str(row["warehouse_code"] or ""), "label": str(row["warehouse_name_clean"] or "")}
                for row in warehouse_rows
                if str(row["warehouse_code"] or "").strip()
            ],
        }
    )


@router.put("/bi-dashboard/api/inventory-flows/rules")
async def save_inventory_flow_rules(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    rules = normalize_inventory_flow_rule_payload(payload.get("rules"))
    current_engine = get_engine()
    with current_engine.begin() as conn:
        status_lookup, warehouse_lookup = resolve_inventory_master_lookups(conn)
        existing_ids = {int(row[0]) for row in conn.execute(text("SELECT id FROM bi_inventory_flow_rule")).fetchall()}
        submitted_ids = {item["id"] for item in rules if item["id"] > 0}
        for item in rules:
            record = {
                **item,
                "source_status_name": status_lookup.get(item["source_status_id"], ""),
                "target_status_name": status_lookup.get(item["target_status_id"], ""),
                "source_warehouse_name": warehouse_lookup.get(item["source_warehouse_code"], ""),
                "target_warehouse_name": warehouse_lookup.get(item["target_warehouse_code"], ""),
                "updated_by": username,
            }
            if item["id"] > 0:
                result = conn.execute(
                    text(
                        """
                        UPDATE bi_inventory_flow_rule
                        SET
                            rule_name = :rule_name,
                            trigger_source = :trigger_source,
                            trigger_condition = :trigger_condition,
                            action_type = :action_type,
                            source_status_id = :source_status_id,
                            source_status_name = :source_status_name,
                            target_status_id = :target_status_id,
                            target_status_name = :target_status_name,
                            source_warehouse_code = :source_warehouse_code,
                            source_warehouse_name = :source_warehouse_name,
                            target_warehouse_code = :target_warehouse_code,
                            target_warehouse_name = :target_warehouse_name,
                            priority = :priority,
                            auto_create_task = :auto_create_task,
                            is_enabled = :is_enabled,
                            sort_order = :sort_order,
                            note = :note,
                            updated_by = :updated_by
                        WHERE id = :id
                        """
                    ),
                    record,
                )
                if result.rowcount:
                    continue
            record["created_by"] = username
            conn.execute(
                text(
                    """
                    INSERT INTO bi_inventory_flow_rule(
                        rule_name, trigger_source, trigger_condition, action_type,
                        source_status_id, source_status_name, target_status_id, target_status_name,
                        source_warehouse_code, source_warehouse_name, target_warehouse_code, target_warehouse_name,
                        priority, auto_create_task, is_enabled, sort_order, note, created_by, updated_by
                    ) VALUES (
                        :rule_name, :trigger_source, :trigger_condition, :action_type,
                        :source_status_id, :source_status_name, :target_status_id, :target_status_name,
                        :source_warehouse_code, :source_warehouse_name, :target_warehouse_code, :target_warehouse_name,
                        :priority, :auto_create_task, :is_enabled, :sort_order, :note, :created_by, :updated_by
                    )
                    """
                ),
                record,
            )
        disabled_ids = sorted(existing_ids - submitted_ids)
        if disabled_ids:
            placeholders = ", ".join(f":rid_{idx}" for idx, _ in enumerate(disabled_ids))
            params_disabled = {f"rid_{idx}": row_id for idx, row_id in enumerate(disabled_ids)}
            params_disabled["updated_by"] = username
            conn.execute(
                text(
                    f"""
                    UPDATE bi_inventory_flow_rule
                    SET is_enabled = 0, updated_by = :updated_by
                    WHERE id IN ({placeholders})
                    """
                ),
                params_disabled,
            )
    record_dashboard_audit(
        module_key="inventory_flow",
        module_name="库存流转",
        action_key="rules.save",
        action_name="保存库存流转规则",
        target_type="inventory_flow_rule",
        target_name="库存状态流转与调拨规则",
        detail_summary=f"保存 {len(rules)} 条规则",
        detail={
            "rule_count": len(rules),
            "rule_names": audit_preview_values(rules, "rule_name"),
            "enabled_rule_count": sum(1 for item in rules if item["is_enabled"]),
            "auto_rule_count": sum(1 for item in rules if item["auto_create_task"]),
        },
        triggered_by=username,
        source_path=INVENTORY_FLOW_RULE_API_PATH,
        source_method="PUT",
        affected_count=len(rules),
    )
    return JSONResponse({"saved": True, "rule_count": len(rules)})


@router.post("/bi-dashboard/api/inventory-flows/tasks")
async def save_inventory_flow_task(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    task = normalize_inventory_flow_task_payload(payload)
    current_engine = get_engine()
    with current_engine.begin() as conn:
        duplicate = conn.execute(
            text(
                """
                SELECT id
                FROM bi_inventory_flow_task
                WHERE task_no = :task_no AND id <> :id
                LIMIT 1
                """
            ),
            {"task_no": task["task_no"], "id": task["id"]},
        ).fetchone()
        if duplicate:
            raise HTTPException(status_code=400, detail=f"库存流转任务号已存在：{task['task_no']}")
        status_lookup, warehouse_lookup = resolve_inventory_master_lookups(conn)
        record = {
            **task,
            "source_status_name": status_lookup.get(task["source_status_id"], ""),
            "target_status_name": status_lookup.get(task["target_status_id"], ""),
            "source_warehouse_name": warehouse_lookup.get(task["source_warehouse_code"], ""),
            "target_warehouse_name": warehouse_lookup.get(task["target_warehouse_code"], ""),
            "updated_by": username,
        }
        saved_id = int(task["id"] or 0)
        created = False
        if saved_id > 0:
            result = conn.execute(
                text(
                    """
                    UPDATE bi_inventory_flow_task
                    SET
                        task_no = :task_no,
                        source_record_type = :source_record_type,
                        source_record_id = :source_record_id,
                        source_record_no = :source_record_no,
                        trigger_source = :trigger_source,
                        action_type = :action_type,
                        task_status = :task_status,
                        priority = :priority,
                        sku_code = :sku_code,
                        sku_name = :sku_name,
                        request_qty = :request_qty,
                        confirmed_qty = :confirmed_qty,
                        source_status_id = :source_status_id,
                        source_status_name = :source_status_name,
                        target_status_id = :target_status_id,
                        target_status_name = :target_status_name,
                        source_warehouse_code = :source_warehouse_code,
                        source_warehouse_name = :source_warehouse_name,
                        target_warehouse_code = :target_warehouse_code,
                        target_warehouse_name = :target_warehouse_name,
                        planned_execute_date = :planned_execute_date,
                        reason_text = :reason_text,
                        note = :note,
                        updated_by = :updated_by
                    WHERE id = :id
                    """
                ),
                record,
            )
            if not result.rowcount:
                saved_id = 0
        if saved_id <= 0:
            created = True
            record["created_by"] = username
            insert_result = conn.execute(
                text(
                    """
                    INSERT INTO bi_inventory_flow_task(
                        task_no, source_record_type, source_record_id, source_record_no, trigger_source,
                        action_type, task_status, priority, sku_code, sku_name, request_qty, confirmed_qty,
                        source_status_id, source_status_name, target_status_id, target_status_name,
                        source_warehouse_code, source_warehouse_name, target_warehouse_code, target_warehouse_name,
                        planned_execute_date, reason_text, note, created_by, updated_by, sort_order
                    ) VALUES (
                        :task_no, :source_record_type, :source_record_id, :source_record_no, :trigger_source,
                        :action_type, :task_status, :priority, :sku_code, :sku_name, :request_qty, :confirmed_qty,
                        :source_status_id, :source_status_name, :target_status_id, :target_status_name,
                        :source_warehouse_code, :source_warehouse_name, :target_warehouse_code, :target_warehouse_name,
                        :planned_execute_date, :reason_text, :note, :created_by, :updated_by, 100
                    )
                    """
                ),
                record,
            )
            saved_id = int(insert_result.lastrowid or 0)
        saved_row = conn.execute(
            text(
                """
                SELECT
                    id, task_no, source_record_type, source_record_id, source_record_no, trigger_source,
                    action_type, task_status, priority, sku_code, sku_name, request_qty, confirmed_qty,
                    source_status_id, source_status_name, target_status_id, target_status_name,
                    source_warehouse_code, source_warehouse_name, target_warehouse_code, target_warehouse_name,
                    planned_execute_date, reason_text, note, created_by, updated_by, sort_order, created_at, updated_at
                FROM bi_inventory_flow_task
                WHERE id = :id
                """
            ),
            {"id": saved_id},
        ).mappings().first()
        task_center_sync = sync_task_center_snapshot(conn, username)
        reconciliation_sync = sync_reconciliation_snapshot(conn, username)
    item = serialize_inventory_flow_task_row(saved_row)
    record_dashboard_audit(
        module_key="inventory_flow",
        module_name="鎼存挸鐡ㄥù浣芥祮",
        action_key="task.upsert",
        action_name="娣囨繂鐡ㄦ惔鎾崇摠濞翠浇娴嗘禒璇插",
        target_type="inventory_flow_task",
        target_id=item["task_no"],
        target_name=item["sku_name"],
        detail_summary=f"{item['task_status']} / {item['action_type']} / {item['request_qty']}",
        detail=item,
        triggered_by=username,
        source_path=INVENTORY_FLOW_TASK_API_PATH,
        source_method="POST",
        affected_count=1,
    )
    return JSONResponse(
        {
            "saved": True,
            "created": created,
            "item": item,
            "task_center_sync": task_center_sync,
            "reconciliation_sync": reconciliation_sync,
        }
    )


@router.get("/bi-dashboard/api/task-center")
async def list_task_center_items(
    task_status: str | None = Query(None),
    source_module: str | None = Query(None),
    priority: str | None = Query(None),
    keyword: str | None = Query(None),
    limit: int = Query(160, ge=1, le=400),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    conditions = ["1 = 1"]
    params: Dict[str, Any] = {}
    if task_status:
        conditions.append("task_status = :task_status")
        params["task_status"] = str(task_status).strip()
    if source_module:
        conditions.append("source_module = :source_module")
        params["source_module"] = str(source_module).strip()
    if priority:
        conditions.append("priority = :priority")
        params["priority"] = str(priority).strip()
    if keyword:
        conditions.append(
            "(task_title LIKE :keyword OR source_no LIKE :keyword OR summary_text LIKE :keyword OR note LIKE :keyword)"
        )
        params["keyword"] = f"%{str(keyword).strip()}%"
    where_sql = " AND ".join(conditions)
    current_engine = get_engine()
    with current_engine.begin() as conn:
        sync_task_center_snapshot(conn, "system")
        rows = conn.execute(
            text(
                f"""
                SELECT
                    id, source_module, source_type, source_id, source_no, task_title, task_category,
                    task_status, priority, owner_name, owner_role, due_date, source_status,
                    source_detail_status, summary_text, note, source_snapshot_json,
                    sort_order, created_by, updated_by, created_at, updated_at
                FROM bi_task_center_item
                WHERE {where_sql}
                ORDER BY FIELD(task_status, 'blocked', 'open', 'in_progress', 'completed'),
                         FIELD(priority, 'high', 'normal', 'low'),
                         CASE WHEN due_date IS NULL THEN 1 ELSE 0 END,
                         due_date ASC,
                         sort_order ASC,
                         updated_at DESC,
                         id DESC
                LIMIT :limit
                """
            ),
            {**params, "limit": int(limit)},
        ).mappings().all()
    items = [serialize_task_center_row(row) for row in rows]
    summary = {
        "total_count": len(items),
        "open_count": sum(1 for item in items if item["task_status"] == "open"),
        "in_progress_count": sum(1 for item in items if item["task_status"] == "in_progress"),
        "blocked_count": sum(1 for item in items if item["task_status"] == "blocked"),
        "completed_count": sum(1 for item in items if item["task_status"] == "completed"),
        "overdue_count": sum(1 for item in items if item["is_overdue"]),
        "procurement_count": sum(1 for item in items if item["source_module"] == "procurement"),
        "inventory_flow_count": sum(1 for item in items if item["source_module"] == "inventory_flow"),
        "high_priority_count": sum(1 for item in items if item["priority"] == "high"),
        "latest_updated_at": max((item["updated_at"] for item in items if item["updated_at"]), default=None),
    }
    return JSONResponse(
        {
            "items": items,
            "summary": summary,
            "task_status_options": task_center_status_options(),
            "source_module_options": task_center_source_module_options(),
            "priority_options": inventory_flow_priority_options(),
            "category_options": task_center_category_options(),
        }
    )


@router.post("/bi-dashboard/api/task-center/items")
async def save_task_center_item(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    record = normalize_task_center_payload(payload)
    current_engine = get_engine()
    with current_engine.begin() as conn:
        existing = conn.execute(
            text(
                """
                SELECT
                    id, source_module, source_type, source_id, source_no, task_title, task_category,
                    task_status, priority, owner_name, owner_role, due_date, source_status,
                    source_detail_status, summary_text, note, source_snapshot_json,
                    sort_order, created_by, updated_by, created_at, updated_at
                FROM bi_task_center_item
                WHERE id = :id
                LIMIT 1
                """
            ),
            {"id": record["id"]},
        ).mappings().first()
        if not existing:
            raise HTTPException(status_code=404, detail="任务中心记录不存在")

        conn.execute(
            text(
                """
                UPDATE bi_task_center_item
                SET
                    task_status = :task_status,
                    priority = :priority,
                    owner_name = :owner_name,
                    owner_role = :owner_role,
                    due_date = :due_date,
                    note = :note,
                    updated_by = :updated_by
                WHERE id = :id
                """
            ),
            {
                **record,
                "updated_by": username,
            },
        )

        if str(existing["source_module"] or "") == "inventory_flow":
            inventory_status = {
                "open": "pending",
                "in_progress": "pending",
                "blocked": "blocked",
                "completed": "completed",
            }.get(record["task_status"], "pending")
            conn.execute(
                text(
                    """
                    UPDATE bi_inventory_flow_task
                    SET
                        task_status = :task_status,
                        priority = :priority,
                        planned_execute_date = :planned_execute_date,
                        updated_by = :updated_by
                    WHERE id = :id
                    """
                ),
                {
                    "id": parse_int_or_default(existing["source_id"], 0),
                    "task_status": inventory_status,
                    "priority": record["priority"],
                    "planned_execute_date": record["due_date"],
                    "updated_by": username,
                },
            )

        sync_task_center_snapshot(conn, username)
        saved_row = conn.execute(
            text(
                """
                SELECT
                    id, source_module, source_type, source_id, source_no, task_title, task_category,
                    task_status, priority, owner_name, owner_role, due_date, source_status,
                    source_detail_status, summary_text, note, source_snapshot_json,
                    sort_order, created_by, updated_by, created_at, updated_at
                FROM bi_task_center_item
                WHERE id = :id
                LIMIT 1
                """
            ),
            {"id": record["id"]},
        ).mappings().first()
    item = serialize_task_center_row(saved_row)
    record_dashboard_audit(
        module_key="task_center",
        module_name="任务中心",
        action_key="item.update",
        action_name="更新任务中心待办",
        target_type="task_center_item",
        target_id=item["source_no"],
        target_name=item["task_title"],
        detail_summary=f"{item['task_status_label']} / {item['priority_label']} / {item['owner_name'] or item['owner_role'] or '未分配'}",
        detail=item,
        triggered_by=username,
        source_path=TASK_CENTER_ITEM_API_PATH,
        source_method="POST",
        affected_count=1,
    )
    return JSONResponse({"saved": True, "item": item})


@router.get("/bi-dashboard/api/reconciliation-center")
async def list_reconciliation_cases(
    case_status: str | None = Query(None),
    case_type: str | None = Query(None),
    severity: str | None = Query(None),
    keyword: str | None = Query(None),
    limit: int = Query(160, ge=1, le=400),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    conditions = ["1 = 1"]
    params: Dict[str, Any] = {}
    if case_status:
        conditions.append("case_status = :case_status")
        params["case_status"] = str(case_status).strip()
    if case_type:
        conditions.append("case_type = :case_type")
        params["case_type"] = str(case_type).strip()
    if severity:
        conditions.append("severity = :severity")
        params["severity"] = str(severity).strip()
    if keyword:
        conditions.append(
            "(case_title LIKE :keyword OR source_no LIKE :keyword OR diff_summary LIKE :keyword OR compensation_note LIKE :keyword)"
        )
        params["keyword"] = f"%{str(keyword).strip()}%"
    where_sql = " AND ".join(conditions)
    current_engine = get_engine()
    with current_engine.begin() as conn:
        sync_reconciliation_snapshot(conn, "system")
        rows = conn.execute(
            text(
                f"""
                SELECT
                    id, source_module, source_type, source_id, source_no, case_type, case_title,
                    case_status, severity, diff_summary, owner_name, owner_role, due_date,
                    expected_snapshot_json, actual_snapshot_json, last_compensation_action,
                    compensation_note, compensated_at, compensated_by,
                    sort_order, created_by, updated_by, created_at, updated_at
                FROM bi_reconciliation_case
                WHERE {where_sql}
                ORDER BY FIELD(case_status, 'open', 'compensating', 'resolved', 'ignored'),
                         FIELD(severity, 'high', 'normal', 'low'),
                         CASE WHEN due_date IS NULL THEN 1 ELSE 0 END,
                         due_date ASC,
                         sort_order ASC,
                         updated_at DESC,
                         id DESC
                LIMIT :limit
                """
            ),
            {**params, "limit": int(limit)},
        ).mappings().all()
    items = [serialize_reconciliation_case_row(row) for row in rows]
    summary = {
        "total_count": len(items),
        "open_count": sum(1 for item in items if item["case_status"] == "open"),
        "compensating_count": sum(1 for item in items if item["case_status"] == "compensating"),
        "resolved_count": sum(1 for item in items if item["case_status"] == "resolved"),
        "ignored_count": sum(1 for item in items if item["case_status"] == "ignored"),
        "high_severity_count": sum(1 for item in items if item["severity"] == "high"),
        "document_sync_count": sum(1 for item in items if item["case_type"] == "document_sync"),
        "inventory_missing_count": sum(1 for item in items if item["case_type"] == "inventory_task_missing"),
        "blocked_count": sum(1 for item in items if item["case_type"] == "inventory_task_blocked"),
        "overdue_count": sum(1 for item in items if item["is_overdue"]),
        "latest_updated_at": max((item["updated_at"] for item in items if item["updated_at"]), default=None),
    }
    return JSONResponse(
        {
            "items": items,
            "summary": summary,
            "case_status_options": reconciliation_case_status_options(),
            "case_type_options": reconciliation_case_type_options(),
            "severity_options": inventory_flow_priority_options(),
            "source_module_options": task_center_source_module_options(),
            "compensation_action_options": reconciliation_compensation_action_options(),
        }
    )


@router.post("/bi-dashboard/api/reconciliation-center/cases")
async def save_reconciliation_case(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    record = normalize_reconciliation_case_payload(payload)
    current_engine = get_engine()
    with current_engine.begin() as conn:
        existing = conn.execute(
            text(
                """
                SELECT
                    id, source_module, source_type, source_id, source_no, case_type, case_title,
                    case_status, severity, diff_summary, owner_name, owner_role, due_date,
                    expected_snapshot_json, actual_snapshot_json, last_compensation_action,
                    compensation_note, compensated_at, compensated_by,
                    sort_order, created_by, updated_by, created_at, updated_at
                FROM bi_reconciliation_case
                WHERE id = :id
                LIMIT 1
                """
            ),
            {"id": record["id"]},
        ).mappings().first()
        if not existing:
            raise HTTPException(status_code=404, detail="对账案例不存在")

        compensation_action = record["compensation_action"]
        case_status = record["case_status"]
        if compensation_action == "mark_resolved":
            case_status = "resolved"
        elif compensation_action == "ignore_case":
            case_status = "ignored"
        elif compensation_action:
            case_status = "compensating"

        conn.execute(
            text(
                """
                UPDATE bi_reconciliation_case
                SET
                    case_status = :case_status,
                    owner_name = :owner_name,
                    owner_role = :owner_role,
                    due_date = :due_date,
                    compensation_note = :compensation_note,
                    last_compensation_action = :last_compensation_action,
                    compensated_at = :compensated_at,
                    compensated_by = :compensated_by,
                    updated_by = :updated_by
                WHERE id = :id
                """
            ),
            {
                "id": record["id"],
                "case_status": case_status,
                "owner_name": record["owner_name"],
                "owner_role": record["owner_role"],
                "due_date": record["due_date"],
                "compensation_note": record["compensation_note"],
                "last_compensation_action": compensation_action or str(existing["last_compensation_action"] or ""),
                "compensated_at": datetime.now() if compensation_action else existing["compensated_at"],
                "compensated_by": username if compensation_action else existing["compensated_by"],
                "updated_by": username,
            },
        )

        compensation_result: Dict[str, Any] = {}
        if compensation_action == "retry_document_sync" and str(existing["source_module"] or "") == "procurement":
            procurement_item = fetch_procurement_item_by_id(conn, existing["source_id"])
            if not procurement_item:
                raise HTTPException(status_code=404, detail="来源采购到货单不存在")
            next_remark = str(procurement_item.get("remark") or "").strip()
            appended = "补偿动作：重试单据编排"
            next_remark = appended if not next_remark else f"{next_remark}\n{appended}"
            conn.execute(
                text(
                    """
                    UPDATE bi_procurement_arrival
                    SET document_status = 'generated', remark = :remark, updated_by = :updated_by
                    WHERE id = :id
                    """
                ),
                {"id": procurement_item["id"], "remark": next_remark, "updated_by": username},
            )
            compensation_result = {
                "action": "retry_document_sync",
                "document_status": "generated",
                "arrival_no": procurement_item["arrival_no"],
            }
        elif compensation_action == "resync_inventory_tasks" and str(existing["source_module"] or "") == "procurement":
            procurement_item = fetch_procurement_item_by_id(conn, existing["source_id"])
            if not procurement_item:
                raise HTTPException(status_code=404, detail="来源采购到货单不存在")
            compensation_result = sync_procurement_inventory_flow_tasks(conn, procurement_item, username)
            compensation_result["action"] = "resync_inventory_tasks"
        elif compensation_action == "reopen_inventory_task" and str(existing["source_module"] or "") == "inventory_flow":
            task_item = fetch_inventory_flow_task_by_id(conn, existing["source_id"])
            if not task_item:
                raise HTTPException(status_code=404, detail="来源库存流转任务不存在")
            next_note = str(task_item.get("note") or "").strip()
            appended = "补偿动作：解除阻塞并重新进入待执行"
            next_note = appended if not next_note else f"{next_note}\n{appended}"
            conn.execute(
                text(
                    """
                    UPDATE bi_inventory_flow_task
                    SET
                        task_status = 'pending',
                        planned_execute_date = :planned_execute_date,
                        note = :note,
                        updated_by = :updated_by
                    WHERE id = :id
                    """
                ),
                {
                    "id": task_item["id"],
                    "planned_execute_date": record["due_date"] or task_item.get("planned_execute_date"),
                    "note": next_note,
                    "updated_by": username,
                },
            )
            compensation_result = {"action": "reopen_inventory_task", "task_no": task_item["task_no"], "task_status": "pending"}

        sync_task_center_snapshot(conn, username)
        sync_reconciliation_snapshot(conn, username)
        saved_row = conn.execute(
            text(
                """
                SELECT
                    id, source_module, source_type, source_id, source_no, case_type, case_title,
                    case_status, severity, diff_summary, owner_name, owner_role, due_date,
                    expected_snapshot_json, actual_snapshot_json, last_compensation_action,
                    compensation_note, compensated_at, compensated_by,
                    sort_order, created_by, updated_by, created_at, updated_at
                FROM bi_reconciliation_case
                WHERE id = :id
                LIMIT 1
                """
            ),
            {"id": record["id"]},
        ).mappings().first()
    item = serialize_reconciliation_case_row(saved_row)
    record_dashboard_audit(
        module_key="reconciliation",
        module_name="对账补偿",
        action_key="case.compensate" if record["compensation_action"] else "case.update",
        action_name="执行对账补偿" if record["compensation_action"] else "更新对账案例",
        target_type="reconciliation_case",
        target_id=item["source_no"],
        target_name=item["case_title"],
        detail_summary=f"{item['case_status_label']} / {item['severity_label']} / {item['last_compensation_action_label'] or '仅保存编排'}",
        detail={"item": item, "compensation_result": compensation_result},
        triggered_by=username,
        source_path=RECONCILIATION_CASE_API_PATH,
        source_method="POST",
        affected_count=1,
    )
    return JSONResponse({"saved": True, "item": item, "compensation_result": compensation_result})


@router.get("/bi-dashboard/api/return-unpack-attendance")
async def list_return_unpack_attendance(
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    limit: int = Query(60, ge=1, le=365),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    start = parse_date_or_none(start_date)
    end = parse_date_or_none(end_date)
    current_engine = get_engine()
    with current_engine.connect() as conn:
        rows = return_unpack_attendance_summaries(conn, start_date=start, end_date=end, limit=limit)
        latest_cleaning = to_plain(latest_date(conn, "sales_cleaning"))
        latest_raw = to_plain(latest_date(conn, "sales"))
    return JSONResponse(
        {
            "rows": rows,
            "latest_cleaning_date": latest_cleaning,
            "latest_raw_date": latest_raw,
        }
    )


@router.post("/bi-dashboard/api/return-unpack-attendance")
async def upsert_return_unpack_attendance(
    payload: Dict[str, Any] = Body(default={}),
    username: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    biz_date = parse_date_or_none(str(payload.get("biz_date") or ""))
    if biz_date is None:
        raise HTTPException(status_code=400, detail="鏃ユ湡涓嶈兘涓虹┖")
    attendance_count = parse_decimal_or_raise(payload.get("attendance_count"), "退货拆包出勤人数")
    current_engine = get_engine()
    try:
        saved_count = save_return_unpack_attendance(current_engine, biz_date, attendance_count)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    with current_engine.connect() as conn:
        rows = return_unpack_attendance_summaries(conn, start_date=biz_date, end_date=biz_date, limit=1)
    summary = rows[0] if rows else {
        "biz_date": biz_date.isoformat(),
        "attendance_count": float(saved_count),
        "sales_return_warehouse": 0.0,
        "total_return_qty": 0.0,
        "total_sales_qty": 0.0,
        "return_unpack_efficiency": 0.0,
    }
    record_dashboard_audit(
        module_key="attendance",
        module_name="出勤记录",
        action_key="return_unpack.upsert",
        action_name="保存退货拆包出勤",
        target_type="return_unpack_attendance",
        target_id=biz_date.isoformat(),
        target_name=f"退货拆包出勤 {biz_date.isoformat()}",
        detail_summary=f"保存退货拆包出勤 {float(saved_count):.2f} 人",
        detail={
            "biz_date": biz_date,
            "attendance_count": float(saved_count),
            "summary": summary,
        },
        triggered_by=username,
        source_path="/financial/bi-dashboard/api/return-unpack-attendance",
        source_method="POST",
        affected_count=1,
    )
    return JSONResponse(
        {
            "saved": True,
            "biz_date": biz_date.isoformat(),
            "attendance_count": float(saved_count),
            "summary": summary,
        }
    )


@router.get("/bi-dashboard/api/meta")
async def bi_meta(_auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        latest_by_dataset = {key: to_plain(latest_date(conn, key)) for key in DATASETS}
        latest_values = [parse_date_or_none(value) for value in latest_by_dataset.values() if value]
    return JSONResponse(
        {
            "widget_types": [{"key": key, "label": label} for key, label in WIDGET_TYPES.items()],
            "widget_type_map": WIDGET_TYPES,
            "datasets": [{"key": key, "label": DATASETS[key]["label"]} for key in DATASETS],
            "dataset_map": {key: DATASETS[key]["label"] for key in DATASETS},
            "dataset_fields": {key: DATASETS[key]["fields"] for key in DATASETS},
            "aggregations": ordered_keys(AGGREGATIONS, AGGREGATION_ORDER),
            "aggregation_options": [{"key": key, "label": AGGREGATION_LABELS[key]} for key in ordered_keys(AGGREGATIONS, AGGREGATION_ORDER)],
            "aggregation_label_map": AGGREGATION_LABELS,
            "filter_operators": ordered_keys(FILTER_OPERATORS, FILTER_OPERATOR_ORDER),
            "filter_operator_options": [{"key": key, "label": FILTER_OPERATOR_LABELS[key]} for key in ordered_keys(FILTER_OPERATORS, FILTER_OPERATOR_ORDER)],
            "filter_operator_label_map": FILTER_OPERATOR_LABELS,
            "layout_heights": sorted(LAYOUT_HEIGHTS),
            "latest_by_dataset": latest_by_dataset,
            "latest_overall": max(latest_values).isoformat() if latest_values else None,
        }
    )


@router.get("/bi-dashboard/api/filter-options")
async def bi_filter_options(
    dataset: str = Query(...),
    field: str = Query(...),
    biz_date: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    keyword: str | None = Query(None),
    limit: int = Query(200, ge=1, le=500),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    if dataset not in DATASETS:
        raise HTTPException(status_code=400, detail=f"涓嶆敮鎸佺殑鏁版嵁闆嗭細{dataset}")
    fields = DATASETS[dataset]["fields"]
    if field not in fields or not fields[field].get("filterable", False):
        raise HTTPException(status_code=400, detail=f"瀛楁涓嶅彲绛涢€夛細{field}")
    selected_date = parse_date_or_none(biz_date)
    range_start = parse_date_or_none(start_date)
    range_end = parse_date_or_none(end_date)
    current_engine = get_engine()
    with current_engine.connect() as conn:
        options, scope = query_filter_options(
            conn,
            dataset=dataset,
            field=field,
            selected_date=selected_date,
            start_date=range_start,
            end_date=range_end,
            keyword=keyword,
            limit=limit,
        )
    return JSONResponse(
        {
            "dataset": dataset,
            "field": field,
            "biz_date": to_plain(selected_date),
            "start_date": to_plain(range_start),
            "end_date": to_plain(range_end),
            "keyword": str(keyword or "").strip() or None,
            "scope": scope,
            "options": options,
        }
    )


@router.get("/bi-dashboard/api/views")
async def list_views(_auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT
                    v.id, v.name, v.description, v.global_filters_json, v.created_at, v.updated_at,
                    COUNT(w.id) AS widget_count
                FROM bi_dashboard_view v
                LEFT JOIN bi_dashboard_widget w ON w.view_id = v.id
                GROUP BY v.id, v.name, v.description, v.global_filters_json, v.created_at, v.updated_at
                ORDER BY v.id
                """
            )
        ).mappings().all()
    return JSONResponse(
        {
            "views": [
                {
                    "id": int(row["id"]),
                    "name": row["name"],
                    "description": row["description"],
                    "global_filters": normalize_global_filters(json_loads(row.get("global_filters_json"), [])),
                    "widget_count": int(row["widget_count"] or 0),
                    "created_at": to_plain(row["created_at"]),
                    "updated_at": to_plain(row["updated_at"]),
                }
                for row in rows
            ]
        }
    )


@router.post("/bi-dashboard/api/views")
async def create_view(payload: Dict[str, Any] = Body(default={}), _auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    name = str(payload.get("name") or "").strip()
    description = str(payload.get("description") or "").strip()
    global_filters = normalize_global_filters(payload.get("global_filters"))
    if not name:
        raise HTTPException(status_code=400, detail="鐪嬫澘鍚嶇О涓嶈兘涓虹┖")
    current_engine = get_engine()
    with current_engine.begin() as conn:
        result = conn.execute(
            text(
                """
                INSERT INTO bi_dashboard_view(name, description, global_filters_json)
                VALUES (:name, :description, :global_filters_json)
                """
            ),
            {
                "name": name,
                "description": description,
                "global_filters_json": json_dumps(global_filters),
            },
        )
    return JSONResponse({"id": int(result.lastrowid)})


@router.get("/bi-dashboard/api/views/{view_id}")
async def get_view(view_id: int, _auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        return JSONResponse(view_detail(conn, view_id))


@router.put("/bi-dashboard/api/views/{view_id}")
async def update_view(view_id: int, payload: Dict[str, Any] = Body(default={}), _auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    name = str(payload.get("name") or "").strip()
    description = str(payload.get("description") or "").strip()
    global_filters = normalize_global_filters(payload.get("global_filters"))
    if not name:
        raise HTTPException(status_code=400, detail="鐪嬫澘鍚嶇О涓嶈兘涓虹┖")
    current_engine = get_engine()
    with current_engine.begin() as conn:
        load_view(conn, view_id)
        conn.execute(
            text(
                """
                UPDATE bi_dashboard_view
                SET name = :name, description = :description, global_filters_json = :global_filters_json
                WHERE id = :id
                """
            ),
            {
                "id": view_id,
                "name": name,
                "description": description,
                "global_filters_json": json_dumps(global_filters),
            },
        )
    return JSONResponse({"id": view_id, "name": name, "description": description, "global_filters": global_filters})


@router.get("/bi-dashboard/api/views/{view_id}/export")
async def export_view(
    view_id: int,
    biz_date: str | None = Query(None),
    _auth: str = Depends(require_auth),
) -> Response:
    ensure_schema()
    selected_date = parse_date_or_none(biz_date)
    current_engine = get_engine()
    with current_engine.connect() as conn:
        detail = view_detail(conn, view_id)
        widgets = [
            {
                "id": widget["id"],
                "title": widget["title"],
                "widget_type": widget["widget_type"],
                "dataset": widget["dataset"],
                "layout": widget["layout"],
                "data": query_widget_data(conn, widget, selected_date, detail["global_filters"]),
            }
            for widget in detail["widgets"]
        ]
    content = json.dumps(
        {
            "view": {
                "id": detail["id"],
                "name": detail["name"],
                "description": detail["description"],
                "global_filters": detail["global_filters"],
            },
            "selected_date": to_plain(selected_date),
            "exported_at": datetime.now().isoformat(timespec="seconds"),
            "widgets": widgets,
        },
        ensure_ascii=False,
        indent=2,
    )
    filename = f"dashboard-view-{view_id}-{selected_date or date.today()}.json"
    return Response(
        content=content,
        media_type="application/json; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/bi-dashboard/api/layout-templates")
async def list_layout_templates(_auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT id, name, description, layout_json, created_at, updated_at
                FROM bi_dashboard_layout_template
                ORDER BY updated_at DESC, id DESC
                """
            )
        ).mappings().all()
    return JSONResponse({"templates": [layout_template_from_row(dict(row)) for row in rows]})


@router.post("/bi-dashboard/api/layout-templates")
async def create_layout_template(
    payload: Dict[str, Any] = Body(default={}),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    name = str(payload.get("name") or "").strip()
    description = str(payload.get("description") or "").strip()
    view_id = int(payload.get("view_id") or 0)
    if not name:
        raise HTTPException(status_code=400, detail="妯℃澘鍚嶇О涓嶈兘涓虹┖")
    if view_id <= 0:
        raise HTTPException(status_code=400, detail="缂哄皯鏈夋晥鐨勭湅鏉?ID")
    current_engine = get_engine()
    with current_engine.begin() as conn:
        detail = view_detail(conn, view_id)
        layout_payload = build_layout_template_payload(detail["widgets"])
        result = conn.execute(
            text(
                """
                INSERT INTO bi_dashboard_layout_template(name, description, layout_json)
                VALUES (:name, :description, :layout_json)
                """
            ),
            {
                "name": name,
                "description": description,
                "layout_json": json_dumps(layout_payload),
            },
        )
    return JSONResponse({"id": int(result.lastrowid)})


@router.post("/bi-dashboard/api/layout-templates/{template_id}/apply")
async def apply_layout_template(
    template_id: int,
    payload: Dict[str, Any] = Body(default={}),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    view_id = int(payload.get("view_id") or 0)
    if view_id <= 0:
        raise HTTPException(status_code=400, detail="缂哄皯鏈夋晥鐨勭湅鏉?ID")
    current_engine = get_engine()
    with current_engine.begin() as conn:
        template = load_layout_template(conn, template_id)
        load_view(conn, view_id)
        template_widgets = template["layout_payload"]["widgets"]
        if not template_widgets:
            raise HTTPException(status_code=400, detail="璇ュ竷灞€妯℃澘娌℃湁鍙鐢ㄧ殑鍗＄墖甯冨眬")
        widget_rows = conn.execute(
            text(
                """
                SELECT id, layout_json, sort_order
                FROM bi_dashboard_widget
                WHERE view_id = :view_id
                ORDER BY sort_order, id
                """
            ),
            {"view_id": view_id},
        ).mappings().all()
        placed_layouts: List[Dict[str, Any]] = []
        template_bottom = max((item["layout"]["y"] + item["layout"]["h"] for item in template_widgets), default=0)
        overflow_y = template_bottom
        for index, row in enumerate(widget_rows):
            if index < len(template_widgets):
                layout = normalize_layout(template_widgets[index]["layout"])
            else:
                current_layout = normalize_layout(json_loads(row.get("layout_json"), default_layout()))
                layout = normalize_layout({**current_layout, "x": 0, "y": overflow_y})
            while True:
                overlap = next((item for item in placed_layouts if layouts_collide(layout, item)), None)
                if not overlap:
                    break
                layout = normalize_layout({**layout, "y": overlap["y"] + overlap["h"]})
            placed_layouts.append(layout)
            overflow_y = max(overflow_y, layout["y"] + layout["h"])
            conn.execute(
                text(
                    """
                    UPDATE bi_dashboard_widget
                    SET sort_order = :sort_order, layout_json = :layout_json
                    WHERE id = :id
                    """
                ),
                {
                    "id": int(row["id"]),
                    "sort_order": (layout["y"] * 100) + layout["x"] + index,
                    "layout_json": json_dumps(layout),
                },
            )
    return JSONResponse({"applied": True, "template_id": template_id, "view_id": view_id})


@router.delete("/bi-dashboard/api/layout-templates/{template_id}")
async def delete_layout_template(template_id: int, _auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.begin() as conn:
        load_layout_template(conn, template_id)
        conn.execute(text("DELETE FROM bi_dashboard_layout_template WHERE id = :id"), {"id": template_id})
    return JSONResponse({"deleted": True})


@router.post("/bi-dashboard/api/views/{view_id}/widgets")
async def create_widget(
    view_id: int,
    payload: Dict[str, Any] = Body(default={}),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    title = str(payload.get("title") or "鏂板缓缁勪欢").strip()
    widget_type = str(payload.get("widget_type") or "bar")
    dataset = str(payload.get("dataset") or "sales_cleaning")
    if widget_type not in WIDGET_TYPES:
        raise HTTPException(status_code=400, detail=f"不支持的图表类型：{widget_type}")
    if dataset not in DATASETS:
        raise HTTPException(status_code=400, detail=f"涓嶆敮鎸佺殑鏁版嵁闆嗭細{dataset}")
    config = normalize_widget_config(widget_type, default_widget_config(widget_type, dataset))
    current_engine = get_engine()
    with current_engine.begin() as conn:
        load_view(conn, view_id)
        max_sort = int(
            conn.execute(
                text("SELECT COALESCE(MAX(sort_order), 0) FROM bi_dashboard_widget WHERE view_id = :view_id"),
                {"view_id": view_id},
            ).scalar()
            or 0
        )
        result = conn.execute(
            text(
                """
                INSERT INTO bi_dashboard_widget(
                    view_id, title, widget_type, dataset, config_json, layout_json, sort_order, analysis_text
                ) VALUES (
                    :view_id, :title, :widget_type, :dataset, :config_json, :layout_json, :sort_order, :analysis_text
                )
                """
            ),
            {
                "view_id": view_id,
                "title": title,
                "widget_type": widget_type,
                "dataset": dataset,
                "config_json": json_dumps(config),
                "layout_json": json_dumps(default_layout()),
                "sort_order": max_sort + 10,
                "analysis_text": "",
            },
        )
    return JSONResponse({"id": int(result.lastrowid)})


@router.post("/bi-dashboard/api/widgets/{widget_id}/duplicate")
async def duplicate_widget(widget_id: int, _auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.begin() as conn:
        widget = load_widget(conn, widget_id)
        duplicated_layout = normalize_layout(
            {
                **widget["layout"],
                "x": min(GRID_COLUMNS - widget["layout"].get("w", GRID_COLUMNS // 2), widget["layout"].get("x", 0) + 1),
                "y": widget["layout"].get("y", 0) + 1,
            }
        )
        max_sort = int(
            conn.execute(
                text("SELECT COALESCE(MAX(sort_order), 0) FROM bi_dashboard_widget WHERE view_id = :view_id"),
                {"view_id": widget["view_id"]},
            ).scalar()
            or 0
        )
        result = conn.execute(
            text(
                """
                INSERT INTO bi_dashboard_widget(
                    view_id, title, widget_type, dataset, config_json, layout_json, sort_order, analysis_text
                ) VALUES (
                    :view_id, :title, :widget_type, :dataset, :config_json, :layout_json, :sort_order, :analysis_text
                )
                """
            ),
            {
                "view_id": widget["view_id"],
                "title": f"{widget['title']} 鍓湰",
                "widget_type": widget["widget_type"],
                "dataset": widget["dataset"],
                "config_json": json_dumps(widget["config"]),
                "layout_json": json_dumps(duplicated_layout),
                "sort_order": max_sort + 10,
                "analysis_text": widget["analysis_text"],
            },
        )
    return JSONResponse({"id": int(result.lastrowid)})


@router.put("/bi-dashboard/api/views/{view_id}/layout")
async def save_view_layout(
    view_id: int,
    payload: Dict[str, Any] = Body(default={}),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    widgets = payload.get("widgets") or []
    if not isinstance(widgets, list):
        raise HTTPException(status_code=400, detail="布局数据格式错误，widgets 必须是数组")
    current_engine = get_engine()
    with current_engine.begin() as conn:
        load_view(conn, view_id)
        valid_ids = {
            int(row[0])
            for row in conn.execute(
                text("SELECT id FROM bi_dashboard_widget WHERE view_id = :view_id"),
                {"view_id": view_id},
            ).fetchall()
        }
        for item in widgets:
            if not isinstance(item, dict):
                continue
            widget_id = int(item.get("id", 0) or 0)
            if widget_id not in valid_ids:
                continue
            conn.execute(
                text(
                    """
                    UPDATE bi_dashboard_widget
                    SET sort_order = :sort_order, layout_json = :layout_json
                    WHERE id = :id
                    """
                ),
                {
                    "id": widget_id,
                    "sort_order": int(item.get("sort_order", 0) or 0),
                    "layout_json": json_dumps(normalize_layout(item.get("layout"))),
                },
            )
    return JSONResponse({"saved": True})


@router.put("/bi-dashboard/api/widgets/{widget_id}")
async def update_widget(
    widget_id: int,
    payload: Dict[str, Any] = Body(default={}),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.begin() as conn:
        widget = load_widget(conn, widget_id)
        widget_type = str(payload.get("widget_type") or widget["widget_type"])
        dataset = str(payload.get("dataset") or widget["dataset"])
        if widget_type not in WIDGET_TYPES:
            raise HTTPException(status_code=400, detail=f"不支持的图表类型：{widget_type}")
        if dataset not in DATASETS:
            raise HTTPException(status_code=400, detail=f"涓嶆敮鎸佺殑鏁版嵁闆嗭細{dataset}")
        config_raw = payload.get("config", widget["config"])
        if isinstance(config_raw, dict):
            config_raw = dict(config_raw)
            config_raw["dataset"] = dataset
        conn.execute(
            text(
                """
                UPDATE bi_dashboard_widget
                SET
                    title = :title,
                    widget_type = :widget_type,
                    dataset = :dataset,
                    config_json = :config_json,
                    layout_json = :layout_json,
                    sort_order = :sort_order,
                    analysis_text = :analysis_text
                WHERE id = :id
                """
            ),
            {
                "id": widget_id,
                "title": str(payload.get("title") or widget["title"]).strip(),
                "widget_type": widget_type,
                "dataset": dataset,
                "config_json": json_dumps(normalize_widget_config(widget_type, config_raw if isinstance(config_raw, dict) else {})),
                "layout_json": json_dumps(normalize_layout(payload.get("layout", widget["layout"]))),
                "sort_order": int(payload.get("sort_order", widget["sort_order"])),
                "analysis_text": str(payload.get("analysis_text", widget["analysis_text"])),
            },
        )
    return JSONResponse({"id": widget_id})


@router.delete("/bi-dashboard/api/widgets/{widget_id}")
async def delete_widget(widget_id: int, _auth: str = Depends(require_auth)) -> JSONResponse:
    ensure_schema()
    current_engine = get_engine()
    with current_engine.begin() as conn:
        load_widget(conn, widget_id)
        conn.execute(text("DELETE FROM bi_dashboard_widget WHERE id = :id"), {"id": widget_id})
    return JSONResponse({"deleted": True})


@router.get("/bi-dashboard/api/widgets/{widget_id}/data")
async def widget_data(
    widget_id: int,
    biz_date: str | None = Query(None),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    selected_date = parse_date_or_none(biz_date)
    current_engine = get_engine()
    with current_engine.connect() as conn:
        widget = load_widget(conn, widget_id)
        view = load_view(conn, widget["view_id"])
        payload = query_widget_data(conn, widget, selected_date, view["global_filters"])
    payload.update({"widget_id": widget["id"], "widget_type": widget["widget_type"], "title": widget["title"]})
    return JSONResponse(payload)


@router.get("/bi-dashboard/api/views/{view_id}/widget-data")
async def view_widget_data(
    view_id: int,
    biz_date: str | None = Query(None),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    selected_date = parse_date_or_none(biz_date)
    current_engine = get_engine()
    with current_engine.connect() as conn:
        detail = view_detail(conn, view_id)
        items: List[Dict[str, Any]] = []
        for widget in detail["widgets"]:
            payload = query_widget_data(conn, widget, selected_date, detail["global_filters"])
            payload.update({"widget_id": widget["id"], "widget_type": widget["widget_type"], "title": widget["title"]})
            items.append(payload)
    return JSONResponse(
        {
            "view_id": view_id,
            "biz_date": to_plain(selected_date),
            "items": items,
        }
    )


@router.post("/bi-dashboard/api/widgets/{widget_id}/share-dingtalk")
async def share_widget_dingtalk(
    widget_id: int,
    payload: Dict[str, Any] = Body(default={}),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    selected_date = parse_date_or_none(payload.get("biz_date"))
    group_name = str(payload.get("group_name") or "供应链数据同步群").strip() or "供应链数据同步群"
    message_tag = str(payload.get("message_tag") or "北极星系统测试消息").strip() or "北极星系统测试消息"
    current_engine = get_engine()
    with current_engine.connect() as conn:
        widget = load_widget(conn, widget_id)
        view = load_view(conn, widget["view_id"])
        widget_payload = query_widget_data(conn, widget, selected_date, view["global_filters"])
        dataset_label = DATASETS[widget["dataset"]]["label"]

    share_result = await share_dashboard_widget(
        widget_title=widget["title"],
        dataset_label=dataset_label,
        widget_type=widget["widget_type"],
        target_date=str(widget_payload.get("target_date") or "--"),
        dimensions=list(widget_payload.get("dimensions") or []),
        metrics=list(widget_payload.get("metrics") or []),
        rows=list(widget_payload.get("rows") or []),
        series_field=str(widget_payload.get("series_field") or ""),
        series_groups=list(widget_payload.get("series_groups") or []),
        group_name=group_name,
        message_tag=message_tag,
    )
    share_result["widget_id"] = widget_id
    return JSONResponse(share_result)


@router.post("/bi-dashboard/api/widgets/{widget_id}/ai-analysis")
async def widget_ai_analysis(
    widget_id: int,
    biz_date: str | None = Query(None),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    selected_date = parse_date_or_none(biz_date)
    current_engine = get_engine()
    with current_engine.begin() as conn:
        widget = load_widget(conn, widget_id)
        view = load_view(conn, widget["view_id"])
        payload = query_widget_data(conn, widget, selected_date, view["global_filters"])
        analysis_text = generate_widget_ai_text(widget, payload)
        conn.execute(
            text("UPDATE bi_dashboard_widget SET analysis_text = :analysis_text WHERE id = :id"),
            {"id": widget_id, "analysis_text": analysis_text},
        )
    return JSONResponse({"widget_id": widget_id, "analysis_text": analysis_text})


@router.put("/bi-dashboard/api/widgets/{widget_id}/analysis")
async def update_widget_analysis(
    widget_id: int,
    payload: Dict[str, Any] = Body(default={}),
    _auth: str = Depends(require_auth),
) -> JSONResponse:
    ensure_schema()
    analysis_text = str(payload.get("analysis_text") or "")
    current_engine = get_engine()
    with current_engine.begin() as conn:
        load_widget(conn, widget_id)
        conn.execute(
            text("UPDATE bi_dashboard_widget SET analysis_text = :analysis_text WHERE id = :id"),
            {"id": widget_id, "analysis_text": analysis_text},
        )
    return JSONResponse({"widget_id": widget_id, "analysis_text": analysis_text})
